from fastapi import FastAPI, Form, UploadFile, File, Request, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from typing import Optional, List
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Adicionado estilos
from openpyxl.utils import get_column_letter                           # Adicionado utilitário
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import sqlite3
import os
import re
import uuid
import asyncio
import time
import math
from collections import deque
import qrcode
import base64
import html
import secrets
import json
import urllib.parse
import unicodedata

app = FastAPI()

DATABASE = "database.db"
UPLOAD_DIR = "static/imagens"

os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

RACAS = ["Branca", "Preta", "Parda", "Amarela", "Indígena"]
ANOS = ["6º ano", "7º ano", "8º ano", "9º ano"]

# Tipos de questão suportados. Cada um vira um fluxo de cadastro/resposta diferente.
TIPOS_QUESTAO = {
    "multipla_escolha": {"label": "Múltipla escolha (A/B/C/D)", "icone": "🔘"},
    "discursiva":       {"label": "Discursiva (resposta livre)", "icone": "📝"},
    "vf":               {"label": "Verdadeiro ou Falso (afirmações)", "icone": "✓✗"},
    "associacao":       {"label": "Associação de colunas", "icone": "↔"},
}

# Limites pra cartão impresso (mantém legibilidade)
VF_MAX_AFIRMACOES = 5      # até 5 afirmações por questão V/F
ASSOC_MAX_PARES = 5         # até 5×5 (5 itens × 5 letras) na associação

# === Autenticação ===
# Variáveis de ambiente esperadas em produção. Em dev, defaults permitem testar.
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
SESSION_SECRET_KEY = os.environ.get("SESSION_SECRET_KEY", "dev-key-CHANGE-IN-PRODUCTION-" + secrets.token_hex(8))
BASE_URL = os.environ.get("BASE_URL", "http://localhost:8000")
ALLOWED_EMAIL_DOMAIN = os.environ.get("ALLOWED_EMAIL_DOMAIN", "smevr.com.br")
# Modo dev: se não tem credenciais OAuth, libera login fake só com email
DEV_MODE = (os.environ.get("DEV_MODE", "1") == "1") and not GOOGLE_CLIENT_ID
SESSION_COOKIE = "corretor_session"
SESSION_MAX_AGE = 60 * 60 * 24 * 30  # 30 dias

_session_serializer = URLSafeTimedSerializer(SESSION_SECRET_KEY, salt="session-v1")


GOOGLE_DRIVE_FOLDER_ID = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")
GOOGLE_DRIVE_CREDENTIALS_JSON = os.environ.get("GOOGLE_DRIVE_CREDENTIALS_JSON", "")

TIPOS_AFASTAMENTO = {
    "atestado_medico": "Atestado médico",
    "permissao_ausencia": "Permissão de ausência",
    "abono_1_3": "Abono 1/3",
    "abono_2_3": "Abono 2/3",
    "folga_tre": "Folga TRE",
    "abono_integral": "Abono Integral",
    "declaracao_comparecimento": "Declaração de comparecimento",
}
# Tipos que precisam de horário específico (não é o dia inteiro) — 26/08/2026
TIPOS_COM_HORARIO = {"permissao_ausencia", "declaracao_comparecimento"}


def _extrair_matricula(email: str) -> str:
    """Extrai a matrícula do e-mail de cadastro do profissional. Formato usado na rede:
    nome.MATRICULA@smevr.com.br — pega o último trecho separado por ponto, se for só
    dígitos. Se não achar (email fora do padrão), retorna '—' (25/08/2026)."""
    if not email or "@" not in email:
        return "—"
    local = email.split("@")[0]
    partes = local.split(".")
    ultima = partes[-1] if partes else ""
    return ultima if ultima.isdigit() else "—"


def _drive_upload_arquivo(nome_arquivo: str, conteudo_bytes: bytes, mime_type: str):
    """Sobe um arquivo pra pasta do Google Drive configurada (GOOGLE_DRIVE_FOLDER_ID).

    Usa a IDENTIDADE DA PRÓPRIA VM (Application Default Credentials) em vez de uma chave
    JSON baixada — não precisa de service account key (a organização pode bloquear a
    criação dessas chaves por política de segurança, como aconteceu aqui em 25/08/2026).
    A VM precisa ter o escopo do Drive habilitado (drive, não só cloud-platform — o
    Drive fica de fora do cloud-platform) e a pasta compartilhada com o e-mail da conta
    de serviço da própria VM. Se GOOGLE_DRIVE_CREDENTIALS_JSON estiver definida, ainda é
    aceita como alternativa (ambientes fora do GCP, ou se a política mudar).

    IMPORTANTE sobre o escopo: usamos 'drive' (completo) e não 'drive.file'. O escopo
    'drive.file' só enxerga arquivos/pastas que O PRÓPRIO APP criou (ou que foram abertos
    via seletor do Google) — uma pasta compartilhada manualmente pela interface do Drive
    fica invisível pra esse escopo, mesmo com a permissão de Editor certinha. Descobrimos
    isso em produção em 25/08/2026 (erro 404 'File not found' na pasta, mesmo compartilhada
    corretamente) — trocar pra 'drive' resolve.

    Retorna (file_id, link, erro). Se não der pra autenticar, retorna erro claro em vez de
    quebrar — o registro do afastamento é salvo de qualquer forma."""
    if not GOOGLE_DRIVE_FOLDER_ID:
        return None, None, "Google Drive não configurado ainda (falta GOOGLE_DRIVE_FOLDER_ID no servidor)."
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        import io as _io

        DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]
        creds = None
        if GOOGLE_DRIVE_CREDENTIALS_JSON:
            from google.oauth2 import service_account
            info = json.loads(GOOGLE_DRIVE_CREDENTIALS_JSON)
            creds = service_account.Credentials.from_service_account_info(info, scopes=DRIVE_SCOPES)
        else:
            import google.auth
            creds, _ = google.auth.default(scopes=DRIVE_SCOPES)

        service = build("drive", "v3", credentials=creds)
        metadata = {"name": nome_arquivo, "parents": [GOOGLE_DRIVE_FOLDER_ID]}
        media = MediaIoBaseUpload(_io.BytesIO(conteudo_bytes), mimetype=mime_type, resumable=False)
        # supportsAllDrives=True é obrigatório se a pasta estiver dentro de um Drive
        # Compartilhado (não "Meu Drive") — sem isso a API nem enxerga a pasta e retorna
        # 404 "File not found", mesmo com permissão certa (achado em produção, 25/08/2026).
        arquivo = service.files().create(
            body=metadata, media_body=media, fields="id, webViewLink", supportsAllDrives=True
        ).execute()
        return arquivo.get("id"), arquivo.get("webViewLink"), None
    except ImportError:
        return None, None, "Biblioteca do Google Drive não instalada no servidor (google-api-python-client / google-auth)."
    except Exception as e:
        import traceback
        # Algumas exceções do google-auth/googleapiclient têm str(e) vazio — cai pro
        # repr(e) (mostra ao menos o tipo da exceção) e, se for erro HTTP da API,
        # extrai o motivo (25/08/2026, depois de ver esse caso em produção).
        detalhe = str(e).strip() or repr(e)
        try:
            from googleapiclient.errors import HttpError
            if isinstance(e, HttpError):
                motivo = e._get_reason().strip() if hasattr(e, "_get_reason") else ""
                detalhe = f"HTTP {e.resp.status} — {motivo or e.content}"
        except Exception:
            pass
        # Log completo (com traceback) vai pro journalctl, mesmo que a mensagem mostrada
        # ao usuário seja curta — é o que a gente olha pra diagnosticar de verdade.
        print(f"[Drive upload] Falha ao enviar '{nome_arquivo}': {detalhe}")
        print(traceback.format_exc())
        return None, None, f"Erro ao enviar pro Google Drive: {detalhe}"


def _pode_editar_questao(prof: Optional[dict], questao_criador_id: Optional[int]) -> bool:
    """Autor da questão OU admin podem editar. Questões legadas (sem dono) só admin edita."""
    if not prof:
        return False
    if prof.get("is_admin"):
        return True
    if questao_criador_id is None:
        return False
    return prof["id"] == questao_criador_id


def _redimensionar_imagem(data: bytes, max_width: int = 800) -> bytes:
    """Redimensiona imagem para no máximo max_width px de largura, convertendo para JPEG."""
    try:
        from PIL import Image as _PilImage
        import io as _io
        img = _PilImage.open(_io.BytesIO(data))
        if img.mode in ("RGBA", "P"): img = img.convert("RGB")
        w, h = img.size
        if w > max_width:
            img = img.resize((max_width, int(h * max_width / w)), _PilImage.LANCZOS)
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=85, optimize=True)
        return buf.getvalue()
    except Exception:
        return data


def _sanitizar_html_enunciado(html: str) -> str:
    """Permite apenas tags básicas de formatação no enunciado. Remove scripts, iframes, handlers JS.
    Tags permitidas: strong/b, em/i, u, br, p, div (só com style text-align), span (só com style text-align), ul, ol, li, blockquote.
    Atributos permitidos: apenas style com text-align."""
    import re as _re
    if not html:
        return ""
    # Remove tags perigosas completas (com conteúdo)
    html = _re.sub(r'<(script|style|iframe|object|embed|form|input|button|textarea|select|link|meta)\b[^>]*>.*?</\1>',
                   '', html, flags=_re.IGNORECASE | _re.DOTALL)
    html = _re.sub(r'<(script|style|iframe|object|embed|form|input|button|textarea|select|link|meta)\b[^>]*/?>',
                   '', html, flags=_re.IGNORECASE)
    # Remove atributos on* (onclick, onerror, etc.) e javascript: em href/src
    html = _re.sub(r'\son[a-z]+\s*=\s*"[^"]*"', '', html, flags=_re.IGNORECASE)
    html = _re.sub(r"\son[a-z]+\s*=\s*'[^']*'", '', html, flags=_re.IGNORECASE)
    html = _re.sub(r'\son[a-z]+\s*=\s*[^\s>]+', '', html, flags=_re.IGNORECASE)
    html = _re.sub(r'(href|src)\s*=\s*["\']?\s*javascript:[^"\'>\s]*["\']?', '', html, flags=_re.IGNORECASE)
    # Whitelist de tags - remove qualquer tag que não esteja na lista
    permitidas = {"strong", "b", "em", "i", "u", "br", "p", "div", "span", "ul", "ol", "li", "blockquote",
                   "table", "thead", "tbody", "tr", "th", "td", "img", "figure", "figcaption"}
    def _filtrar_tag(m):
        tag_full = m.group(0)
        tag_name = m.group(1).lower()
        if tag_name not in permitidas:
            return ""
        # img: mantém src e alt, remove outros atributos perigosos
        if tag_name == "img":
            src = _re.search(r'src\s*=\s*["\']([^"\'>]+)["\']', tag_full)
            if not src: return ""
            alt = _re.search(r'alt\s*=\s*["\']([^"\'>]*)["\' ]', tag_full)
            alt_val = alt.group(1) if alt else ""
            return f'<img src="{src.group(1)}" alt="{alt_val}" style="max-width:100%; height:auto;">'
        # table/th/td: mantém style de bordas
        if tag_name in ("table", "th", "td"):
            style = _re.search(r'style\s*=\s*["\']([^"\'>]+)["\' ]', tag_full)
            style_attr = f' style="{style.group(1)}"' if style else ""
            if tag_full.startswith("</"): return f"</{tag_name}>"
            return f'<{tag_name}{style_attr}>'
        # thead/tbody/tr/figure/figcaption: sem atributos
        if tag_name in ("thead", "tbody", "tr", "figure", "figcaption"):
            if tag_full.startswith("</"): return f"</{tag_name}>"
            return f"<{tag_name}>"
        # Para div/span/p, mantém só style com text-align
        if tag_name in ("div", "span", "p"):
            ta_match = _re.search(r'style\s*=\s*["\']([^"\']*text-align\s*:\s*(left|center|right|justify)[^"\']*)["\']', tag_full, _re.IGNORECASE)
            if ta_match:
                align_val = ta_match.group(2).lower()
                return f'<{tag_name} style="text-align:{align_val};">' if not tag_full.startswith("</") else f"</{tag_name}>"
        # Outras tags: sem atributos
        if tag_full.startswith("</"):
            return f"</{tag_name}>"
        return f"<{tag_name}>"
    html = _re.sub(r'</?([a-zA-Z][a-zA-Z0-9]*)\b[^>]*>', _filtrar_tag, html)
    return html.strip()




_JS_MATH_BUTTONS = r"""
            function _inserirTexto(editor, sync, txt) {
                editor.focus();
                var sel = window.getSelection();
                if (sel && sel.rangeCount) {
                    var rng = sel.getRangeAt(0);
                    if (!editor.contains(rng.commonAncestorContainer)) {
                        rng = document.createRange();
                        rng.selectNodeContents(editor);
                        rng.collapse(false);
                    }
                    rng.deleteContents();
                    var node = document.createTextNode(txt);
                    rng.insertNode(node);
                    rng.setStartAfter(node);
                    rng.collapse(true);
                    sel.removeAllRanges();
                    sel.addRange(rng);
                } else {
                    document.execCommand("insertText", false, txt);
                }
                sync();
            }
            var btnFrac = toolbar.querySelector(".btn-insert-frac");
            if (btnFrac) {
                btnFrac.addEventListener("click", function(e) {
                    e.preventDefault();
                    var num = prompt("Numerador da fração:");
                    if (num === null) return;
                    var den = prompt("Denominador da fração:");
                    if (den === null) return;
                    _inserirTexto(editor, sync, "$\\frac{" + num + "}{" + den + "}$");
                });
            }
            var btnPot = toolbar.querySelector(".btn-insert-pot");
            if (btnPot) {
                btnPot.addEventListener("click", function(e) {
                    e.preventDefault();
                    var base = prompt("Base (ex: 2, x, 2x):");
                    if (base === null) return;
                    var expoente = prompt("Expoente (ex: 2, 3, n):");
                    if (expoente === null) return;
                    _inserirTexto(editor, sync, "$" + base + "^{" + expoente + "}$");
                });
            }
            var btnTab = toolbar.querySelector(".btn-insert-tab");
            if (btnTab) {
                btnTab.addEventListener("click", function(e) {
                    e.preventDefault();
                    var nlin = parseInt(prompt("Número de linhas:", "3"));
                    if (!nlin || nlin < 1) return;
                    var ncol = parseInt(prompt("Número de colunas:", "3"));
                    if (!ncol || ncol < 1) return;
                    var tbl = '<table style="border-collapse:collapse;width:100%;margin:8px 0;">';
                    tbl += "<thead><tr>";
                    for (var c = 0; c < ncol; c++) {
                        tbl += '<th style="border:1px solid #999;padding:6px 10px;background:#f0f0f0;font-weight:600;">Col ' + (c+1) + "</th>";
                    }
                    tbl += "</tr></thead><tbody>";
                    for (var r = 0; r < nlin - 1; r++) {
                        tbl += "<tr>";
                        for (var c2 = 0; c2 < ncol; c2++) {
                            tbl += '<td style="border:1px solid #999;padding:6px 10px;">&nbsp;</td>';
                        }
                        tbl += "</tr>";
                    }
                    tbl += "</tbody></table><p></p>";
                    document.execCommand("insertHTML", false, tbl);
                    sync();
                });
            }
"""

_JS_DETECTAR_ALTS = r"""
            function detectarAlternativas(texto) {
                texto = texto.replace(/\r\n/g, '\n').replace(/\u00A0/g, ' ').trim();
                var padrao = /(?:^|\n)[ \t]*[(]?([A-Da-d])[)]?[ \t]*[-).,:][ \t]*/g;
                var matches = Array.from(texto.matchAll(padrao));
                var idxA=-1, idxB=-1, idxC=-1, idxD=-1;
                for (var mi=0; mi<matches.length; mi++) {
                    var letra = matches[mi][1].toUpperCase();
                    var pos = matches[mi].index;
                    if (letra==='A' && idxA===-1) idxA=pos;
                    else if (letra==='B' && idxB===-1 && idxA!==-1 && pos>idxA) idxB=pos;
                    else if (letra==='C' && idxC===-1 && idxB!==-1 && pos>idxB) idxC=pos;
                    else if (letra==='D' && idxD===-1 && idxC!==-1 && pos>idxC) idxD=pos;
                }
                if (idxA===-1 || idxB===-1 || idxC===-1 || idxD===-1) return null;
                var enunciado = texto.slice(0, idxA).trim();
                function ext(s,e) { return texto.slice(s,e).replace(/^\n?[ \t]*[(]?[A-Da-d][)]?[ \t]*[-).,:][ \t]*/, "").trim(); }
                return { enunciado:enunciado, alternativas:[ext(idxA,idxB),ext(idxB,idxC),ext(idxC,idxD),ext(idxD,texto.length)] };
            }
            function aplicarAlternativas(texto) {
                var r = detectarAlternativas(texto);
                if (!r) { document.execCommand("insertText", false, texto); return; }
                var trunc = function(s) { return s.length > 60 ? s.slice(0,60)+"..." : s; };
                var nl = "\n";
                var msg = "Detectei 4 alternativas. Aplicar automaticamente?" + nl + nl
                        + (r.enunciado ? "Enunciado: " + trunc(r.enunciado) + nl : "")
                        + "A) " + trunc(r.alternativas[0]) + nl
                        + "B) " + trunc(r.alternativas[1]) + nl
                        + "C) " + trunc(r.alternativas[2]) + nl
                        + "D) " + trunc(r.alternativas[3]);
                if (!confirm(msg)) { document.execCommand("insertText", false, texto); return; }
                editor.innerHTML = r.enunciado ? r.enunciado.replace(/\n/g, "<br>") : "";
                hidden.value = editor.innerHTML;
                refreshPlaceholder();
                ["a","b","c","d"].forEach(function(letra, idx) {
                    var altEd = document.querySelector(".editor-content[data-target=\"alt_"+letra+"\"]");
                    var altHid = document.getElementById("alt_"+letra+"_hidden");
                    if (altEd && altHid) {
                        altEd.innerHTML = r.alternativas[idx].replace(/\n/g, "<br>");
                        altHid.value = altEd.innerHTML;
                        altEd.removeAttribute("data-ph-shown");
                    }
                });
            }
            editor.addEventListener("paste", function(e) {
                var cb = e.clipboardData || window.clipboardData;
                if (!cb) return;
                var items = cb.items ? Array.from(cb.items) : [];
                var imgItem = items.find(function(it) { return it.type.startsWith("image/"); });
                if (imgItem) {
                    e.preventDefault();
                    var blob = imgItem.getAsFile();
                    if (!blob) return;
                    var fd = new FormData();
                    fd.append("arquivo", blob, "imagem_colada.png");
                    fetch("/upload-imagem-inline", { method: "POST", body: fd })
                        .then(function(r) { return r.json(); })
                        .then(function(data) {
                            if (data.url) {
                                document.execCommand("insertHTML", false,
                                    "<img src=\"" + data.url + "\" style=\"max-width:100%; height:auto; display:block; margin:4px 0;\" alt=\"\">");
                                sync();
                            }
                        })
                        .catch(function() { alert("Erro ao fazer upload da imagem."); });
                    return;
                }
                var texto = cb.getData("text/plain") || "";
                if (!texto) return;
                e.preventDefault();
                aplicarAlternativas(texto);
            });
"""

def _editor_enunciado_html(name: str = "enunciado", valor_inicial: str = "", required: bool = True,
                            label: str = "Enunciado", compact: bool = False, min_height: int = 120,
                            placeholder: str = "", detectar_alternativas: bool = False) -> str:
    """Editor WYSIWYG com toolbar EMBAIXO do conteúdo (estilo Slack/Discord).
    - compact=True mostra só B / I / U / limpar (pra campos curtos como alternativas).
    - placeholder aparece DENTRO da caixa quando vazia, some ao digitar.
    - detectar_alternativas=True: ao colar texto com "A) ... B) ... C) ... D) ...",
      oferece extrair as alternativas pros campos alt_a/alt_b/alt_c/alt_d automaticamente.
    O HTML editado é sincronizado num <textarea hidden> que vai no submit."""
    import html as _html
    valor_escapado_textarea = _html.escape(valor_inicial or "")
    req_attr = " required" if required else ""

    # Toolbar: botões variam conforme compact
    btn_style = "padding:3px 7px; background:transparent; border:1px solid var(--border); border-radius:3px; cursor:pointer; font-family:inherit; font-size:12px; color:inherit;"
    bot_basicos = (
        f'<button type="button" data-cmd="bold" title="Negrito (Ctrl+B)" style="{btn_style} font-weight:700; min-width:26px;">B</button>'
        f'<button type="button" data-cmd="italic" title="Itálico (Ctrl+I)" style="{btn_style} font-style:italic; min-width:26px;">I</button>'
        f'<button type="button" data-cmd="underline" title="Sublinhado (Ctrl+U)" style="{btn_style} text-decoration:underline; min-width:26px;">U</button>'
    )
    sep = '<span style="border-left:1px solid var(--border); margin:0 2px;"></span>'
    bot_extra = (
        f'<button type="button" data-cmd="justifyLeft" title="Alinhar à esquerda" style="{btn_style}">⇤</button>'
        f'<button type="button" data-cmd="justifyCenter" title="Centralizar" style="{btn_style}">⇔</button>'
        f'<button type="button" data-cmd="justifyRight" title="Alinhar à direita" style="{btn_style}">⇥</button>'
        f'{sep}'
        f'<button type="button" data-cmd="insertUnorderedList" title="Lista" style="{btn_style}">• Lista</button>'
        f'<button type="button" data-cmd="formatBlock" data-arg="blockquote" title="Citação" style="{btn_style}">❝ Citação</button>'
        f'{sep}'
    )
    bot_limpar = f'<button type="button" data-cmd="removeFormat" title="Limpar formatação" style="{btn_style} color:var(--text-muted);">⌫ limpar</button>'
    bot_fracao = f'<button type="button" class="btn-insert-frac" title="Inserir fração como $\\frac{{num}}{{den}}$" style="{btn_style}">½ fração</button>'
    bot_potencia = f'<button type="button" class="btn-insert-pot" title="Inserir potência como $base^{{exp}}$" style="{btn_style}">x² potência</button>'
    bot_tabela = f'<button type="button" class="btn-insert-tab" title="Inserir tabela" style="{btn_style}">⊞ tabela</button>'

    toolbar_buttons = bot_basicos + sep + bot_fracao + bot_potencia + bot_tabela + sep + bot_limpar if compact else bot_basicos + sep + bot_extra + bot_fracao + bot_potencia + bot_tabela + sep + bot_limpar

    placeholder_attr = f' data-placeholder="{_html.escape(placeholder, quote=True)}"' if placeholder else ""

    return f"""
        <style>
            .editor-content[data-placeholder]:empty::before {{
                content: attr(data-placeholder);
                color: var(--text-muted);
                opacity: 0.7;
                pointer-events: none;
                font-style: italic;
            }}
            .ed-wrap:focus-within {{ box-shadow: 0 0 0 2px rgba(59,130,246,0.3); border-color: var(--accent); }}
            .editor-content blockquote {{ margin: 8px 0; padding: 6px 14px; border-left: 3px solid var(--border); color: var(--text-muted); font-style: italic; }}
            .editor-content ul {{ margin: 6px 0 6px 22px; }}
        </style>
        <label style="display:block; margin:8px 0;">{label}
            <div class="ed-wrap" style="border:1px solid var(--border); border-radius:5px; background:var(--bg); overflow:hidden;">
                <div class="editor-content" contenteditable="true" data-target="{name}"{placeholder_attr} style="min-height:{min_height}px; padding:10px 12px; outline:none; font-family:inherit; font-size:14px; line-height:1.5;">{valor_inicial}</div>
                <div class="editor-toolbar" style="display:flex; gap:3px; flex-wrap:wrap; align-items:center; padding:5px 7px; background:var(--bg-subtle); border-top:1px solid var(--border);">
                    {toolbar_buttons}
                </div>
            </div>
            <textarea name="{name}" id="{name}_hidden" style="display:none;"{req_attr}>{valor_escapado_textarea}</textarea>
        </label>
        <script>
        (function() {{
            const editor = document.querySelector('.editor-content[data-target="{name}"]');
            const hidden = document.getElementById('{name}_hidden');
            if (!editor || !hidden) return;
            function sync() {{ hidden.value = editor.innerHTML; }}
            editor.addEventListener('input', sync);
            editor.addEventListener('blur', sync);
            const form = editor.closest('form');
            if (form) form.addEventListener('submit', sync);

            // Placeholder: mostra quando vazio (via CSS :empty já cobre em alguns browsers; aqui garantimos)
            const ph = editor.getAttribute('data-placeholder') || '';
            function refreshPlaceholder() {{
                const isEmpty = editor.innerHTML.trim() === '' || editor.innerHTML.trim() === '<br>';
                if (isEmpty && ph && !editor.hasAttribute('data-ph-shown')) {{
                    editor.setAttribute('data-ph-shown', '1');
                    editor.style.position = 'relative';
                }}
                if (!isEmpty) editor.removeAttribute('data-ph-shown');
            }}
            editor.addEventListener('input', refreshPlaceholder);
            refreshPlaceholder();

            const toolbar = editor.parentNode.querySelector('.editor-toolbar');
            if (toolbar) {{
                toolbar.querySelectorAll('button[data-cmd]').forEach(btn => {{
                    btn.addEventListener('click', e => {{
                        e.preventDefault();
                        const cmd = btn.getAttribute('data-cmd');
                        const arg = btn.getAttribute('data-arg') || null;
                        editor.focus();
                        try {{ document.execCommand(cmd, false, arg); }} catch(err) {{}}
                        sync();
                        refreshPlaceholder();
                    }});
                }});
            }}
            {_JS_MATH_BUTTONS}

            {_JS_DETECTAR_ALTS if detectar_alternativas else ""}

            // Paste de imagem (todos os campos, incluindo alternativas)
            editor.addEventListener('paste', function(e) {{
                var cb = e.clipboardData || window.clipboardData;
                if (!cb) return;
                var items = cb.items ? Array.from(cb.items) : [];
                var imgItem = items.find(function(it) {{ return it.type.startsWith('image/'); }});
                if (!imgItem) return;  // texto é tratado pelo handler acima (se existir) ou pelo browser
                e.preventDefault();
                var blob = imgItem.getAsFile();
                if (!blob) return;
                var fd = new FormData();
                fd.append('arquivo', blob, 'imagem_colada.png');
                fetch('/upload-imagem-inline', {{ method: 'POST', body: fd }})
                    .then(function(r) {{ return r.json(); }})
                    .then(function(data) {{
                        if (data.url) {{
                            document.execCommand('insertHTML', false,
                                '<img src="' + data.url + '" style="max-width:100%; height:auto; display:block; margin:4px 0;" alt="">');
                            sync();
                        }}
                    }})
                    .catch(function() {{ alert('Erro ao fazer upload da imagem.'); }});
            }}, true);  // capture=true para rodar antes do handler de texto
        }})();
        </script>
    """
    """Autor da questão OU admin podem editar. Questões legadas (sem dono) só admin edita."""
    if not prof:
        return False
    if prof["is_admin"]:
        return True
    if questao_criador_id is None:
        return False
    return prof["id"] == questao_criador_id


def _require_admin_or_403(request: Request) -> HTMLResponse:
    """Retorna None se admin, ou HTMLResponse 403 se não. Helper p/ rotas internas."""
    prof = get_current_professor(request)
    if not prof or not prof["is_admin"]:
        return HTMLResponse(render_page(
            "Acesso restrito",
            '<div class="page-header"><h1>🔒 Acesso restrito</h1></div>'
            '<div style="background:var(--red-bg); color:var(--red); border:1px solid var(--red); padding:16px; border-radius:6px;">'
            '<p>Apenas o administrador da escola pode criar, editar ou excluir <strong>turmas e estudantes</strong>.</p>'
            '<p>Se você precisa de uma turma cadastrada, fale com o administrador.</p>'
            '</div>'
            '<div class="page-actions" style="margin-top:16px;"><a href="/turmas" class="btn">← Ver turmas</a></div>',
            active="turmas"
        ), status_code=403)
    return None


# ContextVar pra propagar prof logado pra render_page sem mudar 50 assinaturas
import contextvars
_current_prof_ctx: contextvars.ContextVar[Optional[dict]] = contextvars.ContextVar("current_prof", default=None)


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS disciplinas (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL UNIQUE);
        CREATE TABLE IF NOT EXISTS questoes (id INTEGER PRIMARY KEY AUTOINCREMENT, disciplina_id INTEGER NOT NULL, enunciado TEXT NOT NULL, FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id));
        CREATE TABLE IF NOT EXISTS alternativas (id INTEGER PRIMARY KEY AUTOINCREMENT, questao_id INTEGER NOT NULL, letra TEXT NOT NULL, texto TEXT NOT NULL, correta INTEGER NOT NULL DEFAULT 0, FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE);
        CREATE TABLE IF NOT EXISTS textos_apoio (id INTEGER PRIMARY KEY AUTOINCREMENT, questao_id INTEGER NOT NULL, conteudo TEXT NOT NULL, fonte TEXT, ordem INTEGER NOT NULL DEFAULT 0, FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE);
        CREATE TABLE IF NOT EXISTS imagens (id INTEGER PRIMARY KEY AUTOINCREMENT, questao_id INTEGER NOT NULL, caminho TEXT NOT NULL, legenda TEXT, fonte TEXT, ordem INTEGER NOT NULL DEFAULT 0, FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE);
        CREATE TABLE IF NOT EXISTS provas (id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, descricao TEXT, criada_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS prova_questoes (id INTEGER PRIMARY KEY AUTOINCREMENT, prova_id INTEGER NOT NULL, questao_id INTEGER NOT NULL, ordem INTEGER NOT NULL DEFAULT 0, FOREIGN KEY (prova_id) REFERENCES provas(id) ON DELETE CASCADE, FOREIGN KEY (questao_id) REFERENCES questoes(id));
        CREATE TABLE IF NOT EXISTS habilidades_bncc (id INTEGER PRIMARY KEY AUTOINCREMENT, codigo TEXT NOT NULL UNIQUE, descricao TEXT);
        CREATE TABLE IF NOT EXISTS questao_habilidades (id INTEGER PRIMARY KEY AUTOINCREMENT, questao_id INTEGER NOT NULL, habilidade_id INTEGER NOT NULL, UNIQUE(questao_id, habilidade_id), FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE, FOREIGN KEY (habilidade_id) REFERENCES habilidades_bncc(id));
        CREATE TABLE IF NOT EXISTS turmas (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL, ano_letivo INTEGER NOT NULL, criada_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS alunos (id INTEGER PRIMARY KEY AUTOINCREMENT, turma_id INTEGER NOT NULL, nome TEXT NOT NULL, numero INTEGER, codigo_unico TEXT NOT NULL UNIQUE, raca TEXT, email TEXT, data_nascimento TEXT, FOREIGN KEY (turma_id) REFERENCES turmas(id) ON DELETE CASCADE);
        CREATE TABLE IF NOT EXISTS aplicacoes (id INTEGER PRIMARY KEY AUTOINCREMENT, prova_id INTEGER NOT NULL, turma_id INTEGER NOT NULL, modo TEXT NOT NULL DEFAULT 'online', titulo TEXT, aberta INTEGER NOT NULL DEFAULT 1, criada_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (prova_id) REFERENCES provas(id), FOREIGN KEY (turma_id) REFERENCES turmas(id));
        CREATE TABLE IF NOT EXISTS respostas (id INTEGER PRIMARY KEY AUTOINCREMENT, aplicacao_id INTEGER NOT NULL, aluno_id INTEGER NOT NULL, questao_id INTEGER NOT NULL, alternativa_letra TEXT, respondida_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, UNIQUE(aplicacao_id, aluno_id, questao_id), FOREIGN KEY (aplicacao_id) REFERENCES aplicacoes(id) ON DELETE CASCADE, FOREIGN KEY (aluno_id) REFERENCES alunos(id), FOREIGN KEY (questao_id) REFERENCES questoes(id));
        CREATE TABLE IF NOT EXISTS entregas (id INTEGER PRIMARY KEY AUTOINCREMENT, aplicacao_id INTEGER NOT NULL, aluno_id INTEGER NOT NULL, finalizada_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, UNIQUE(aplicacao_id, aluno_id), FOREIGN KEY (aplicacao_id) REFERENCES aplicacoes(id) ON DELETE CASCADE, FOREIGN KEY (aluno_id) REFERENCES alunos(id));
        CREATE TABLE IF NOT EXISTS professores (id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT NOT NULL UNIQUE, nome TEXT NOT NULL, foto_url TEXT, is_admin INTEGER NOT NULL DEFAULT 0, criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP, ultimo_acesso TIMESTAMP);
    """)
    cols = {row[1] for row in conn.execute("PRAGMA table_info(alunos)").fetchall()}
    if "raca" not in cols:
        conn.execute("ALTER TABLE alunos ADD COLUMN raca TEXT")
    if "email" not in cols:
        conn.execute("ALTER TABLE alunos ADD COLUMN email TEXT")
    if "data_nascimento" not in cols:
        conn.execute("ALTER TABLE alunos ADD COLUMN data_nascimento TEXT")
    if "sexo" not in cols:
        conn.execute("ALTER TABLE alunos ADD COLUMN sexo TEXT")
    if "codigo_rede" not in cols:
        # Código do aluno na plataforma oficial da rede (e-cidade). Populado na 1ª importação
        # do Conselho de Classe por casamento de nome; usado como chave estável nas seguintes (24/08/2026).
        conn.execute("ALTER TABLE alunos ADD COLUMN codigo_rede TEXT")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alunos_codigo_rede ON alunos(codigo_rede)")

    cols_prof = {row[1] for row in conn.execute("PRAGMA table_info(professores)").fetchall()}
    if "papel" not in cols_prof:
        # 'docente', 'gestao' ou 'apoio' (Apoio Educacional) — preenchido no onboarding
        # obrigatório do primeiro acesso (admin nunca precisa, já vê a escola toda) — 24/08/2026.
        conn.execute("ALTER TABLE professores ADD COLUMN papel TEXT")

    # Módulo Administrativo: solicitações de afastamento (atestado/permissão/abono) com
    # anexo salvo no Google Drive — 25/08/2026.
    conn.execute("""CREATE TABLE IF NOT EXISTS afastamentos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        professor_id INTEGER NOT NULL,
        tipo TEXT NOT NULL,
        data_inicio TEXT NOT NULL,
        data_fim TEXT NOT NULL,
        observacao TEXT,
        arquivo_nome TEXT,
        arquivo_drive_id TEXT,
        arquivo_drive_link TEXT,
        status_upload TEXT NOT NULL DEFAULT 'pendente',
        criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (professor_id) REFERENCES professores(id) ON DELETE CASCADE
    )""")
    cols_afast = {row[1] for row in conn.execute("PRAGMA table_info(afastamentos)").fetchall()}
    if "horario_inicio" not in cols_afast:
        # Só preenchido quando o tipo é 'Permissão de ausência' — a pessoa saiu/voltou
        # num horário específico do dia, não o dia inteiro (26/08/2026).
        conn.execute("ALTER TABLE afastamentos ADD COLUMN horario_inicio TEXT")
    if "horario_fim" not in cols_afast:
        conn.execute("ALTER TABLE afastamentos ADD COLUMN horario_fim TEXT")

    cols_q = {row[1] for row in conn.execute("PRAGMA table_info(questoes)").fetchall()}
    if "ano" not in cols_q:
        conn.execute("ALTER TABLE questoes ADD COLUMN ano TEXT")
    if "criada_por_professor_id" not in cols_q:
        conn.execute("ALTER TABLE questoes ADD COLUMN criada_por_professor_id INTEGER")
    if "tipo" not in cols_q:
        conn.execute("ALTER TABLE questoes ADD COLUMN tipo TEXT DEFAULT 'multipla_escolha'")
        # Questões antigas viram múltipla escolha (que era o único tipo até agora)
        conn.execute("UPDATE questoes SET tipo = 'multipla_escolha' WHERE tipo IS NULL")
    if "anulada" not in cols_q:
        conn.execute("ALTER TABLE questoes ADD COLUMN anulada INTEGER DEFAULT 0")
    if "gabarito_original" not in cols_q:
        conn.execute("ALTER TABLE questoes ADD COLUMN gabarito_original TEXT")

    # Migração: respostas ganham coluna pra V/F e Associação (JSON)
    cols_resp = {row[1] for row in conn.execute("PRAGMA table_info(respostas)").fetchall()}
    if "dados_extra" not in cols_resp:
        conn.execute("ALTER TABLE respostas ADD COLUMN dados_extra TEXT")
    if "credito_anulacao" not in cols_resp:
        conn.execute("ALTER TABLE respostas ADD COLUMN credito_anulacao INTEGER DEFAULT 0")
    if "resposta_texto" not in cols_resp:
        # Texto livre do aluno pra questões discursivas na aplicação online (12/08/2026)
        conn.execute("ALTER TABLE respostas ADD COLUMN resposta_texto TEXT")

    # Tabelas pra V ou F
    conn.execute("""CREATE TABLE IF NOT EXISTS vf_afirmacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        questao_id INTEGER NOT NULL,
        ordem INTEGER NOT NULL,
        texto TEXT NOT NULL,
        gabarito TEXT NOT NULL CHECK(gabarito IN ('V','F')),
        FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE
    )""")

    # Tabelas pra Associação de colunas
    conn.execute("""CREATE TABLE IF NOT EXISTS assoc_itens_a (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        questao_id INTEGER NOT NULL,
        ordem INTEGER NOT NULL,
        texto TEXT NOT NULL,
        gabarito_letra TEXT NOT NULL,
        FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS assoc_itens_b (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        questao_id INTEGER NOT NULL,
        letra TEXT NOT NULL,
        texto TEXT NOT NULL,
        FOREIGN KEY (questao_id) REFERENCES questoes(id) ON DELETE CASCADE
    )""")
    # Migrations multi-prof + gestão
    cols_p = {row[1] for row in conn.execute("PRAGMA table_info(provas)").fetchall()}
    if "criada_por_professor_id" not in cols_p:
        conn.execute("ALTER TABLE provas ADD COLUMN criada_por_professor_id INTEGER")
    if "status_revisao" not in cols_p:
        conn.execute("ALTER TABLE provas ADD COLUMN status_revisao TEXT NOT NULL DEFAULT 'rascunho'")
    if "obs_gestao" not in cols_p:
        conn.execute("ALTER TABLE provas ADD COLUMN obs_gestao TEXT")
    if "revisado_por_id" not in cols_p:
        conn.execute("ALTER TABLE provas ADD COLUMN revisado_por_id INTEGER")
    if "revisado_em" not in cols_p:
        conn.execute("ALTER TABLE provas ADD COLUMN revisado_em TIMESTAMP")
    cols_a = {row[1] for row in conn.execute("PRAGMA table_info(aplicacoes)").fetchall()}
    if "criada_por_professor_id" not in cols_a:
        conn.execute("ALTER TABLE aplicacoes ADD COLUMN criada_por_professor_id INTEGER")
    if "mostrar_resultado_aluno" not in cols_a:
        # Flag: professor decide se aluno vê nota/gabarito ao entregar (12/08/2026).
        # Default 1 (mostra) — preserva o comportamento atual para aplicações já existentes.
        conn.execute("ALTER TABLE aplicacoes ADD COLUMN mostrar_resultado_aluno INTEGER NOT NULL DEFAULT 1")
    cols_prof = {row[1] for row in conn.execute("PRAGMA table_info(professores)").fetchall()}
    if "is_gestor" not in cols_prof:
        conn.execute("ALTER TABLE professores ADD COLUMN is_gestor INTEGER NOT NULL DEFAULT 0")
    if "status" not in cols_prof:
        conn.execute("ALTER TABLE professores ADD COLUMN status TEXT NOT NULL DEFAULT 'ativo'")
        conn.execute("UPDATE professores SET status = 'ativo' WHERE status IS NULL OR status = ''")
    # Simulado
    conn.execute("""CREATE TABLE IF NOT EXISTS simulados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        trimestre INTEGER NOT NULL,
        ano INTEGER NOT NULL,
        ano_escolaridade INTEGER,
        pontuacao_total REAL NOT NULL DEFAULT 10.0,
        status TEXT NOT NULL DEFAULT 'montagem',
        dia INTEGER NOT NULL DEFAULT 1,
        ordem INTEGER NOT NULL DEFAULT 0,
        criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        criado_por_professor_id INTEGER,
        FOREIGN KEY (criado_por_professor_id) REFERENCES professores(id)
    )""")
    # Migration para bancos existentes
    cols_sim = {row[1] for row in conn.execute("PRAGMA table_info(simulados)").fetchall()}
    if "ano_escolaridade" not in cols_sim:
        conn.execute("ALTER TABLE simulados ADD COLUMN ano_escolaridade INTEGER")
    if "ordem" not in cols_sim:
        conn.execute("ALTER TABLE simulados ADD COLUMN ordem INTEGER NOT NULL DEFAULT 0")
        sims = conn.execute("SELECT id FROM simulados ORDER BY ano DESC, trimestre DESC, id DESC").fetchall()
        for i, s in enumerate(sims):
            conn.execute("UPDATE simulados SET ordem = ? WHERE id = ?", (i, s["id"]))
    if "dia" not in cols_sim:
        conn.execute("ALTER TABLE simulados ADD COLUMN dia INTEGER NOT NULL DEFAULT 1")
    if "turma_id" in cols_sim:
        pass  # mantém coluna legada sem remover
    conn.execute("""CREATE TABLE IF NOT EXISTS simulado_blocos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        simulado_id INTEGER NOT NULL,
        numero INTEGER NOT NULL,
        disciplina_id INTEGER NOT NULL,
        n_questoes INTEGER NOT NULL DEFAULT 10,
        tempo_minutos INTEGER NOT NULL DEFAULT 25,
        status TEXT NOT NULL DEFAULT 'aguardando',
        professor_id INTEGER,
        FOREIGN KEY (simulado_id) REFERENCES simulados(id) ON DELETE CASCADE,
        FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id),
        FOREIGN KEY (professor_id) REFERENCES professores(id)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS simulado_questoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bloco_id INTEGER NOT NULL,
        questao_id INTEGER NOT NULL,
        ordem INTEGER NOT NULL DEFAULT 0,
        FOREIGN KEY (bloco_id) REFERENCES simulado_blocos(id) ON DELETE CASCADE,
        FOREIGN KEY (questao_id) REFERENCES questoes(id)
    )""")

    # ── BOLETIM / CONSELHO DE CLASSE (incorporado 05/08/2026) ──────────────
    # Reaproveita alunos/turmas/disciplinas/professores já existentes. As 3
    # tabelas abaixo guardam os dados por TRIMESTRE+ANO desde o início, pra
    # não sobrescrever um trimestre com o seguinte.
    conn.execute("""CREATE TABLE IF NOT EXISTS boletim_medias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        aluno_id INTEGER NOT NULL,
        disciplina_id INTEGER NOT NULL,
        trimestre INTEGER NOT NULL,
        ano INTEGER NOT NULL,
        nota REAL,
        nota_texto TEXT,
        UNIQUE(aluno_id, disciplina_id, trimestre, ano),
        FOREIGN KEY (aluno_id) REFERENCES alunos(id) ON DELETE CASCADE,
        FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS boletim_faltas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        aluno_id INTEGER NOT NULL,
        disciplina_id INTEGER NOT NULL,
        trimestre INTEGER NOT NULL,
        ano INTEGER NOT NULL,
        faltas INTEGER NOT NULL DEFAULT 0,
        UNIQUE(aluno_id, disciplina_id, trimestre, ano),
        FOREIGN KEY (aluno_id) REFERENCES alunos(id) ON DELETE CASCADE,
        FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS boletim_analise (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        aluno_id INTEGER NOT NULL,
        disciplina_id INTEGER NOT NULL,
        professor_id INTEGER,
        trimestre INTEGER NOT NULL,
        ano INTEGER NOT NULL,
        emocional TEXT,
        apoio INTEGER NOT NULL DEFAULT 0,
        alfabetizacao INTEGER NOT NULL DEFAULT 0,
        faltoso INTEGER NOT NULL DEFAULT 0,
        faltoso_json TEXT,
        observacao TEXT,
        atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(aluno_id, disciplina_id, trimestre, ano),
        FOREIGN KEY (aluno_id) REFERENCES alunos(id) ON DELETE CASCADE,
        FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id),
        FOREIGN KEY (professor_id) REFERENCES professores(id)
    )""")

    conn.execute("""CREATE TABLE IF NOT EXISTS boletim_professor_turma (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        professor_id INTEGER NOT NULL,
        turma_id INTEGER NOT NULL,
        disciplina_id INTEGER NOT NULL,
        UNIQUE(professor_id, turma_id, disciplina_id),
        FOREIGN KEY (professor_id) REFERENCES professores(id) ON DELETE CASCADE,
        FOREIGN KEY (turma_id) REFERENCES turmas(id) ON DELETE CASCADE,
        FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id)
    )""")

    # Migrações — cobre bancos onde essas tabelas já existiam ANTES dessas colunas serem
    # adicionadas (CREATE TABLE IF NOT EXISTS não altera uma tabela que já existe).
    cols_bm = {row[1] for row in conn.execute("PRAGMA table_info(boletim_medias)").fetchall()}
    if "nota_texto" not in cols_bm:
        conn.execute("ALTER TABLE boletim_medias ADD COLUMN nota_texto TEXT")
    cols_ba = {row[1] for row in conn.execute("PRAGMA table_info(boletim_analise)").fetchall()}
    if "faltoso" not in cols_ba:
        conn.execute("ALTER TABLE boletim_analise ADD COLUMN faltoso INTEGER NOT NULL DEFAULT 0")

    conn.commit()
    conn.close()


init_db()


# ==========================================
#  FILA GLOBAL DE ESCANEAMENTO (OMR EM SEGUNDO PLANO)
# ==========================================
# Processa os cartões de TODOS os professores em segundo plano, sem travar o
# servidor. Cada foto vira um "item" numa fila global (FIFO); alguns
# "trabalhadores" (workers) vão pegando item por item e processando em
# threads separadas (usando a correção já aplicada com asyncio.to_thread).
#
# IMPORTANTE: essa fila fica em memória. Se o serviço for reiniciado
# (systemctl restart) no meio de um processamento, os lotes em andamento se
# perdem e precisam ser reenviados pelo professor.

N_WORKERS_ESCANEAMENTO = 2  # processos "trabalhadores" simultâneos; pode subir para 3-4 numa VM maior (e2-medium+)

FILAS_ESCANEAMENTO: dict = {}          # lote_id -> job (dict)
FILA_GLOBAL_ESCANEAMENTO: "asyncio.Queue" = None  # criada no startup do app (ou sob demanda, ver _garantir_fila_escaneamento)
TEMPOS_GLOBAIS_ESCANEAMENTO = deque(maxlen=30)    # últimos tempos de processamento (segundos), pra estimar ETA
_WORKERS_ESCANEAMENTO_INICIADOS = False


def _garantir_fila_escaneamento():
    """Cria a fila global e os workers se ainda não existirem. Serve tanto pro evento
    de startup normal quanto como proteção: se por algum motivo o startup não tiver
    rodado (ex.: variação de como o servidor foi iniciado), a fila se cria sozinha na
    primeira vez que alguém tentar escanear, em vez de quebrar com 'internal server error'."""
    global FILA_GLOBAL_ESCANEAMENTO, _WORKERS_ESCANEAMENTO_INICIADOS
    if FILA_GLOBAL_ESCANEAMENTO is None:
        FILA_GLOBAL_ESCANEAMENTO = asyncio.Queue()
    if not _WORKERS_ESCANEAMENTO_INICIADOS:
        _WORKERS_ESCANEAMENTO_INICIADOS = True
        for i in range(N_WORKERS_ESCANEAMENTO):
            asyncio.create_task(_worker_escaneamento(i))


def _novo_lote_escaneamento(tipo: str, contexto: dict, arquivos: list) -> str:
    """Cria um lote de escaneamento e enfileira cada foto como um item.
    arquivos: lista de tuplas (filename, image_bytes). Retorna o lote_id."""
    _garantir_fila_escaneamento()
    lote_id = uuid.uuid4().hex[:12]
    agora = time.monotonic()
    itens = []
    for i, (filename, image_bytes) in enumerate(arquivos):
        itens.append({
            "idx": i, "filename": filename, "bytes": image_bytes,
            "status": "pendente", "resultado": None,
            "seq": agora + i * 1e-6,
        })
    job = {
        "id": lote_id, "tipo": tipo, "contexto": contexto,
        "itens": itens, "total": len(itens), "processados": 0,
        "criado_em": datetime.now(), "concluido": len(itens) == 0,
    }
    FILAS_ESCANEAMENTO[lote_id] = job
    for item in itens:
        FILA_GLOBAL_ESCANEAMENTO.put_nowait((lote_id, item["idx"]))
    return lote_id


async def _enriquecer_item_simulado(job: dict, item: dict):
    """Para simulados: anota nome do aluno, se já tem entrega e se é duplicata dentro do
    próprio lote — SEM gravar nada no banco ainda. A gravação só acontece quando o
    professor revisa e confirma (igual já funciona na prova normal)."""
    resultado = item["resultado"]
    ctx = job["contexto"]
    if not resultado or not resultado.get("success"):
        return
    conn = get_db()
    try:
        aluno_id = resultado["aluno_id"]
        aluno = conn.execute("SELECT * FROM alunos WHERE id = ?", (aluno_id,)).fetchone()
        if not aluno:
            item["resultado"] = {"success": False, "error": f"Aluno {aluno_id} não encontrado.",
                                  "filename": item["filename"]}
            return
        item["resultado"]["aluno_nome"] = aluno["nome"]
        item["resultado"]["aluno_numero"] = aluno["numero"]
        item["resultado"]["aluno_codigo"] = aluno["codigo_unico"]
        item["resultado"]["ja_entregue"] = conn.execute(
            "SELECT id FROM entregas WHERE aplicacao_id = ? AND aluno_id = ?",
            (ctx["app_id"], aluno_id)
        ).fetchone() is not None
        item["resultado"]["duplicado"] = aluno_id in ctx.setdefault("_alunos_no_lote", set())
        ctx["_alunos_no_lote"].add(aluno_id)
    finally:
        conn.close()


async def _enriquecer_item_prova(job: dict, item: dict):
    """Para prova normal: só anota o nome do aluno no resultado (não mexe em mais
    nada — diferente do simulado, a prova normal já trata duplicata/entrega em
    outro lugar). Serve pra mostrar o nome na lista ao vivo de progresso do
    escaneamento, em vez de só um contador (28/07/2026)."""
    resultado = item["resultado"]
    if not resultado or not resultado.get("success"):
        return
    aluno_id = resultado.get("aluno_id")
    if not aluno_id:
        return
    conn = get_db()
    try:
        aluno = conn.execute("SELECT nome FROM alunos WHERE id = ?", (aluno_id,)).fetchone()
        if aluno:
            item["resultado"]["aluno_nome"] = aluno["nome"]
    finally:
        conn.close()


async def _worker_escaneamento(worker_num: int):
    """Loop infinito: pega o próximo item da fila global e processa (numa thread separada,
    sem travar o servidor pros outros usuários)."""
    while True:
        lote_id, item_idx = await FILA_GLOBAL_ESCANEAMENTO.get()
        job = FILAS_ESCANEAMENTO.get(lote_id)
        if job is None:
            FILA_GLOBAL_ESCANEAMENTO.task_done()
            continue
        item = job["itens"][item_idx]
        item["status"] = "processando"
        t0 = time.monotonic()
        try:
            if not item["bytes"]:
                item["resultado"] = {"success": False, "error": "Arquivo vazio.", "filename": item["filename"]}
            elif job["tipo"] == "prova":
                ctx = job["contexto"]
                item["resultado"] = await asyncio.to_thread(
                    _processar_cartao_resposta, item["bytes"], ctx["n_questoes"],
                    filename=item["filename"] or "", questoes_info=ctx["questoes_info"]
                )
                await _enriquecer_item_prova(job, item)
            else:  # simulado
                ctx = job["contexto"]
                item["resultado"] = await asyncio.to_thread(
                    _processar_cartao_simulado, item["bytes"], ctx["blocos_info"], item["filename"] or ""
                )
                await _enriquecer_item_simulado(job, item)
        except Exception as e:
            item["resultado"] = {"success": False, "error": f"Erro inesperado no processamento: {e}",
                                  "filename": item["filename"]}
        finally:
            item["bytes"] = None  # libera a memória da foto assim que processa
            item["status"] = "concluido"
            dt = time.monotonic() - t0
            TEMPOS_GLOBAIS_ESCANEAMENTO.append(dt)
            job["processados"] += 1
            if job["processados"] >= job["total"]:
                job["concluido"] = True
            FILA_GLOBAL_ESCANEAMENTO.task_done()


@app.on_event("startup")
async def _iniciar_workers_escaneamento():
    _garantir_fila_escaneamento()


def _status_lote_escaneamento(lote_id: str) -> Optional[dict]:
    """Monta o payload de status/progresso de um lote: quantos faltam, posição na
    fila global (cartões de outros lotes ainda na frente) e tempo estimado."""
    job = FILAS_ESCANEAMENTO.get(lote_id)
    if not job:
        return None

    tempo_medio = (sum(TEMPOS_GLOBAIS_ESCANEAMENTO) / len(TEMPOS_GLOBAIS_ESCANEAMENTO)) if TEMPOS_GLOBAIS_ESCANEAMENTO else 2.5

    pendentes_deste_lote = [it for it in job["itens"] if it["status"] == "pendente"]
    posicao_fila_global = 0
    if pendentes_deste_lote:
        primeiro_seq = pendentes_deste_lote[0]["seq"]
        for outro_job in FILAS_ESCANEAMENTO.values():
            for it in outro_job["itens"]:
                if it["status"] in ("pendente", "processando") and it["seq"] < primeiro_seq:
                    posicao_fila_global += 1

    itens_restantes = job["total"] - job["processados"]
    eta_segundos = math.ceil((posicao_fila_global + itens_restantes) / N_WORKERS_ESCANEAMENTO * tempo_medio) if itens_restantes else 0

    # Lista dos últimos cartões concluídos, pra mostrar um checklist ao vivo em vez
    # de só uma porcentagem abstrata (melhoria de UX, 28/07/2026 — inspirado em como
    # apps de mercado tipo ZipGrade mostram feedback item a item, não só um total).
    concluidos = [it for it in job["itens"] if it["status"] == "concluido"]
    concluidos_recentes = []
    for it in concluidos[-8:]:
        r = it.get("resultado") or {}
        ok = bool(r.get("success"))
        nome = r.get("aluno_nome") or it.get("filename") or "cartão"
        concluidos_recentes.append({
            "nome": nome,
            "ok": ok,
            "erro": (r.get("error") or "não foi possível ler") if not ok else None,
        })

    return {
        "lote_id": lote_id,
        "tipo": job["tipo"],
        "total": job["total"],
        "processados": job["processados"],
        "concluido": job["concluido"],
        "posicao_fila": posicao_fila_global,
        "eta_segundos": eta_segundos,
        "concluidos_recentes": concluidos_recentes,
        "redirect_url": job["contexto"].get("revisar_url") if job["concluido"] else None,
    }


# ==========================================
#  AUTENTICAÇÃO MULTI-PROFESSOR (D1)
# ==========================================

def _criar_sessao(professor_id: int, email: str) -> str:
    return _session_serializer.dumps({"professor_id": professor_id, "email": email})


def _ler_sessao(token: str) -> Optional[dict]:
    if not token:
        return None
    try:
        return _session_serializer.loads(token, max_age=SESSION_MAX_AGE)
    except (BadSignature, SignatureExpired):
        return None


def get_current_professor(request: Request) -> Optional[dict]:
    """Retorna dict com id/email/nome/is_admin do professor logado, ou None."""
    token = request.cookies.get(SESSION_COOKIE)
    payload = _ler_sessao(token)
    if not payload:
        return None
    conn = get_db()
    prof = conn.execute("SELECT * FROM professores WHERE id = ?", (payload["professor_id"],)).fetchone()
    conn.close()
    if not prof:
        return None
    return {
        "id": prof["id"], "email": prof["email"], "nome": prof["nome"],
        "foto_url": prof["foto_url"], "is_admin": bool(prof["is_admin"]), "is_gestor": bool(prof["is_gestor"] if "is_gestor" in prof.keys() else 0), "status": (prof["status"] if "status" in prof.keys() else "ativo"),
        "papel": (prof["papel"] if "papel" in prof.keys() else None),
    }


# Rotas públicas (sem login)
PUBLIC_PATHS = {"/login", "/auth/google", "/auth/google/callback", "/auth/dev-login", "/logout", "/acesso-pendente", "/acesso-bloqueado"}
PUBLIC_PREFIXES = ("/static/", "/responder/")
# Rotas que continuam acessíveis mesmo pra quem ainda não completou o onboarding
# (senão o próprio onboarding fica inacessível — causaria loop de redirecionamento).
ONBOARDING_ISENTAS = {"/onboarding", "/logout"}


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in PUBLIC_PATHS or any(path.startswith(p) for p in PUBLIC_PREFIXES):
        return await call_next(request)
    prof = get_current_professor(request)
    if not prof:
        from_url = path + ("?" + request.url.query if request.url.query else "")
        return RedirectResponse(f"/login?next={urllib.parse.quote(from_url)}", status_code=303)
    status_prof = prof.get("status", "ativo")
    if status_prof == "pendente" and path != "/acesso-pendente":
        return RedirectResponse("/acesso-pendente", status_code=303)
    if status_prof == "bloqueado" and path != "/acesso-bloqueado":
        return RedirectResponse("/acesso-bloqueado", status_code=303)
    # Onboarding obrigatório (24/08/2026): professor sem papel definido (e que não é
    # admin) precisa escolher docente/gestão antes de usar o resto do sistema. Admin
    # nunca passa por isso — já enxerga a escola toda.
    if not prof.get("is_admin") and not prof.get("papel") and path not in ONBOARDING_ISENTAS:
        return RedirectResponse("/onboarding", status_code=303)
    # Apoio Educacional (Cuidadores/Agentes/Biblioteca/Apoio) só acessa a solicitação de
    # afastamento — nenhuma outra área do sistema (25/08/2026, a pedido).
    if prof.get("papel") == "apoio" and not prof.get("is_admin") and not prof.get("is_gestor"):
        rotas_permitidas_apoio = {
            "/", "/administrativo/afastamentos/novo", "/administrativo/afastamentos",
        }
        if path not in rotas_permitidas_apoio and path not in ONBOARDING_ISENTAS:
            return RedirectResponse("/administrativo/afastamentos/novo", status_code=303)
    request.state.professor = prof
    token = _current_prof_ctx.set(prof)
    try:
        return await call_next(request)
    finally:
        _current_prof_ctx.reset(token)


def _upsert_professor(email: str, nome: str, foto_url: Optional[str] = None) -> dict:
    """Cria ou atualiza professor. Primeiro = admin ativo. Demais = pendente até aprovação."""
    conn = get_db()
    existing = conn.execute("SELECT * FROM professores WHERE email = ?", (email,)).fetchone()
    if existing:
        conn.execute("UPDATE professores SET nome = ?, foto_url = ?, ultimo_acesso = CURRENT_TIMESTAMP WHERE id = ?",
                     (nome, foto_url, existing["id"]))
        prof_id = existing["id"]
        is_admin = bool(existing["is_admin"])
        is_gestor = bool(existing["is_gestor"] if "is_gestor" in existing.keys() else 0)
        status = existing["status"] if "status" in existing.keys() else "ativo"
    else:
        total = conn.execute("SELECT COUNT(*) AS c FROM professores").fetchone()["c"]
        is_admin_val = 1 if total == 0 else 0
        status_val = "ativo" if is_admin_val == 1 else "pendente"
        c = conn.execute(
            "INSERT INTO professores (email, nome, foto_url, is_admin, is_gestor, status, ultimo_acesso) VALUES (?, ?, ?, ?, 0, ?, CURRENT_TIMESTAMP)",
            (email, nome, foto_url, is_admin_val, status_val)
        )
        prof_id = c.lastrowid
        is_admin = bool(is_admin_val)
        is_gestor = False
        status = status_val
        if is_admin_val == 1:
            conn.execute("UPDATE provas SET criada_por_professor_id = ? WHERE criada_por_professor_id IS NULL", (prof_id,))
            conn.execute("UPDATE aplicacoes SET criada_por_professor_id = ? WHERE criada_por_professor_id IS NULL", (prof_id,))
            conn.execute("UPDATE questoes SET criada_por_professor_id = ? WHERE criada_por_professor_id IS NULL", (prof_id,))
    conn.commit()
    conn.close()
    return {"id": prof_id, "email": email, "nome": nome, "is_admin": is_admin, "is_gestor": is_gestor, "status": status}


@app.get("/onboarding", response_class=HTMLResponse)
def form_onboarding(request: Request):
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    if prof.get("is_admin") or prof.get("papel"):
        return RedirectResponse("/", status_code=303)

    conn = get_db()
    disciplinas = conn.execute("SELECT id, nome FROM disciplinas WHERE nome != 'Geral' ORDER BY nome").fetchall()
    turmas = conn.execute("SELECT id, nome, ano_letivo FROM turmas ORDER BY ano_letivo DESC, nome").fetchall()
    conn.close()

    opts_disc = "".join(
        f'<label style="display:flex; align-items:center; gap:8px; padding:8px 10px; border:1px solid var(--border); border-radius:6px; margin-bottom:6px; cursor:pointer;">'
        f'<input type="checkbox" name="disciplina_id" value="{d["id"]}" style="width:auto;"> {d["nome"]}</label>'
        for d in disciplinas
    )
    opts_turma = "".join(
        f'<label style="display:flex; align-items:center; gap:8px; padding:8px 10px; border:1px solid var(--border); border-radius:6px; margin-bottom:6px; cursor:pointer;">'
        f'<input type="checkbox" name="turma_id" value="{t["id"]}" style="width:auto;"> {t["nome"]} ({t["ano_letivo"]})</label>'
        for t in turmas
    )

    content = f"""
        <div style="max-width:560px; margin:40px auto; padding:0 16px;">
            <div class="page-header">
                <h1>👋 Bem-vindo(a), {prof["nome"]}</h1>
                <p class="subtitle">Antes de continuar, conta pra gente seu papel na escola — isso personaliza o que você vê na tela inicial.</p>
            </div>
            <form action="/onboarding" method="post">
                <fieldset>
                    <legend>Qual é o seu papel?</legend>
                    <label style="display:flex; align-items:center; gap:8px; padding:10px; border:1px solid var(--border); border-radius:6px; margin-bottom:8px; cursor:pointer;">
                        <input type="radio" name="papel" value="docente" required style="width:auto;" onchange="document.getElementById('bloco-docente').style.display='block';"> <strong>Docente</strong> — dou aula em turmas específicas
                    </label>
                    <label style="display:flex; align-items:center; gap:8px; padding:10px; border:1px solid var(--border); border-radius:6px; margin-bottom:8px; cursor:pointer;">
                        <input type="radio" name="papel" value="gestao" required style="width:auto;" onchange="document.getElementById('bloco-docente').style.display='none';"> <strong>Gestão</strong> — acompanho a escola toda
                    </label>
                    <label style="display:flex; align-items:center; gap:8px; padding:10px; border:1px solid var(--border); border-radius:6px; margin-bottom:8px; cursor:pointer;">
                        <input type="radio" name="papel" value="apoio" required style="width:auto;" onchange="document.getElementById('bloco-docente').style.display='none';"> <strong>Apoio Educacional</strong> — Cuidadores, Agentes Escolares, Biblioteca e Apoio
                    </label>
                </fieldset>

                <div id="bloco-docente" style="display:none;">
                    <div class="tip" style="margin-top:14px;">Marque as disciplinas e turmas que você leciona. Isso filtra o que aparece pra você — não afeta o que outros professores veem.</div>
                    <fieldset style="margin-top:10px;">
                        <legend>Disciplina(s) que você leciona</legend>
                        {opts_disc}
                    </fieldset>
                    <fieldset style="margin-top:10px;">
                        <legend>Turma(s) em que você leciona</legend>
                        {opts_turma}
                    </fieldset>
                </div>

                <div class="page-actions">
                    <button type="submit" class="btn btn-primary">Continuar</button>
                </div>
            </form>
        </div>
    """
    return HTMLResponse(render_page("Bem-vindo", content, active=""))


@app.post("/onboarding")
async def salvar_onboarding(request: Request):
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    if prof.get("is_admin") or prof.get("papel"):
        return RedirectResponse("/", status_code=303)

    form = await request.form()
    papel = form.get("papel")
    if papel not in ("docente", "gestao", "apoio"):
        return RedirectResponse("/onboarding", status_code=303)

    conn = get_db()
    conn.execute("UPDATE professores SET papel = ? WHERE id = ?", (papel, prof["id"]))

    if papel == "docente":
        disciplina_ids = [int(v) for v in form.getlist("disciplina_id") if v.strip().isdigit()]
        turma_ids = [int(v) for v in form.getlist("turma_id") if v.strip().isdigit()]
        # Grava o produto disciplina × turma selecionado — assume que o professor leciona
        # a(s) mesma(s) disciplina(s) em todas as turmas marcadas (caso comum; quem dá
        # disciplinas diferentes em turmas diferentes pode ajustar depois, se necessário).
        for did in disciplina_ids:
            for tid in turma_ids:
                conn.execute("""INSERT OR IGNORE INTO boletim_professor_turma (professor_id, turma_id, disciplina_id)
                                 VALUES (?, ?, ?)""", (prof["id"], tid, did))

    conn.commit()
    conn.close()
    return RedirectResponse("/", status_code=303)


# ============ MÓDULO ADMINISTRATIVO — afastamentos (25/08/2026) ============

@app.get("/administrativo/afastamentos/novo", response_class=HTMLResponse)
def form_novo_afastamento(request: Request):
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)

    opts_tipo = "".join(f'<option value="{k}">{v}</option>' for k, v in TIPOS_AFASTAMENTO.items())
    tipos_com_horario_js = json.dumps(list(TIPOS_COM_HORARIO))
    content = f"""
        <div class="page-header">
            <h1>📄 Nova Justificativa para o Ponto</h1>
            <p class="subtitle">Atestado médico, permissão de ausência, abono ou outro documento. Anexe o arquivo — ele é enviado direto pro Drive da escola.</p>
        </div>
        <form action="/administrativo/afastamentos/novo" method="post" enctype="multipart/form-data">
            <label>Tipo de documento
                <select name="tipo" id="sel-tipo-afastamento" required onchange="_toggleHorarioAfastamento(this.value)">
                    <option value="">— selecione —</option>
                    {opts_tipo}
                </select>
            </label>
            <div style="display:flex; flex-wrap:wrap; gap:14px;">
                <label style="flex:1 1 200px;">Data de início
                    <input type="date" name="data_inicio" required>
                </label>
                <label style="flex:1 1 200px;">Data de término
                    <input type="date" name="data_fim" required>
                </label>
            </div>
            <p style="font-size:12px; color:var(--text-muted); margin-top:-8px;">Pra um único dia, use a mesma data nos dois campos.</p>
            <div id="bloco-horario-afastamento" style="display:none;">
                <div class="tip" style="margin-bottom:10px;">Permissão de ausência: informe o horário de saída e de retorno nesse dia.</div>
                <div style="display:flex; flex-wrap:wrap; gap:14px;">
                    <label style="flex:1 1 200px;">Horário de saída
                        <input type="time" name="horario_inicio" id="inp-horario-inicio">
                    </label>
                    <label style="flex:1 1 200px;">Horário de retorno
                        <input type="time" name="horario_fim" id="inp-horario-fim">
                    </label>
                </div>
            </div>
            <label>Observação (opcional)
                <textarea name="observacao" rows="3" placeholder="Algum detalhe adicional, se necessário"></textarea>
            </label>
            <label>Documento (PDF, foto ou imagem do atestado/permissão)
                <input type="file" name="arquivo" accept=".pdf,.jpg,.jpeg,.png" required>
            </label>
            <div class="page-actions">
                <button type="submit" class="btn btn-primary">Enviar solicitação</button>
                <a href="/administrativo/afastamentos" class="btn">Ver minhas solicitações</a>
            </div>
        </form>
        <script>
        const _tiposComHorario = {tipos_com_horario_js};
        function _toggleHorarioAfastamento(valor) {{
            const bloco = document.getElementById('bloco-horario-afastamento');
            const precisa = _tiposComHorario.includes(valor);
            bloco.style.display = precisa ? 'block' : 'none';
            document.getElementById('inp-horario-inicio').required = precisa;
            document.getElementById('inp-horario-fim').required = precisa;
        }}
        </script>
    """
    return HTMLResponse(render_page("Nova Justificativa", content, active="administrativo-novo"))


@app.post("/administrativo/afastamentos/novo", response_class=HTMLResponse)
async def criar_afastamento(request: Request, tipo: str = Form(...), data_inicio: str = Form(...),
                             data_fim: str = Form(...), observacao: str = Form(""), arquivo: UploadFile = File(...),
                             horario_inicio: str = Form(""), horario_fim: str = Form("")):
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    if tipo not in TIPOS_AFASTAMENTO:
        return HTMLResponse(render_page("Erro", '<div class="page-header"><h1>Erro</h1></div><p>Tipo de documento inválido.</p><a href="/administrativo/afastamentos/novo" class="btn">Voltar</a>', active="administrativo-novo"))

    # Horário só é obrigatório pra tipos como Permissão de ausência (26/08/2026)
    horario_inicio = horario_inicio.strip() or None
    horario_fim = horario_fim.strip() or None
    if tipo in TIPOS_COM_HORARIO and (not horario_inicio or not horario_fim):
        return HTMLResponse(render_page("Erro", f'<div class="page-header"><h1>Erro</h1></div><p>Pra "{TIPOS_AFASTAMENTO[tipo]}" é preciso informar o horário de saída e retorno.</p><a href="/administrativo/afastamentos/novo" class="btn">Voltar</a>', active="administrativo-novo"))

    conteudo = await arquivo.read()
    ext = os.path.splitext(arquivo.filename or "")[1] or ".pdf"
    nome_no_drive = f"{prof['nome']} - {TIPOS_AFASTAMENTO[tipo]} - {data_inicio}{ext}"
    mime = arquivo.content_type or "application/octet-stream"

    drive_id, drive_link, erro_drive = _drive_upload_arquivo(nome_no_drive, conteudo, mime)
    status_upload = "enviado" if drive_id else "erro"

    conn = get_db()
    conn.execute("""
        INSERT INTO afastamentos (professor_id, tipo, data_inicio, data_fim, observacao, arquivo_nome, arquivo_drive_id, arquivo_drive_link, status_upload, horario_inicio, horario_fim)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (prof["id"], tipo, data_inicio, data_fim, observacao.strip() or None, arquivo.filename, drive_id, drive_link, status_upload, horario_inicio, horario_fim))
    conn.commit()
    conn.close()

    aviso_html = ""
    if erro_drive:
        aviso_html = f'<div class="tip" style="background:var(--orange-bg); border-color:var(--orange); margin-top:14px;"><strong>Atenção:</strong> a solicitação foi registrada, mas o documento NÃO foi enviado ao Drive ainda — {erro_drive} Avise o administrador; o arquivo original fica só com você por enquanto.</div>'

    horario_str = f' · Horário: {horario_inicio} às {horario_fim}' if horario_inicio and horario_fim else ""
    content = f"""
        <div class="page-header"><h1>✅ Justificativa enviada</h1></div>
        <p>Tipo: <strong>{TIPOS_AFASTAMENTO[tipo]}</strong> · Período: {data_inicio} a {data_fim}{horario_str}</p>
        {aviso_html}
        <div class="page-actions" style="margin-top:16px;">
            <a href="/administrativo/afastamentos" class="btn btn-primary">Ver minhas Justificativas</a>
            <a href="/administrativo/afastamentos/novo" class="btn">Nova Justificativa</a>
        </div>
    """
    return HTMLResponse(render_page("Justificativa enviada", content, active="administrativo-novo"))


@app.get("/administrativo/afastamentos", response_class=HTMLResponse)
def listar_meus_afastamentos(request: Request):
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)

    conn = get_db()
    registros = conn.execute("""
        SELECT * FROM afastamentos WHERE professor_id = ? ORDER BY data_inicio DESC
    """, (prof["id"],)).fetchall()
    conn.close()

    if not registros:
        linhas = '<tr><td colspan="6" style="padding:16px; text-align:center; color:var(--text-muted);">Nenhuma Justificativa ainda.</td></tr>'
    else:
        linhas = ""
        for r in registros:
            dias = (date.fromisoformat(r["data_fim"]) - date.fromisoformat(r["data_inicio"])).days + 1
            status_badge = (
                '<span style="color:var(--green);">✓ Anexado</span>' if r["status_upload"] == "enviado"
                else '<span style="color:var(--orange);">⚠ Pendente de envio</span>'
            )
            link_html = f' · <a href="{r["arquivo_drive_link"]}" target="_blank">Ver documento</a>' if r["arquivo_drive_link"] else ""
            horario_col = f'{r["horario_inicio"]} às {r["horario_fim"]}' if ("horario_inicio" in r.keys() and r["horario_inicio"]) else "—"
            linhas += f"""<tr>
                <td style="padding:8px;">{TIPOS_AFASTAMENTO.get(r["tipo"], r["tipo"])}</td>
                <td style="padding:8px;">{r["data_inicio"]} a {r["data_fim"]}</td>
                <td style="padding:8px; text-align:center;">{horario_col}</td>
                <td style="padding:8px; text-align:center;">{dias}</td>
                <td style="padding:8px;">{status_badge}{link_html}</td>
            </tr>"""

    content = f"""
        <div class="page-header">
            <h1>📄 Minhas Justificativas para o Ponto</h1>
        </div>
        <div class="page-actions" style="margin-bottom:14px;">
            <a href="/administrativo/afastamentos/novo" class="btn btn-primary">+ Nova Justificativa</a>
        </div>
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <thead><tr style="background:var(--bg-subtle);">
                <th style="padding:8px; text-align:left;">Tipo</th>
                <th style="padding:8px; text-align:left;">Período</th>
                <th style="padding:8px;">Horário</th>
                <th style="padding:8px;">Dias</th>
                <th style="padding:8px; text-align:left;">Documento</th>
            </tr></thead>
            <tbody>{linhas}</tbody>
        </table>
    """
    return HTMLResponse(render_page("Minhas Justificativas", content, active="administrativo-minhas"))


@app.get("/administrativo/relatorio", response_class=HTMLResponse)
def relatorio_afastamentos(request: Request, mes: Optional[int] = None, ano: Optional[int] = None):
    prof = get_current_professor(request)
    if not prof or not (prof.get("is_admin") or prof.get("is_gestor")):
        return RedirectResponse("/", status_code=303)

    hoje = date.today()
    mes = mes or hoje.month
    ano = ano or hoje.year

    conn = get_db()
    registros = conn.execute("""
        SELECT a.*, p.nome AS prof_nome, p.email AS prof_email
        FROM afastamentos a
        JOIN professores p ON p.id = a.professor_id
        WHERE strftime('%Y-%m', a.data_inicio) = ?
        ORDER BY p.nome, a.data_inicio
    """, (f"{ano:04d}-{mes:02d}",)).fetchall()
    conn.close()

    if not registros:
        linhas = '<tr><td colspan="7" style="padding:16px; text-align:center; color:var(--text-muted);">Nenhuma Justificativa registrada nesse mês.</td></tr>'
    else:
        linhas = ""
        for r in registros:
            dias = (date.fromisoformat(r["data_fim"]) - date.fromisoformat(r["data_inicio"])).days + 1
            matricula = _extrair_matricula(r["prof_email"])
            link_html = f'<a href="{r["arquivo_drive_link"]}" target="_blank">Ver</a>' if r["arquivo_drive_link"] else "—"
            horario_col = f'{r["horario_inicio"]} às {r["horario_fim"]}' if ("horario_inicio" in r.keys() and r["horario_inicio"]) else "—"
            linhas += f"""<tr>
                <td style="padding:8px;">{r["prof_nome"]}</td>
                <td style="padding:8px; text-align:center;">{matricula}</td>
                <td style="padding:8px;">{TIPOS_AFASTAMENTO.get(r["tipo"], r["tipo"])}</td>
                <td style="padding:8px;">{r["data_inicio"]} a {r["data_fim"]}</td>
                <td style="padding:8px; text-align:center;">{horario_col}</td>
                <td style="padding:8px; text-align:center;">{dias}</td>
                <td style="padding:8px;">{link_html}</td>
            </tr>"""

    meses_nomes = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    opts_mes = "".join(f'<option value="{i}"{" selected" if i==mes else ""}>{meses_nomes[i]}</option>' for i in range(1, 13))

    content = f"""
        <div class="page-header">
            <h1>📊 Relatório de Justificativas para o Ponto — {meses_nomes[mes]}/{ano}</h1>
        </div>
        <form method="get" style="display:flex; gap:12px; align-items:flex-end; margin-bottom:18px;">
            <label style="margin:0;">Mês <select name="mes">{opts_mes}</select></label>
            <label style="margin:0;">Ano <input type="number" name="ano" value="{ano}"></label>
            <button type="submit" class="btn">Filtrar</button>
            <a href="/administrativo/relatorio/exportar?mes={mes}&ano={ano}" class="btn">📥 Baixar Excel</a>
        </form>
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <thead><tr style="background:var(--bg-subtle);">
                <th style="padding:8px; text-align:left;">Nome</th>
                <th style="padding:8px;">Matrícula</th>
                <th style="padding:8px; text-align:left;">Tipo</th>
                <th style="padding:8px; text-align:left;">Período</th>
                <th style="padding:8px;">Horário</th>
                <th style="padding:8px;">Dias</th>
                <th style="padding:8px; text-align:left;">Documento</th>
            </tr></thead>
            <tbody>{linhas}</tbody>
        </table>
    """
    return HTMLResponse(render_page("Relatório de Justificativas", content, active="administrativo-relatorio"))


@app.get("/administrativo/relatorio/exportar")
def exportar_relatorio_afastamentos(request: Request, mes: Optional[int] = None, ano: Optional[int] = None):
    prof = get_current_professor(request)
    if not prof or not (prof.get("is_admin") or prof.get("is_gestor")):
        return RedirectResponse("/", status_code=303)

    hoje = date.today()
    mes = mes or hoje.month
    ano = ano or hoje.year

    conn = get_db()
    registros = conn.execute("""
        SELECT a.*, p.nome AS prof_nome, p.email AS prof_email
        FROM afastamentos a
        JOIN professores p ON p.id = a.professor_id
        WHERE strftime('%Y-%m', a.data_inicio) = ?
        ORDER BY p.nome, a.data_inicio
    """, (f"{ano:04d}-{mes:02d}",)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Justificativas"
    cabecalho = ["Nome", "Matrícula", "Tipo", "Data início", "Data fim", "Horário saída", "Horário retorno", "Dias", "Observação", "Link do documento"]
    ws.append(cabecalho)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in registros:
        dias = (date.fromisoformat(r["data_fim"]) - date.fromisoformat(r["data_inicio"])).days + 1
        matricula = _extrair_matricula(r["prof_email"])
        horario_inicio_r = r["horario_inicio"] if ("horario_inicio" in r.keys() and r["horario_inicio"]) else ""
        horario_fim_r = r["horario_fim"] if ("horario_fim" in r.keys() and r["horario_fim"]) else ""
        ws.append([
            r["prof_nome"], matricula, TIPOS_AFASTAMENTO.get(r["tipo"], r["tipo"]),
            r["data_inicio"], r["data_fim"], horario_inicio_r, horario_fim_r, dias,
            r["observacao"] or "", r["arquivo_drive_link"] or "",
        ])

    for i, largura in enumerate([28, 12, 22, 12, 12, 13, 14, 8, 30, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    meses_nomes_arq = ["", "janeiro", "fevereiro", "marco", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
    filename = f"justificativas_ponto_{meses_nomes_arq[mes]}_{ano}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/login", response_class=HTMLResponse)
def pagina_login(request: Request, next: str = "/", erro: str = ""):
    prof = get_current_professor(request)
    if prof:
        return RedirectResponse(next or "/", status_code=303)

    import html as _html
    erro_html = f'<div style="background:var(--red-bg); color:var(--red); border:1px solid var(--red); padding:12px; border-radius:6px; margin-bottom:16px;">{_html.escape(erro)}</div>' if erro else ""

    if DEV_MODE:
        botao_login = f"""
            <div style="background:var(--orange-bg); color:var(--orange); border:1px solid var(--orange); padding:12px; border-radius:6px; margin-bottom:16px; font-size:13px;">
                ⚙ <strong>Modo de desenvolvimento</strong> ativo (sem credenciais OAuth Google).
                Em produção, este botão será substituído por "Entrar com Google".
            </div>
            <form action="/auth/dev-login" method="post">
                <input type="hidden" name="next" value="{_html.escape(next, quote=True)}">
                <label>Email institucional<input type="email" name="email" required placeholder="seu.nome@{ALLOWED_EMAIL_DOMAIN}" autofocus></label>
                <label>Nome<input type="text" name="nome" required placeholder="Seu nome"></label>
                <button type="submit" class="btn btn-primary" style="width:100%; margin-top:10px;">Entrar (dev)</button>
            </form>
        """
    else:
        botao_login = f"""
            <a href="/auth/google?next={urllib.parse.quote(next)}" class="btn btn-primary" style="display:flex; align-items:center; justify-content:center; gap:10px; width:100%; padding:12px; font-size:15px;">
                <svg width="20" height="20" viewBox="0 0 24 24"><path fill="#4285F4" d="M22.56 12.25c0-.78-.07-1.53-.2-2.25H12v4.26h5.92c-.26 1.37-1.04 2.53-2.21 3.31v2.77h3.57c2.08-1.92 3.28-4.74 3.28-8.09z"/><path fill="#34A853" d="M12 23c2.97 0 5.46-.98 7.28-2.66l-3.57-2.77c-.98.66-2.23 1.06-3.71 1.06-2.86 0-5.29-1.93-6.16-4.53H2.18v2.84C3.99 20.53 7.7 23 12 23z"/><path fill="#FBBC05" d="M5.84 14.09c-.22-.66-.35-1.36-.35-2.09s.13-1.43.35-2.09V7.07H2.18C1.43 8.55 1 10.22 1 12s.43 3.45 1.18 4.93l2.85-2.22.81-.62z"/><path fill="#EA4335" d="M12 5.38c1.62 0 3.06.56 4.21 1.64l3.15-3.15C17.45 2.09 14.97 1 12 1 7.7 1 3.99 3.47 2.18 7.07l3.66 2.84c.87-2.6 3.3-4.53 6.16-4.53z"/></svg>
                Entrar com Google
            </a>
            <p class="muted-line" style="font-size:12px; text-align:center; margin-top:14px;">Use sua conta institucional <strong>@{ALLOWED_EMAIL_DOMAIN}</strong>. Acesso de outros domínios será recusado.</p>
        """

    content = f"""
    <div style="max-width:420px; margin:60px auto; padding:30px; background:var(--bg); border:1px solid var(--border); border-radius:8px;">
        <div style="text-align:center; margin-bottom:18px;">
            <img src="/static/imagens/logo_walmir.png" alt="E.M. Walmir de Freitas Monteiro" style="max-width:200px; height:auto; display:block; margin:0 auto;">
        </div>
        <h1 style="margin:0 0 6px 0; text-align:center; font-size:22px;">Sistema Pedagógico</h1>
        <p class="muted-line" style="margin:0 0 24px 0; text-align:center;">E.M. Walmir de Freitas Monteiro</p>
        {erro_html}
        {botao_login}
    </div>
    """
    return render_page("Entrar", content, active="", standalone=True)


@app.get("/auth/google")
def auth_google_redirect(next: str = "/"):
    if DEV_MODE:
        return RedirectResponse(f"/login?next={urllib.parse.quote(next)}", status_code=303)
    state = _session_serializer.dumps({"next": next})
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "response_type": "code",
        "scope": "openid email profile",
        "redirect_uri": f"{BASE_URL}/auth/google/callback",
        "state": state,
        "hd": ALLOWED_EMAIL_DOMAIN,
        "access_type": "online",
    }
    url = "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)
    return RedirectResponse(url, status_code=303)


@app.get("/auth/google/callback")
async def auth_google_callback(code: str = "", state: str = "", error: str = ""):
    if error:
        return RedirectResponse(f"/login?erro={urllib.parse.quote('Login cancelado: ' + error)}", status_code=303)
    if not code:
        return RedirectResponse("/login?erro=Código%20de%20autorização%20ausente", status_code=303)

    try:
        state_data = _session_serializer.loads(state, max_age=600)
        next_url = state_data.get("next", "/")
    except Exception:
        next_url = "/"

    import httpx
    async with httpx.AsyncClient(timeout=15.0) as client:
        token_resp = await client.post("https://oauth2.googleapis.com/token", data={
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "code": code,
            "grant_type": "authorization_code",
            "redirect_uri": f"{BASE_URL}/auth/google/callback",
        })
        if token_resp.status_code != 200:
            return RedirectResponse(f"/login?erro={urllib.parse.quote('Falha no token: ' + token_resp.text[:80])}", status_code=303)
        tokens = token_resp.json()
        userinfo_resp = await client.get(
            "https://www.googleapis.com/oauth2/v3/userinfo",
            headers={"Authorization": f"Bearer {tokens['access_token']}"}
        )
        if userinfo_resp.status_code != 200:
            return RedirectResponse("/login?erro=Falha%20ao%20obter%20userinfo", status_code=303)
        userinfo = userinfo_resp.json()

    email = (userinfo.get("email") or "").lower().strip()
    if not email:
        return RedirectResponse("/login?erro=Email%20ausente", status_code=303)
    if not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
        return RedirectResponse(
            f"/login?erro={urllib.parse.quote(f'Apenas contas @{ALLOWED_EMAIL_DOMAIN} são aceitas. Você entrou com {email}.')}",
            status_code=303
        )

    nome = userinfo.get("name") or email.split("@")[0]
    foto = userinfo.get("picture")
    prof = _upsert_professor(email, nome, foto)

    token = _criar_sessao(prof["id"], prof["email"])
    response = RedirectResponse(next_url or "/", status_code=303)
    response.set_cookie(
        key=SESSION_COOKIE, value=token,
        max_age=SESSION_MAX_AGE, httponly=True, samesite="lax",
        secure=BASE_URL.startswith("https://"),
    )
    return response


@app.post("/auth/dev-login")
def auth_dev_login(email: str = Form(...), nome: str = Form(...), next: str = Form("/")):
    if not DEV_MODE:
        return RedirectResponse("/login?erro=Modo%20dev%20desabilitado", status_code=303)
    email = email.lower().strip()
    if not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
        return RedirectResponse(
            f"/login?erro={urllib.parse.quote(f'Apenas contas @{ALLOWED_EMAIL_DOMAIN}')}",
            status_code=303
        )
    prof = _upsert_professor(email, nome.strip())
    token = _criar_sessao(prof["id"], prof["email"])
    response = RedirectResponse(next or "/", status_code=303)
    response.set_cookie(
        key=SESSION_COOKIE, value=token,
        max_age=SESSION_MAX_AGE, httponly=True, samesite="lax", secure=False,
    )
    return response


@app.post("/logout")
@app.get("/logout")
def logout():
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


MATHJAX = """
<script>
window.MathJax = {
  tex: { inlineMath: [['$', '$']], displayMath: [['$$', '$$']], processEscapes: true },
  svg: { fontCache: 'global' },
  options: {
    skipHtmlTags: ['script','noscript','style','textarea','pre','code']
  },
  startup: {
    ready: function() {
      MathJax.startup.defaultReady();
      // Após carregar, re-processa elementos visíveis (ignora display:none)
      var visíveis = Array.from(document.querySelectorAll('.questao-card-preview'))
                         .filter(function(el) { return el.offsetParent !== null; });
      if (visíveis.length) MathJax.typesetPromise(visíveis);
    }
  }
};
</script>
<script async src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js"></script>
"""

# Versão para páginas de EDIÇÃO: não auto-renderiza nos campos editáveis,
# mas processa os cards de questão da listagem (preview do banco)
MATHJAX_EDIT = """
<script>
window.MathJax = {
  tex: { inlineMath: [['$', '$']], displayMath: [['$$', '$$']], processEscapes: true },
  svg: { fontCache: 'global' },
  options: {
    skipHtmlTags: ['script','noscript','style','textarea','pre','code'],
    ignoreHtmlClass: 'ed-wrap'
  },
  startup: {
    typeset: false,
    ready: function() {
      MathJax.startup.defaultReady();
      // Renderiza cards de questão mas não os campos editáveis (.ed-wrap)
      var targets = document.querySelectorAll('.questao-card-preview');
      if (targets.length) MathJax.typesetPromise(Array.from(targets));
    }
  }
};
</script>
<script defer src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js"></script>
"""

INTER_FONT = '<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Sora:wght@400;500;600;700;800&display=swap" rel="stylesheet">'

def _css_version():
    """Hash curto do app.css pra cache-busting automático. Muda sempre que o CSS muda."""
    import hashlib
    try:
        with open(os.path.join("static", "css", "app.css"), "rb") as f:
            return hashlib.md5(f.read()).hexdigest()[:8]
    except Exception:
        return "0"

CSS_VERSION = _css_version()
CSS_LINK = f'<link rel="stylesheet" href="/static/css/app.css?v={CSS_VERSION}">'

# Tags de PWA (Progressive Web App) — permitem "Adicionar à tela inicial" no
# celular, abrindo em tela cheia (sem barra do navegador), com ícone próprio.
# Não faz o site funcionar offline de propósito (ver nota em static/sw.js) —
# só melhora a experiência de abrir o sistema no dia a dia.
PWA_HEAD_TAGS = """<link rel="manifest" href="/static/manifest.json">
    <meta name="theme-color" content="#4C6EF5">
    <link rel="apple-touch-icon" href="/static/icon-192.png">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="default">
    <meta name="apple-mobile-web-app-title" content="Ped. Walmir">
    <script>
    if ('serviceWorker' in navigator) {
      window.addEventListener('load', function() {
        navigator.serviceWorker.register('/static/sw.js').catch(function(){});
      });
    }
    </script>"""

# Script de tema (claro/escuro) — aplicado em todas as páginas via render_page.
# Lê preferência do localStorage e aplica antes do render pra evitar flash.
THEME_BOOT_SCRIPT = """<script>
(function(){
  try {
    var saved = localStorage.getItem('walmir-theme') || 'light';
    document.documentElement.setAttribute('data-theme', saved);
    var sb = localStorage.getItem('walmir-sidebar') || 'expanded';
    document.documentElement.setAttribute('data-sidebar', sb);
  } catch(e) {
    document.documentElement.setAttribute('data-theme', 'light');
    document.documentElement.setAttribute('data-sidebar', 'expanded');
  }
})();
function _walmirToggleTheme() {
  var html = document.documentElement;
  var cur = html.getAttribute('data-theme') || 'light';
  var next = cur === 'light' ? 'dark' : 'light';
  html.setAttribute('data-theme', next);
  try { localStorage.setItem('walmir-theme', next); } catch(e) {}
  document.querySelectorAll('[data-theme-toggle]').forEach(function(btn){
    btn.innerHTML = next === 'dark' ? '☀️ Tema claro' : '🌙 Tema escuro';
  });
}
function _walmirToggleSidebar() {
  var html = document.documentElement;
  var cur = html.getAttribute('data-sidebar') || 'expanded';
  var next = cur === 'expanded' ? 'collapsed' : 'expanded';
  html.setAttribute('data-sidebar', next);
  try { localStorage.setItem('walmir-sidebar', next); } catch(e) {}
}
function _walmirToggleMobileMenu() {
  var html = document.documentElement;
  var cur = html.getAttribute('data-mobile-menu') || 'closed';
  html.setAttribute('data-mobile-menu', cur === 'open' ? 'closed' : 'open');
}
</script>"""


def render_page(title: str, content: str, active: str = "", head_extra: str = "", standalone: bool = False, professor: Optional[dict] = None) -> str:
    """standalone=True omite sidebar (usado em /login).
    professor: dict do usuário logado pra exibir no rodapé do sidebar. Se None,
    o middleware pode ter colocado em request.state — mas como render_page não
    recebe request, em rotas internas o caller passa explicitamente, OU usamos
    o helper render_page_for(request, ...)."""
    def nav_class(name):
        return ' class="active"' if active == name else ''

    if standalone:
        return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <meta name="color-scheme" content="light dark">
    <title>{title} · Sistema Pedagógico do Walmir</title>
    {INTER_FONT}
    {THEME_BOOT_SCRIPT}
    {CSS_LINK}
    {PWA_HEAD_TAGS}
    {head_extra}
</head>
<body>
    {content}
</body>
</html>"""

    # Rodapé do sidebar com info do prof + toggle de tema
    if professor is None:
        professor = _current_prof_ctx.get()
    user_block = ""
    if professor:
        admin_badge = ' <span style="background:var(--purple); color:white; font-size:9px; padding:1px 5px; border-radius:3px; vertical-align:middle;">ADMIN</span>' if professor.get("is_admin") else ""
        user_block = f"""
            <div class="sidebar-user-footer" style="margin-top:auto; padding:12px; border-top:1px solid var(--border); font-size:12px;">
                <div class="sidebar-user-info">
                    <div style="font-weight:600;">{professor.get("nome", "")}{admin_badge}</div>
                    <div style="color:var(--text-muted); font-size:11px; margin-top:2px; word-break:break-all;">{professor.get("email", "")}</div>
                </div>
                <button data-theme-toggle data-icon="🌙" class="theme-toggle" onclick="_walmirToggleTheme()">🌙 Tema escuro</button>
                <a href="/logout" style="display:inline-block; margin-top:8px; font-size:11px; color:var(--text-muted);">Sair</a>
            </div>
        """

    # Sidebar dinâmico: itens de admin escondidos para professores comuns
    is_admin_view = bool(professor and professor.get("is_admin"))
    # Helper pra montar item da nav: ícone emoji + label (escondida quando collapsed) + data-name (tooltip)
    def nav_item(href, key, icon, label):
        return (
            f'<a href="{href}" data-name="{label}"{nav_class(key)}>'
            f'<span class="nav-icon">{icon}</span>'
            f'<span class="nav-label">{label}</span>'
            f'</a>'
        )

    link_disciplinas = nav_item("/disciplinas", "disciplinas", "📚", "Disciplinas") if is_admin_view else ''
    link_habilidades = nav_item("/habilidades", "habilidades", "🎯", "Habilidades BNCC") if is_admin_view else ''
    link_turmas = nav_item("/turmas", "turmas", "👥", "Turmas") if is_admin_view else ''

    # Apoio Educacional (Cuidadores/Agentes/Biblioteca/Apoio) só vê Início + Administrativo
    # no menu — nenhuma outra área (25/08/2026, a pedido).
    eh_apoio_restrito = bool(professor and professor.get("papel") == "apoio" and not professor.get("is_admin") and not professor.get("is_gestor"))

    if eh_apoio_restrito:
        nav_body = f"""
                {nav_item("/", "home", "🏠", "Início")}
                <div class="sidebar-section">Justificativas para o Ponto</div>
                {nav_item("/administrativo/afastamentos/novo", "administrativo-novo", "📄", "Nova Justificativa")}
                {nav_item("/administrativo/afastamentos", "administrativo-minhas", "📋", "Minhas Justificativas")}
        """
    else:
        # "Gerar Boletim" fica visível pra qualquer profissional (como já era), mas agora
        # agrupado dentro da seção "Boletim" — por isso o cabeçalho e os itens de
        # importação (admin/gestão) são montados separadamente (25/08/2026, reorganização).
        itens_boletim = []
        itens_boletim.append(nav_item("/boletim/boletim-individual", "boletim-individual", "🧾", "Gerar Boletim"))
        if professor and (professor.get("is_admin") or professor.get("is_gestor")):
            itens_boletim.append(nav_item("/boletim/importar-ecidade", "boletim-importar-ecidade", "📥", "Importar notas (e-cidade)"))
            itens_boletim.append(nav_item("/boletim/importar", "boletim-importar", "📥", "Importar planilha"))
        secao_boletim = ('<div class="sidebar-section">Boletim</div>' + "".join(itens_boletim)) if professor else ""

        # "Configurações" reúne o que era cadastro estrutural espalhado (Habilidades BNCC
        # e Turmas saíram de onde estavam) — admin apenas, mesma regra de antes (25/08/2026).
        secao_configuracoes = (
            '<div class="sidebar-section">Configurações</div>' + link_disciplinas + link_habilidades + link_turmas
        ) if is_admin_view else ""

        nav_body = f"""
                {nav_item("/", "home", "🏠", "Início")}
                <div class="sidebar-section">Banco de questões</div>
                {nav_item("/questoes", "questoes", "✏️", "Cadastrar questão")}
                <div class="sidebar-section">Tarefas</div>
                {nav_item("/provas", "provas", "📝", "Cadastrar atividade")}
                {nav_item("/aplicacoes", "aplicacoes", "📤", "Aplicar atividade")}
                {nav_item("/minhas-aplicacoes", "minhas-aplicacoes", "📋", "Minhas aplicações")}
                {nav_item("/painel-gestao", "painel-gestao", "🏛️", "Gestão de tarefas") if (professor and (professor.get("is_admin") or professor.get("is_gestor"))) else ""}
                {nav_item("/escanear", "escanear", "📷", "Digitalizar")}
                {nav_item("/simulados", "simulados", "📊", "Simulados")}
                <div class="sidebar-section">Análises pedagógicas</div>
                {nav_item("/boletim/analise", "boletim-analise", "📝", "Análise COC")}
                {(nav_item("/boletim/dashboard", "boletim-dashboard", "📊", "Dashboard Pedagógico") + nav_item("/boletim/comparativo", "boletim-comparativo", "🔄", "Comparativo") + nav_item("/boletim/estudantes", "boletim-estudantes", "👥", "Mapa de notas") + nav_item("/boletim/relatorio-geral", "boletim-relatorio-geral", "📄", "Relatório Geral") + nav_item("/boletim/relatorio-turma", "boletim-relatorio-turma", "📄", "Relatório por Turma")) if professor else ""}
                {('<div class="sidebar-section">Análise Simulado</div>' + nav_item("/analises-pedagogicas", "analises-pedagogicas", "📈", "Análise Pedagógica") + nav_item("/simulados/relatorio-notas", "simulados-relatorio-notas", "📄", "Relatório de Notas")) if professor else ""}
                {secao_boletim}
                {secao_configuracoes}
                {('<div class="sidebar-section">Justificativas para o Ponto</div>' + nav_item("/administrativo/afastamentos/novo", "administrativo-novo", "📄", "Nova Justificativa") + nav_item("/administrativo/afastamentos", "administrativo-minhas", "📋", "Minhas Justificativas") + (nav_item("/administrativo/relatorio", "administrativo-relatorio", "📊", "Relatório") if (professor.get("is_admin") or professor.get("is_gestor")) else "")) if professor else ""}
                {nav_item("/admin/usuarios", "admin-usuarios", "👥", "Usuários") if (professor and professor.get("is_admin")) else ""}
        """

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <meta name="color-scheme" content="light dark">
    <title>{title} · Sistema Pedagógico do Walmir</title>
    {INTER_FONT}
    {THEME_BOOT_SCRIPT}
    {CSS_LINK}
    {PWA_HEAD_TAGS}
    {head_extra}
</head>
<body>
    <div class="app">
        <div class="mobile-topbar">
            <button class="mobile-menu-btn" onclick="_walmirToggleMobileMenu()" type="button" aria-label="Abrir menu">☰</button>
            <span class="mobile-topbar-title">Sistema Pedagógico</span>
        </div>
        <div class="mobile-backdrop" onclick="document.documentElement.setAttribute('data-mobile-menu','closed')"></div>
        <aside class="sidebar" style="display:flex; flex-direction:column;">
            <button class="sidebar-toggle" onclick="_walmirToggleSidebar()" type="button" title="Recolher/expandir menu" aria-label="Recolher menu">
                <span class="sidebar-toggle-icon">≡</span>
            </button>
            <div class="sidebar-brand" style="text-align:center; padding:8px 6px 4px;">
                <img src="/static/imagens/logo_walmir.png" class="sidebar-logo-full" alt="Walmir" style="max-width:100%; height:auto; max-height:80px; display:block; margin:0 auto;">
                <div class="sidebar-logo-mini" aria-hidden="true">W</div>
                <div class="sidebar-brand-text" style="font-size:11px; color:var(--text-muted); margin-top:6px; font-weight:600; letter-spacing:0.3px;">Sistema Pedagógico</div>
            </div>
            <nav>
                {nav_body}
            </nav>
            {user_block}
        </aside>
        <main class="main">
            {content}
        </main>
    </div>
</body>
</html>"""


def gerar_codigo_aluno(conn):
    for _ in range(10):
        codigo = uuid.uuid4().hex[:8].upper()
        if not conn.execute("SELECT id FROM alunos WHERE codigo_unico = ?", (codigo,)).fetchone():
            return codigo
    raise RuntimeError("Não foi possível gerar código único após 10 tentativas")

def qr_data_uri(texto):
    """Gera um QR Code e retorna como data URI base64 (pra embutir em <img src>)."""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
    )
    qr.add_data(texto)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    return f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode()}"

def get_base_url(request):
    """Constrói a URL base correta, considerando proxy reverso (Codespaces, produção, etc.)."""
    host = request.headers.get("x-forwarded-host") or request.headers.get("host", "localhost")
    proto = request.headers.get("x-forwarded-proto", "http")
    return f"{proto}://{host}"

def format_data_br(iso_str):
    if not iso_str:
        return ""
    try:
        return datetime.fromisoformat(iso_str).strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return iso_str


def _preview_enunciado(enunciado: str, max_chars: int = 160) -> str:
    """Gera texto limpo para preview: remove tags HTML e simplifica notação MathJax."""
    import re as _re, html as _html
    # Remove tabelas inteiras (substituir por [tabela])
    enunciado = _re.sub(r'<table[^>]*>.*?</table>', '[tabela]', enunciado, flags=_re.DOTALL | _re.IGNORECASE)
    # Simplifica fórmulas: remove delimitadores $ mas mantém o conteúdo LaTeX legível
    # ex: $\frac{1}{2}$ → [1/2]   $x^{2}$ → [x^2]
    def _simplificar_formula(m):
        inner = m.group(1).strip()
        # Frações: \frac{a}{b} → (a/b)
        inner = _re.sub(r'\\frac\{([^}]+)\}\{([^}]+)\}', r'(\1/\2)', inner)
        # Potências: a^{b} → a^b
        inner = _re.sub(r'\^\{([^}]+)\}', r'^\1', inner)
        # Remove outras barras LaTeX
        inner = _re.sub(r'\\[a-zA-Z]+', '', inner)
        inner = inner.strip()
        return f'[{inner}]' if inner else '[fórmula]'
    texto = _re.sub(r'\$\$(.+?)\$\$', _simplificar_formula, enunciado, flags=_re.DOTALL)
    texto = _re.sub(r'\$([^$\n]+)\$', _simplificar_formula, texto)
    texto = _re.sub(r'\\\((.+?)\\\)', _simplificar_formula, texto, flags=_re.DOTALL)
    # Remove tags HTML
    texto = _re.sub(r'<[^>]+>', ' ', texto)
    # Decodifica entidades HTML
    texto = _html.unescape(texto)
    # Normaliza espaços
    texto = ' '.join(texto.split())
    return _html.escape(texto[:max_chars]) + ("..." if len(texto) > max_chars else "")


def render_questao_card(conn, q, numero=None, mostrar_acoes=False, compact=False, pode_editar=True, autor_nome=None):
    """
    pode_editar: se False, esconde botões Editar/Excluir mesmo com mostrar_acoes=True.
    autor_nome: se passado, exibe badge 'Por: <nome>' (usado quando admin lista questões alheias).
    """
    textos = conn.execute("SELECT conteudo, fonte FROM textos_apoio WHERE questao_id = ? ORDER BY ordem", (q["id"],)).fetchall()
    imagens = conn.execute("SELECT caminho, legenda, fonte FROM imagens WHERE questao_id = ? ORDER BY ordem", (q["id"],)).fetchall()
    alts = conn.execute("SELECT letra, texto, correta FROM alternativas WHERE questao_id = ? ORDER BY letra", (q["id"],)).fetchall()
    habilidades = conn.execute("SELECT h.codigo FROM questao_habilidades qh JOIN habilidades_bncc h ON h.id = qh.habilidade_id WHERE qh.questao_id = ? ORDER BY h.codigo", (q["id"],)).fetchall()
    ano_q = q["ano"] if "ano" in q.keys() and q["ano"] else None

    textos_html = ""
    for t in textos:
        fonte_html = f'<footer>Fonte: {t["fonte"]}</footer>' if t["fonte"] else ""
        textos_html += f'<blockquote>{t["conteudo"]}{fonte_html}</blockquote>'

    imagens_html = ""
    for img in imagens:
        legenda_html = f'<figcaption>{img["legenda"]}</figcaption>' if img["legenda"] else ""
        fonte_html = f'<figcaption><small>Fonte: {img["fonte"]}</small></figcaption>' if img["fonte"] else ""
        imagens_html += f'<figure><img src="/{img["caminho"]}" alt="">{legenda_html}{fonte_html}</figure>'

    tipo_q = q["tipo"] if "tipo" in q.keys() and q["tipo"] else "multipla_escolha"
    tipo_info = TIPOS_QUESTAO.get(tipo_q, TIPOS_QUESTAO["multipla_escolha"])

    alts_html = ""
    if tipo_q == "multipla_escolha":
        for a in alts:
            cls = ' class="correct"' if a["correta"] else ''
            marca = ' ✓' if a["correta"] else ''
            alts_html += f'<li{cls}><strong>{a["letra"]})</strong> {a["texto"]}{marca}</li>'
    elif tipo_q == "discursiva":
        alts_html = '<li style="list-style:none; padding:8px 12px; background:var(--bg-subtle); border-left:3px solid var(--accent); color:var(--text-muted); font-style:italic;">📝 Questão discursiva — resposta livre (correção manual)</li>'
    elif tipo_q == "vf":
        afirmacoes = conn.execute("SELECT ordem, texto, gabarito FROM vf_afirmacoes WHERE questao_id = ? ORDER BY ordem", (q["id"],)).fetchall()
        items = ""
        for af in afirmacoes:
            cor = "var(--green)" if af["gabarito"] == "V" else "var(--red)"
            items += (
                f'<li style="list-style:none; padding:6px 10px; background:var(--bg-subtle); margin-bottom:4px; border-radius:4px; border-left:3px solid {cor};">'
                f'<strong style="color:{cor};">({af["gabarito"]})</strong> {af["texto"]}'
                f'</li>'
            )
        alts_html = items or '<li style="list-style:none; padding:8px; color:var(--text-muted); font-style:italic;">(sem afirmações cadastradas)</li>'
    elif tipo_q == "associacao":
        itens_a = conn.execute("SELECT ordem, texto, gabarito_letra FROM assoc_itens_a WHERE questao_id = ? ORDER BY ordem", (q["id"],)).fetchall()
        itens_b = conn.execute("SELECT letra, texto FROM assoc_itens_b WHERE questao_id = ? ORDER BY letra", (q["id"],)).fetchall()
        ca_html = "".join(
            f'<li style="margin-bottom:4px;"><strong>{a["ordem"]+1}.</strong> {a["texto"]} '
            f'<span style="font-size:11px; color:var(--green);">→ resposta: ({a["gabarito_letra"]})</span></li>'
            for a in itens_a
        )
        cb_html = "".join(
            f'<li style="margin-bottom:4px;"><strong>({b["letra"]})</strong> {b["texto"]}</li>'
            for b in itens_b
        )
        alts_html = (
            f'<li style="list-style:none; padding:8px; background:var(--bg-subtle); border-radius:4px;">'
            f'<div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:14px;">'
            f'<div><strong style="font-size:12px; text-transform:uppercase; color:var(--text-muted);">Coluna A</strong><ul style="margin:6px 0 0 18px; padding:0;">{ca_html}</ul></div>'
            f'<div><strong style="font-size:12px; text-transform:uppercase; color:var(--text-muted);">Coluna B</strong><ul style="margin:6px 0 0 18px; padding:0;">{cb_html}</ul></div>'
            f'</div></li>'
        )

    habilidades_html = ""
    if habilidades:
        badges = "".join(f'<span class="badge">{h["codigo"]}</span>' for h in habilidades)
        habilidades_html = f'<div class="habilidades-row">{badges}</div>'

    # Badge de tipo (sempre visível)
    cores_tipo = {
        "multipla_escolha": ("var(--accent-bg)", "var(--accent)"),
        "discursiva":       ("var(--orange-bg)", "var(--orange)"),
        "vf":               ("var(--green-bg)", "var(--green)"),
        "associacao":       ("var(--purple-bg)", "var(--purple)"),
    }
    cor_tipo_bg, cor_tipo_fg = cores_tipo.get(tipo_q, cores_tipo["multipla_escolha"])
    tipo_badge = f' · <span class="badge" style="background:{cor_tipo_bg}; color:{cor_tipo_fg}; font-size:10px;">{tipo_info["icone"]} {tipo_info["label"]}</span>'

    cabecalho = f'Questão {numero} · {q["disciplina_nome"]}' if numero else q["disciplina_nome"]
    ano_badge = f' · <span style="color:var(--text-muted); font-weight:400;">{ano_q}</span>' if ano_q else ""
    autor_badge_inline = f' · <span class="badge" style="background:var(--purple-bg); color:var(--purple); font-size:10px;">Por: {autor_nome}</span>' if autor_nome else ""
    anulada = bool(q["anulada"]) if "anulada" in q.keys() else False
    anulada_badge = ' · <span class="badge" style="background:var(--red-bg); color:var(--red); font-size:10px;">🚫 ANULADA — todas as alternativas contam como corretas</span>' if anulada else ""

    acoes_html = ""
    if mostrar_acoes and pode_editar:
        botao_anular = ""
        if tipo_q == "multipla_escolha":
            if anulada:
                botao_anular = (
                    f'<form action="/questoes/{q["id"]}/anular" method="post" style="margin:0;" '
                    f'onsubmit="return confirm(\'Reativar o gabarito original desta questão?\');">'
                    f'<button type="submit" class="btn">✅ Reativar gabarito</button>'
                    f'</form>'
                )
            else:
                botao_anular = (
                    f'<form action="/questoes/{q["id"]}/anular" method="post" style="margin:0;" '
                    f'onsubmit="return confirm(\'Anular esta questão? Todas as alternativas passam a contar como corretas para todos os alunos, em todas as provas/simulados que já usam essa questão.\');">'
                    f'<button type="submit" class="btn" style="background:var(--orange); color:white; border-color:var(--orange);">🚫 Anular questão</button>'
                    f'</form>'
                )
        acoes_html = (
            f'<div class="page-actions" style="margin-top:16px; padding-top:12px; border-top:1px solid var(--border);">'
            f'<a href="/questoes/{q["id"]}/editar" class="btn">Editar</a>'
            f'{botao_anular}'
            f'<form action="/questoes/{q["id"]}/deletar" method="post" style="margin:0;" '
            f'onsubmit="return confirm(\'Excluir esta questão? Se ela for usada em alguma prova, a exclusão será bloqueada.\');">'
            f'<button type="submit" class="btn" style="background:var(--red); color:white; border-color:var(--red);">Excluir</button>'
            f'</form>'
            f'</div>'
        )
    elif mostrar_acoes and not pode_editar:
        acoes_html = (
            f'<div style="margin-top:12px; padding-top:10px; border-top:1px solid var(--border); font-size:11px; color:var(--text-muted);">'
            f'🔒 Questão de outro professor — você pode usá-la em suas provas, mas só o autor ou o administrador podem editá-la.'
            f'</div>'
        )

    if compact:
        # IMPORTANTE: strip de tags HTML antes do slice. Cortar HTML em 160 chars pode
        # deixar uma tag aberta (ex: <p style="..."> sem </p>), quebrando o layout.
        preview = _preview_enunciado(q["enunciado"], max_chars=160)
        habs_inline = ""
        if habilidades:
            habs_inline = " " + "".join(f'<span class="badge" style="font-size:10px;">{h["codigo"]}</span>' for h in habilidades)
        return (
            f'<div class="question questao-card-preview" style="margin-bottom:8px; padding:12px 16px;">'
            f'<div style="display:flex; justify-content:space-between; align-items:flex-start; gap:12px;">'
            f'<div style="flex:1; min-width:0;">'
            f'<div class="question-header" style="margin:0;">Q{q["id"]} · {q["disciplina_nome"]}{ano_badge}{tipo_badge}{autor_badge_inline}{anulada_badge}{habs_inline}</div>'
            f'<div style="margin-top:6px; color:var(--text); font-size:14px; line-height:1.5;">{preview}</div>'
            f'</div>'
            f'<button type="button" onclick="toggleQuestao({q["id"]})" id="q-toggle-{q["id"]}" '
            f'style="background:none; border:1px solid var(--border); border-radius:4px; padding:4px 10px; color:var(--text-muted); cursor:pointer; font-size:11px; white-space:nowrap; font-family:inherit;">'
            f'Ver completa ▾</button>'
            f'</div>'
            f'<div id="q-detalhes-{q["id"]}" style="display:none; margin-top:14px; padding-top:14px; border-top:1px solid var(--border);">'
            f'{textos_html}{imagens_html}'
            f'<div class="enunciado">{q["enunciado"]}</div>'
            f'<ul class="alternativas">{alts_html}</ul>'
            f'{habilidades_html}{acoes_html}'
            f'</div>'
            f'</div>'
        )

    return f'<div class="question questao-card-preview"><div class="question-header">{cabecalho}{ano_badge}{tipo_badge}{anulada_badge}</div>{textos_html}{imagens_html}<div class="enunciado">{q["enunciado"]}</div><ul class="alternativas">{alts_html}</ul>{habilidades_html}{acoes_html}</div>'


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    prof = get_current_professor(request)
    nome_prof = prof["nome"].split()[0] if prof else "professor(a)"
    prof_id = prof["id"] if prof else 0

    conn = get_db()

    # === APLICAÇÕES ABERTAS (do prof; pra ADMIN, é DA ESCOLA inteira) ===
    is_admin = bool(prof and prof["is_admin"])
    if is_admin:
        # Aplicações ABERTAS da ESCOLA (qualquer prof) — tela inicial mostra só essas (24/08/2026)
        minhas_ultimas = conn.execute("""
            SELECT a.id, a.modo, a.aberta,
                   COALESCE(a.titulo, p.titulo) AS titulo,
                   t.nome AS turma_nome, t.ano_letivo,
                   prof.nome AS criador_nome,
                   (SELECT COUNT(*) FROM entregas e WHERE e.aplicacao_id = a.id) AS n_entregas,
                   (SELECT COUNT(*) FROM alunos al WHERE al.turma_id = a.turma_id) AS n_alunos
            FROM aplicacoes a
            JOIN provas p ON p.id = a.prova_id
            JOIN turmas t ON t.id = a.turma_id
            LEFT JOIN professores prof ON prof.id = a.criada_por_professor_id
            WHERE a.aberta = 1
            ORDER BY a.id DESC LIMIT 10
        """).fetchall()
    else:
        # Só as aplicações ABERTAS criadas por esse professor (24/08/2026)
        minhas_ultimas = conn.execute("""
            SELECT a.id, a.modo, a.aberta,
                   COALESCE(a.titulo, p.titulo) AS titulo,
                   t.nome AS turma_nome, t.ano_letivo,
                   NULL AS criador_nome,
                   (SELECT COUNT(*) FROM entregas e WHERE e.aplicacao_id = a.id) AS n_entregas,
                   (SELECT COUNT(*) FROM alunos al WHERE al.turma_id = a.turma_id) AS n_alunos
            FROM aplicacoes a
            JOIN provas p ON p.id = a.prova_id
            JOIN turmas t ON t.id = a.turma_id
            WHERE a.criada_por_professor_id = ? AND a.aberta = 1
            ORDER BY a.id DESC LIMIT 10
        """, (prof_id,)).fetchall()

    # Card "Alunos que precisam de atenção" — precisa do prof completo (papel/id), não só
    # is_admin, então busca antes de fechar a conexão (24/08/2026).
    atencao_dados = _home_alunos_atencao(conn, prof) if prof else None
    conn.close()

    # ----- HTML -----
    # Grid de atalhos tipo "badge" pra tela inicial, agrupado em seções nomeadas (com uma
    # cor de destaque por grupo) em vez de um grid único e monótono — 24/08/2026, revisão
    # depois de feedback de que a versão anterior ficou genérica/sem hierarquia.
    is_gestor = bool(prof and prof.get("is_gestor"))
    eh_apoio_restrito = bool(prof and prof.get("papel") == "apoio" and not is_admin and not is_gestor)

    grupo_administrativo = ("Justificativas para o Ponto", "#0891b2", [
        ("📄", "Nova Justificativa", "/administrativo/afastamentos/novo", True),
        ("📋", "Minhas Justificativas", "/administrativo/afastamentos", True),
        ("📊", "Relatório", "/administrativo/relatorio", is_admin or is_gestor),
    ])

    if eh_apoio_restrito:
        # Apoio Educacional só acessa a solicitação de afastamento — a tela inicial
        # mostra só isso, sem os atalhos de áreas que ele não pode abrir (25/08/2026).
        grupos_config = [grupo_administrativo]
    else:
        grupos_config = [
            ("Ação rápida", "#2563eb", [
                ("✏️", "Nova questão", "/questoes/nova", True),
                ("📝", "Atividades", "/provas", True),
                ("📤", "Aplicar", "/aplicacoes/nova", True),
                ("📋", "Minhas aplic.", "/minhas-aplicacoes", True),
                ("📷", "Digitalizar", "/escanear", True),
                ("📊", "Simulados", "/simulados", True),
                ("📈", "Análises", "/analises-pedagogicas", True),
                ("📊", "Dashboard Pedagógico", "/boletim/dashboard", True),
            ]),
            ("Banco", "#16a34a", [
                ("📚", "Disciplinas", "/disciplinas", is_admin),
                ("🎯", "Habilidades", "/habilidades", is_admin),
                ("👥", "Turmas", "/turmas", is_admin),
            ]),
            grupo_administrativo,
            ("Gestão", "#7c3aed", [
                ("🏛️", "Painel gestão", "/painel-gestao", is_admin or is_gestor),
                ("👤", "Usuários", "/admin/usuarios", is_admin),
            ]),
        ]
    grupos_html = ""
    TILE_PX = 140   # largura-alvo de cada bloco, igual em todas as fileiras
    GAP_PX = 12
    for titulo_grupo, cor_grupo, itens in grupos_config:
        itens_visiveis = [it for it in itens if it[3]]
        if not itens_visiveis:
            continue
        badges_grupo = "".join(
            f'<a href="{href}" class="mobile-badge" style="--badge-accent:{cor_grupo};"><span class="mobile-badge-icon">{icone}</span><span class="mobile-badge-label">{label}</span></a>'
            for icone, label, href, _ in itens_visiveis
        )
        # Largura da fileira = só o que ela precisa pros seus próprios itens (até um teto
        # de 6 por linha) — evita tanto "buraco vazio" quanto blocos gigantes/desproporcionais
        # numa fileira com poucos itens (24/08/2026, ajuste depois de feedback visual).
        n_cols_linha = min(len(itens_visiveis), 6)
        largura_max = n_cols_linha * TILE_PX + (n_cols_linha - 1) * GAP_PX
        grupos_html += f"""
        <div class="mobile-launcher-group">
            <div class="mobile-launcher-group-title" style="color:{cor_grupo};">{titulo_grupo}</div>
            <div class="mobile-launcher-grid" style="max-width:{largura_max}px;">{badges_grupo}</div>
        </div>"""

    mobile_launcher_html = f"""
        <div class="mobile-launcher">
            {grupos_html}
        </div>
        <style>
        .mobile-launcher {{
            display: block;
            margin: 4px 0 28px 0;
            max-width: 900px;
        }}
        .mobile-launcher-group {{ margin-bottom: 18px; }}
        .mobile-launcher-group-title {{
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.6px;
            margin-bottom: 8px;
        }}
        .mobile-launcher-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
            gap: 12px;
        }}
        .mobile-badge {{
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            gap: 8px;
            padding: 18px 10px;
            background: var(--bg-subtle);
            border: 1px solid var(--border);
            border-top: 3px solid var(--badge-accent, var(--border));
            border-radius: 12px;
            text-decoration: none;
            color: inherit;
            text-align: center;
            transition: transform 0.12s ease, box-shadow 0.12s ease;
        }}
        .mobile-badge:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
        }}
        .mobile-badge-icon {{ font-size: 26px; line-height: 1; }}
        .mobile-badge-label {{ font-size: 12px; font-weight: 600; line-height: 1.25; }}
        @media (min-width: 640px) {{
            .mobile-badge {{ padding: 20px 10px; }}
            .mobile-badge-icon {{ font-size: 30px; }}
            .mobile-badge-label {{ font-size: 13px; }}
        }}
        </style>
    """

    # Aplicações abertas do professor — única lista mostrada na tela inicial, além do
    # grid de atalhos (24/08/2026: home simplificada a pedido, tirando Acervo/Painel/stats).
    # Grid de 2 colunas em telas largas — antes era 1 coluna comprida, deixando muito
    # vazio ao lado (24/08/2026, ajuste de distribuição).
    if minhas_ultimas:
        linhas = ""
        for u in minhas_ultimas:
            modo_label = "online" if u["modo"] == "online" else "impressa"
            pct = (u["n_entregas"] / u["n_alunos"] * 100) if u["n_alunos"] > 0 else 0
            autor_inline = f' · <span style="color:var(--purple);">por {u["criador_nome"] or "—"}</span>' if is_admin else ""
            linhas += (
                f'<a href="/aplicacoes/{u["id"]}" style="display:flex; flex-direction:column; gap:6px; padding:12px 14px; border:1px solid var(--border); border-radius:10px; text-decoration:none; color:inherit; background:var(--bg-subtle);">'
                f'<div style="min-width:0;"><span style="color:var(--green);">●</span> <strong>{u["titulo"]}</strong></div>'
                f'<div style="font-size:12px; color:var(--text-muted);">{u["turma_nome"]} · {modo_label}{autor_inline}</div>'
                f'<div style="font-size:12px; color:var(--text-muted);">{u["n_entregas"]}/{u["n_alunos"]} entregas ({pct:.0f}%)</div>'
                f'</a>'
            )
        label_abertas_titulo = "📤 Aplicações abertas na escola" if is_admin else "📤 Suas aplicações abertas"
        aplicacoes_abertas_html = f"""
            <h2 style="margin-top:28px; font-size:15px; text-transform:uppercase; letter-spacing:1px; color:var(--text-muted);">{label_abertas_titulo}</h2>
            <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(280px, 1fr)); gap:10px; max-width:900px;">
                {linhas}
            </div>
        """
    else:
        label_vazio = "Nenhuma aplicação aberta na escola no momento." if is_admin else "Você não tem nenhuma aplicação aberta no momento."
        aplicacoes_abertas_html = f"""
            <h2 style="margin-top:28px; font-size:15px; text-transform:uppercase; letter-spacing:1px; color:var(--text-muted);">📤 Aplicações abertas</h2>
            <p style="color:var(--text-muted); font-size:13px;">{label_vazio}</p>
        """

    # Card "Alunos que precisam de atenção" — reaproveita a mesma regra de risco de
    # repetência do dashboard (qualquer disciplina com média T1+T2 < 5,0) + maiores
    # faltas. Vai na coluna direita da tela inicial, no espaço que ficava vazio (24/08/2026).
    if not atencao_dados:
        atencao_html = ""
    else:
        risco_html = ""
        for al in atencao_dados["risco"]:
            discs_txt = ", ".join(f'{d["nome"]} ({d["media"]:.1f})' for d in al["disciplinas"])
            risco_html += (
                f'<div style="padding:8px 0; border-bottom:1px solid var(--border);">'
                f'<strong style="font-size:13px;">{al["nome"]}</strong> <span style="font-size:11px; color:var(--text-muted);">· {al["turma_nome"]}</span>'
                f'<div style="font-size:11px; color:var(--red);">{discs_txt}</div>'
                f'</div>'
            )
        if not risco_html:
            risco_html = '<p style="font-size:12px; color:var(--text-muted);">Nenhum aluno em risco no momento. 🎉</p>'

        faltas_html = ""
        for al in atencao_dados["faltas"]:
            faltas_html += (
                f'<div style="display:flex; justify-content:space-between; padding:8px 0; border-bottom:1px solid var(--border); font-size:13px;">'
                f'<span><strong>{al["nome"]}</strong> <span style="font-size:11px; color:var(--text-muted);">· {al["turma_nome"]}</span></span>'
                f'<span style="color:var(--orange); font-weight:600;">{al["total"]} faltas</span>'
                f'</div>'
            )
        if not faltas_html:
            faltas_html = '<p style="font-size:12px; color:var(--text-muted);">Sem dados de falta ainda.</p>'

        atencao_html = f"""
        <div class="card" style="padding:16px;">
            <h2 style="margin:0 0 4px 0; font-size:14px; text-transform:uppercase; letter-spacing:1px; color:var(--text-muted);">⚠️ Alunos que precisam de atenção</h2>
            <p style="font-size:11px; color:var(--text-muted); margin:0 0 12px 0;">Ano {atencao_dados["ano"]} · Risco de repetência = média (1º+2º tri.) abaixo de 5,0</p>
            <h3 style="font-size:12px; margin:10px 0 4px 0; color:var(--red);">Risco de repetência</h3>
            {risco_html}
            <h3 style="font-size:12px; margin:14px 0 4px 0; color:var(--orange);">Mais faltas</h3>
            {faltas_html}
        </div>
        """

    content = f"""
        <div class="page-header">
            <h1 style="margin-bottom:4px;">Olá, {nome_prof} 👋</h1>
            <p class="subtitle" style="margin-top:0;">Veja seu panorama atualizado.</p>
        </div>
        <div class="home-layout">
            <div class="home-layout-main">
                {mobile_launcher_html}
                {aplicacoes_abertas_html}
            </div>
            <div class="home-layout-side">
                {atencao_html}
            </div>
        </div>
        <style>
        .home-layout {{ display: flex; gap: 24px; align-items: flex-start; flex-wrap: wrap; }}
        .home-layout-main {{ flex: 1 1 640px; min-width: 0; }}
        .home-layout-side {{ flex: 0 1 320px; min-width: 280px; }}
        </style>
    """
    return render_page("Início", content, active="home")


@app.get("/disciplinas", response_class=HTMLResponse)
def listar_disciplinas(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()
    disciplinas = conn.execute("SELECT * FROM disciplinas ORDER BY nome").fetchall()
    conn.close()
    if disciplinas:
        linhas = "".join(f"<li>{d['nome']}</li>" for d in disciplinas)
        lista_html = f'<ul class="clean">{linhas}</ul>'
    else:
        lista_html = '<div class="empty">Nenhuma disciplina cadastrada ainda.</div>'
    content = f'<div class="page-header"><h1>Disciplinas</h1><div class="page-actions"><a href="/disciplinas/nova" class="btn btn-primary">+ Nova disciplina</a></div></div>{lista_html}'
    return render_page("Disciplinas", content, active="disciplinas")


@app.get("/disciplinas/nova", response_class=HTMLResponse)
def form_nova_disciplina():
    content = '<div class="page-header"><h1>Nova disciplina</h1></div><form action="/disciplinas/nova" method="post"><label>Nome<input type="text" name="nome" required autofocus></label><div class="page-actions"><button type="submit" class="btn btn-primary">Cadastrar</button><a href="/disciplinas" class="btn">Cancelar</a></div></form>'
    return render_page("Nova disciplina", content, active="disciplinas")


@app.post("/disciplinas/nova")
def criar_disciplina(nome: str = Form(...)):
    conn = get_db()
    try:
        conn.execute("INSERT INTO disciplinas (nome) VALUES (?)", (nome.strip(),))
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    finally:
        conn.close()
    return RedirectResponse("/disciplinas", status_code=303)


# ═══════════════════════════════════════════════════════════════
# IMPORTADOR DE BANCO DE QUESTÕES (JSON em batch)
# ═══════════════════════════════════════════════════════════════

@app.get("/admin/importar-questoes", response_class=HTMLResponse)
def form_importar_questoes(request: Request):
    """Tela admin: upload de JSON com questões pra importar em batch."""
    _r = _require_admin_or_403(request)
    if _r is not None: return _r

    content = """
        <div class="page-header">
            <h1>📥 Importar banco de questões</h1>
            <p class="subtitle">Carregue um arquivo <code>.json</code> com questões estruturadas. Útil para popular o banco com provas oficiais (OBMEP, SAEB, OBA, etc.).</p>
        </div>

        <div class="card" style="background:var(--accent-bg); border-left:3px solid var(--accent);">
            <h3 style="margin-top:0;">📋 Formato esperado</h3>
            <p style="font-size:13px;">O arquivo deve conter um objeto JSON com a chave <code>questoes</code> contendo uma lista. Cada questão precisa de:</p>
            <ul style="font-size:13px; margin:8px 0 0 20px;">
                <li><code>disciplina</code> — nome (será criada se não existir)</li>
                <li><code>ano</code> — ex: <code>"6º ano"</code>, <code>"7º ano"</code>, etc.</li>
                <li><code>tipo</code> — por enquanto só <code>"multipla_escolha"</code></li>
                <li><code>enunciado</code> — texto da questão (HTML permitido)</li>
                <li><code>fonte</code> (opcional) — ex: <code>"OBMEP 2019, Nível 1, Q5"</code></li>
                <li><code>habilidade_bncc</code> (opcional) — código (ex: <code>"EF06MA15"</code>)</li>
                <li><code>alternativas</code> — lista com <code>{letra, texto, correta}</code></li>
            </ul>
            <p style="font-size:12px; color:var(--text-muted); margin-top:10px;">
                Veja o arquivo de exemplo: <code>banco_inicial_obmep.json</code> distribuído junto com o sistema.
            </p>
        </div>

        <form method="post" action="/admin/importar-questoes/preview" enctype="multipart/form-data" style="margin-top:18px;">
            <label>
                Arquivo JSON
                <input type="file" name="arquivo" accept=".json,application/json" required>
            </label>
            <div style="margin-top:14px;">
                <button type="submit" class="btn btn-primary">Visualizar preview →</button>
                <a href="/" class="btn">Cancelar</a>
            </div>
        </form>
    """
    return HTMLResponse(render_page("Importar questões", content, active=""))


@app.post("/admin/importar-questoes/preview", response_class=HTMLResponse)
async def preview_importar_questoes(request: Request):
    """Recebe JSON, valida estrutura e mostra preview antes de confirmar."""
    _r = _require_admin_or_403(request)
    if _r is not None: return _r

    form = await request.form()
    arquivo = form.get("arquivo")
    if not arquivo:
        return HTMLResponse(render_page("Erro", '<div class="card" style="background:var(--red-bg); color:var(--red);">Nenhum arquivo enviado.</div><a href="/admin/importar-questoes" class="btn">← Voltar</a>'))

    import json as _json
    try:
        conteudo = await arquivo.read()
        dados = _json.loads(conteudo.decode("utf-8"))
    except Exception as e:
        return HTMLResponse(render_page("Erro", f'<div class="card" style="background:var(--red-bg); color:var(--red);">Erro ao ler JSON: {html.escape(str(e))}</div><a href="/admin/importar-questoes" class="btn">← Voltar</a>'))

    questoes_raw = dados.get("questoes", [])
    if not isinstance(questoes_raw, list) or not questoes_raw:
        return HTMLResponse(render_page("Erro", '<div class="card" style="background:var(--red-bg); color:var(--red);">JSON inválido: chave <code>questoes</code> ausente ou vazia.</div><a href="/admin/importar-questoes" class="btn">← Voltar</a>'))

    # Validação de cada questão
    erros = []
    questoes_validas = []
    for i, q in enumerate(questoes_raw, start=1):
        problemas = []
        if not q.get("disciplina"): problemas.append("disciplina ausente")
        if not q.get("enunciado"): problemas.append("enunciado ausente")
        tipo = q.get("tipo", "multipla_escolha")
        if tipo != "multipla_escolha": problemas.append(f"tipo '{tipo}' não suportado por importação ainda")
        alts = q.get("alternativas", [])
        if not isinstance(alts, list) or len(alts) < 2:
            problemas.append("alternativas insuficientes (mínimo 2)")
        else:
            n_corretas = sum(1 for a in alts if a.get("correta"))
            if n_corretas != 1:
                problemas.append(f"deve ter exatamente 1 alternativa correta ({n_corretas} marcadas)")
        if problemas:
            erros.append(f"<li>Questão #{i}: {', '.join(problemas)}</li>")
        else:
            questoes_validas.append(q)

    # Armazena o JSON na sessão pra confirmar depois (codifica em base64 pra ficar na URL)
    import base64
    payload_b64 = base64.urlsafe_b64encode(_json.dumps({"questoes": questoes_validas}).encode("utf-8")).decode("ascii")

    # Tabela de preview
    rows = ""
    for i, q in enumerate(questoes_validas[:50], start=1):  # mostra até 50
        alt_corr = next((a["letra"] for a in q["alternativas"] if a.get("correta")), "?")
        enun_curto = re.sub(r'<[^>]+>', '', q["enunciado"])[:120]
        fonte = q.get("fonte", "—")
        rows += f"""
            <tr>
                <td>{i}</td>
                <td>{html.escape(q.get('disciplina', '—'))}</td>
                <td>{html.escape(q.get('ano', '—'))}</td>
                <td style="max-width:400px;">{html.escape(enun_curto)}{"..." if len(q["enunciado"]) > 120 else ""}</td>
                <td><span class="badge-success badge">✓ {alt_corr}</span></td>
                <td style="font-size:11px; color:var(--text-muted);">{html.escape(fonte)}</td>
            </tr>
        """
    if len(questoes_validas) > 50:
        rows += f'<tr><td colspan="6" style="text-align:center; color:var(--text-muted); padding:14px;">… e mais {len(questoes_validas) - 50} questões válidas (não exibidas pra economizar espaço)</td></tr>'

    erros_html = ""
    if erros:
        erros_html = f"""
            <div class="card" style="background:var(--orange-bg); border-left:3px solid var(--orange); margin-top:14px;">
                <strong style="color:var(--orange);">⚠️ {len(erros)} questão(ões) com problemas (serão ignoradas):</strong>
                <ul style="margin-top:6px; font-size:13px;">{"".join(erros)}</ul>
            </div>
        """

    content = f"""
        <div class="page-header">
            <h1>👀 Preview da importação</h1>
            <p class="subtitle">{len(questoes_validas)} questões prontas pra importar · {len(erros)} com erros</p>
        </div>

        {erros_html}

        <div style="overflow-x:auto; margin-top:14px;">
            <table>
                <thead>
                    <tr><th>#</th><th>Disciplina</th><th>Ano</th><th>Enunciado (resumo)</th><th>Gabarito</th><th>Fonte</th></tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
        </div>

        <form method="post" action="/admin/importar-questoes/confirmar" style="margin-top:18px;">
            <input type="hidden" name="payload" value="{payload_b64}">
            <button type="submit" class="btn btn-primary" {"disabled" if not questoes_validas else ""}>
                ✓ Confirmar importação de {len(questoes_validas)} questões
            </button>
            <a href="/admin/importar-questoes" class="btn">← Voltar</a>
        </form>
    """
    return HTMLResponse(render_page("Preview da importação", content, active=""))


@app.post("/admin/importar-questoes/confirmar", response_class=HTMLResponse)
async def confirmar_importar_questoes(request: Request):
    """Efetiva a importação: cria disciplinas/habilidades novas se necessário, insere questões + alternativas."""
    _r = _require_admin_or_403(request)
    if _r is not None: return _r

    prof = get_current_professor(request)
    form = await request.form()
    import json as _json, base64
    try:
        payload_b64 = form.get("payload", "")
        dados = _json.loads(base64.urlsafe_b64decode(payload_b64.encode("ascii")).decode("utf-8"))
        questoes = dados.get("questoes", [])
    except Exception as e:
        return HTMLResponse(render_page("Erro", f'<div class="card" style="background:var(--red-bg); color:var(--red);">Erro ao decodificar payload: {html.escape(str(e))}</div>'))

    conn = get_db()
    importadas = 0
    bncc_criadas = 0
    disciplinas_criadas = 0

    try:
        for q in questoes:
            # 1. Disciplina (cria se não existir)
            disc_nome = q["disciplina"].strip()
            row = conn.execute("SELECT id FROM disciplinas WHERE LOWER(nome) = LOWER(?)", (disc_nome,)).fetchone()
            if row:
                disc_id = row["id"]
            else:
                c = conn.execute("INSERT INTO disciplinas (nome) VALUES (?)", (disc_nome,))
                disc_id = c.lastrowid
                disciplinas_criadas += 1

            # 2. Enunciado com fonte apêndice
            enunciado = q["enunciado"]
            fonte = q.get("fonte", "").strip()
            if fonte:
                enunciado = enunciado + f'<p style="font-size:11px; color:var(--text-muted); margin-top:10px; font-style:italic;">📚 Fonte: {html.escape(fonte)}</p>'

            # 3. Insere questão
            c = conn.execute(
                "INSERT INTO questoes (disciplina_id, enunciado, ano, tipo, criada_por_professor_id) VALUES (?, ?, ?, ?, ?)",
                (disc_id, enunciado, q.get("ano", ""), q.get("tipo", "multipla_escolha"), prof["id"])
            )
            qid = c.lastrowid

            # 4. Alternativas
            for a in q["alternativas"]:
                conn.execute(
                    "INSERT INTO alternativas (questao_id, letra, texto, correta) VALUES (?, ?, ?, ?)",
                    (qid, a["letra"].upper(), a["texto"], 1 if a.get("correta") else 0)
                )

            # 5. Habilidade BNCC (vincula se existe; cria fantasma se não existir)
            bncc = q.get("habilidade_bncc", "").strip()
            if bncc:
                row = conn.execute("SELECT id FROM habilidades_bncc WHERE codigo = ?", (bncc,)).fetchone()
                if not row:
                    c2 = conn.execute("INSERT INTO habilidades_bncc (codigo, descricao) VALUES (?, ?)", (bncc, "(importada — sem descrição)"))
                    h_id = c2.lastrowid
                    bncc_criadas += 1
                else:
                    h_id = row["id"]
                conn.execute("INSERT INTO questao_habilidades (questao_id, habilidade_id) VALUES (?, ?)", (qid, h_id))

            importadas += 1

        conn.commit()
    finally:
        conn.close()

    content = f"""
        <div class="page-header">
            <h1>✅ Importação concluída</h1>
        </div>
        <div class="card" style="background:var(--green-bg); border-left:3px solid var(--green);">
            <h3 style="margin-top:0; color:var(--green);">🎉 {importadas} questão(ões) cadastrada(s) com sucesso!</h3>
            <ul style="margin-top:8px; font-size:13px;">
                <li>{importadas} questões inseridas no banco coletivo</li>
                {"<li>" + str(disciplinas_criadas) + " disciplina(s) criada(s) automaticamente</li>" if disciplinas_criadas else ""}
                {"<li>" + str(bncc_criadas) + " código(s) BNCC novo(s) cadastrado(s)</li>" if bncc_criadas else ""}
            </ul>
        </div>
        <div class="page-actions" style="margin-top:18px;">
            <a href="/questoes" class="btn btn-primary">Ver banco de questões →</a>
            <a href="/admin/importar-questoes" class="btn">Importar mais</a>
        </div>
    """
    return HTMLResponse(render_page("Importação concluída", content, active=""))


@app.get("/habilidades", response_class=HTMLResponse)
def listar_habilidades(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()
    habs = conn.execute("SELECT h.id, h.codigo, h.descricao, COUNT(qh.id) AS uso FROM habilidades_bncc h LEFT JOIN questao_habilidades qh ON qh.habilidade_id = h.id GROUP BY h.id ORDER BY h.codigo").fetchall()
    total = len(habs)
    com_desc = sum(1 for h in habs if (h["descricao"] or "").strip())
    sem_desc = total - com_desc
    conn.close()

    acoes_html = (
        f'<div class="page-actions">'
        f'<a href="/habilidades/importar" class="btn btn-primary">📥 Importar BNCC (Excel/CSV)</a>'
        f'</div>'
    )

    metricas_html = ""
    if total > 0:
        metricas_html = f"""
        <div class="metric-grid">
            <div class="metric"><div class="metric-label">Total cadastradas</div><div class="metric-value">{total}</div></div>
            <div class="metric"><div class="metric-label">Com descrição</div><div class="metric-value">{com_desc}</div></div>
            <div class="metric"><div class="metric-label">Sem descrição</div><div class="metric-value">{sem_desc}</div></div>
        </div>"""

    if habs:
        items = ""
        for h in habs:
            desc = h["descricao"] or '<em style="color:var(--text-subtle)">sem descrição</em>'
            items += f'<a href="/habilidades/{h["id"]}/editar" class="card card-link"><div class="card-title"><span class="badge">{h["codigo"]}</span></div><div class="card-meta">{desc}</div><div class="card-meta">{h["uso"]} questões usam essa habilidade</div></a>'
        body = items
    else:
        body = '<div class="empty"><p>Nenhuma habilidade cadastrada ainda.</p><p style="font-size:13px;">Use o botão <strong>Importar BNCC</strong> acima para subir a planilha oficial do MEC (1.408 habilidades do Ensino Fundamental), ou cadastre códigos digitando-os no campo BNCC ao criar questões.</p></div>'
    content = f'<div class="page-header"><h1>Habilidades BNCC</h1><p class="subtitle">Catálogo de códigos da BNCC. Clique numa habilidade para editar a descrição.</p>{acoes_html}</div>{metricas_html}{body}'
    return render_page("Habilidades BNCC", content, active="habilidades")


@app.get("/habilidades/{id}/editar", response_class=HTMLResponse)
def form_editar_habilidade(id: int):
    conn = get_db()
    h = conn.execute("SELECT * FROM habilidades_bncc WHERE id = ?", (id,)).fetchone()
    conn.close()
    if not h:
        return RedirectResponse("/habilidades", status_code=303)
    content = f'<div class="page-header"><h1><span class="badge">{h["codigo"]}</span></h1><p class="subtitle">Editar descrição.</p></div><form action="/habilidades/{id}/editar" method="post"><label>Descrição<textarea name="descricao" rows="4">{h["descricao"] or ""}</textarea></label><div class="page-actions"><button type="submit" class="btn btn-primary">Salvar</button><a href="/habilidades" class="btn">Cancelar</a></div></form>'
    return render_page(f"Editar {h['codigo']}", content, active="habilidades")


@app.post("/habilidades/{id}/editar")
def atualizar_habilidade(id: int, descricao: str = Form("")):
    conn = get_db()
    conn.execute("UPDATE habilidades_bncc SET descricao = ? WHERE id = ?", (descricao.strip() or None, id))
    conn.commit()
    conn.close()
    return RedirectResponse("/habilidades", status_code=303)



@app.post("/upload-imagem-inline")
async def upload_imagem_inline(arquivo: UploadFile = File(...)):
    """Endpoint para upload de imagem colada nas alternativas. Retorna o caminho público."""
    try:
        data = await arquivo.read()
        data = _redimensionar_imagem(data, max_width=600)
        unique_name = f"{uuid.uuid4().hex}.jpg"
        file_path = os.path.join(UPLOAD_DIR, unique_name)
        with open(file_path, "wb") as f:
            f.write(data)
        return {"url": f"/static/imagens/{unique_name}"}
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"erro": str(e)}, status_code=500)


@app.get("/questoes", response_class=HTMLResponse)
def listar_questoes(request: Request, disciplina: Optional[str] = None, ano: Optional[str] = None, bncc: Optional[str] = None, q: Optional[str] = None):
    disciplina_id: Optional[int] = int(disciplina) if (disciplina and disciplina.strip().isdigit()) else None
    prof = get_current_professor(request)
    is_admin = bool(prof and prof["is_admin"])
    conn = get_db()

    # Montar query com filtros aplicados — agora trazendo também o autor
    sql = """
        SELECT DISTINCT q.id, q.enunciado, q.ano, q.criada_por_professor_id, q.tipo,
               d.id AS disciplina_id, d.nome AS disciplina_nome,
               aut.nome AS autor_nome
        FROM questoes q
        JOIN disciplinas d ON d.id = q.disciplina_id
        LEFT JOIN professores aut ON aut.id = q.criada_por_professor_id
        LEFT JOIN questao_habilidades qh ON qh.questao_id = q.id
        LEFT JOIN habilidades_bncc h ON h.id = qh.habilidade_id
        WHERE 1=1
    """
    params = []
    if disciplina_id:
        sql += " AND d.id = ?"
        params.append(disciplina_id)
    if ano:
        sql += " AND q.ano = ?"
        params.append(ano)
    if bncc and bncc.strip():
        sql += " AND h.codigo LIKE ?"
        params.append(f"%{bncc.strip().upper()}%")
    if q and q.strip():
        sql += " AND q.enunciado LIKE ?"
        params.append(f"%{q.strip()}%")
    sql += " ORDER BY d.nome, q.id DESC"

    questoes = conn.execute(sql, params).fetchall()
    disciplinas = conn.execute("SELECT * FROM disciplinas ORDER BY nome").fetchall()
    total_geral = conn.execute("SELECT COUNT(*) AS c FROM questoes").fetchone()["c"]

    # Matriz disciplina × ano com contagens — apenas disciplinas que têm pelo menos 1 questão
    matriz_rows = conn.execute("""
        SELECT d.id AS disc_id, d.nome AS disc_nome, q.ano, COUNT(q.id) AS qtd
        FROM questoes q JOIN disciplinas d ON d.id = q.disciplina_id
        GROUP BY d.id, q.ano
        ORDER BY d.nome COLLATE NOCASE, q.ano IS NULL, q.ano
    """).fetchall()

    # Organiza por disciplina
    from collections import defaultdict
    matriz_por_disc = defaultdict(list)
    nomes_disc = {}
    for r in matriz_rows:
        matriz_por_disc[r["disc_id"]].append((r["ano"] or "", r["qtd"]))
        nomes_disc[r["disc_id"]] = r["disc_nome"]

    matriz_html = ""
    if matriz_por_disc:
        import urllib.parse as _urlp
        linhas = []
        for disc_id, anos_qtds in matriz_por_disc.items():
            total_disc = sum(qtd for _, qtd in anos_qtds)
            aberto_por_filtro = (disciplina_id == disc_id)
            badges_linha = []
            for ano_v, qtd in anos_qtds:
                rotulo = ano_v if ano_v else "Sem ano"
                qs = _urlp.urlencode({"disciplina": disc_id, "ano": ano_v} if ano_v else {"disciplina": disc_id})
                ativo = (disciplina_id == disc_id and ((ano_v and ano == ano_v) or (not ano_v and not ano)))
                cor_bg = "var(--accent)" if ativo else "var(--bg)"
                cor_fg = "white" if ativo else "var(--text)"
                borda = "var(--accent)" if ativo else "var(--border)"
                badges_linha.append(
                    f'<a href="/questoes?{qs}" class="badge" style="background:{cor_bg}; color:{cor_fg}; '
                    f'border:1px solid {borda}; text-decoration:none; padding:3px 9px; font-size:11px; '
                    f'margin:0 6px 6px 0; display:inline-block;">'
                    f'{rotulo}: {qtd}</a>'
                )
            qs_disc = _urlp.urlencode({"disciplina": disc_id})
            display_inicial = "flex" if aberto_por_filtro else "none"
            seta_inicial = "▴" if aberto_por_filtro else "▾"
            linhas.append(
                f'<div style="background:var(--card); border:1px solid var(--border); border-radius:10px; margin-bottom:8px; overflow:hidden;">'
                f'<div onclick="toggleDisciplina({disc_id})" style="display:flex; align-items:center; gap:10px; padding:11px 14px; cursor:pointer;">'
                f'<span style="flex:1; font-size:14px; font-weight:600;">{nomes_disc[disc_id]}</span>'
                f'<a href="/questoes?{qs_disc}" onclick="event.stopPropagation();" style="font-size:12px; color:var(--text-muted); text-decoration:none;">{total_disc}</a>'
                f'<span id="seta-{disc_id}" style="font-size:11px; color:var(--text-muted);">{seta_inicial}</span>'
                f'</div>'
                f'<div id="detalhe-{disc_id}" style="display:{display_inicial}; flex-wrap:wrap; padding:0 14px 10px; border-top:1px solid var(--border); padding-top:10px;">'
                f'{"".join(badges_linha)}'
                f'</div>'
                f'</div>'
            )
        matriz_html = (
            f'<div style="margin-bottom:14px;">'
            f'<div style="font-size:11px; text-transform:uppercase; letter-spacing:0.5px; color:var(--text-muted); margin-bottom:8px;">Visão geral · toque numa disciplina pra ver por ano</div>'
            f'{"".join(linhas)}'
            f'</div>'
            f'<script>'
            f'function toggleDisciplina(id) {{'
            f'  var det = document.getElementById("detalhe-" + id);'
            f'  var seta = document.getElementById("seta-" + id);'
            f'  if (det.style.display === "none" || !det.style.display) {{'
            f'    det.style.display = "flex";'
            f'    seta.textContent = "▴";'
            f'  }} else {{'
            f'    det.style.display = "none";'
            f'    seta.textContent = "▾";'
            f'  }}'
            f'}}'
            f'</script>'
        )

    disciplinas_opts = '<option value="">Todas</option>' + "".join(
        f'<option value="{d["id"]}"{(" selected" if disciplina_id == d["id"] else "")}>{d["nome"]}</option>'
        for d in disciplinas
    )
    anos_opts = '<option value="">Todos</option>' + "".join(
        f'<option value="{a}"{(" selected" if ano == a else "")}>{a}</option>'
        for a in ANOS
    )

    filtros_html = (
        f'<form action="/questoes" method="get" '
        f'style="background:var(--bg-subtle); padding:14px 16px; border-radius:8px; margin-bottom:18px;">'
        f'<div style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">'
        f'<label style="margin:0; flex:1 1 130px;">Disciplina<select name="disciplina">{disciplinas_opts}</select></label>'
        f'<label style="margin:0; flex:1 1 100px;">Ano<select name="ano">{anos_opts}</select></label>'
        f'<label style="margin:0; flex:1 1 130px;">Código BNCC<input type="text" name="bncc" placeholder="EF06MA" value="{bncc or ""}"></label>'
        f'<label style="margin:0; flex:1 1 160px;">Buscar no enunciado<input type="text" name="q" placeholder="palavra-chave" value="{q or ""}"></label>'
        f'<button type="submit" class="btn btn-primary" style="margin:0; flex:0 0 auto;">Filtrar</button>'
        f'<a href="/questoes" class="btn" style="margin:0; flex:0 0 auto;">Limpar</a>'
        f'</div></form>'
    )

    if questoes:
        cards_list = []
        for qx in questoes:
            pode_ed = _pode_editar_questao(prof, qx["criada_por_professor_id"])
            # Badge "Por: X" só pra admin (e quando o autor é diferente do admin logado)
            mostrar_autor = is_admin and qx["autor_nome"] and qx["criada_por_professor_id"] != prof["id"]
            autor_nome_card = qx["autor_nome"] if mostrar_autor else None
            cards_list.append(render_questao_card(
                conn, qx, mostrar_acoes=True, compact=True,
                pode_editar=pode_ed, autor_nome=autor_nome_card
            ))
        cards = "".join(cards_list)
    else:
        cards = '<div class="empty">Nenhuma questão encontrada com os filtros selecionados.</div>'
    conn.close()

    tem_filtro = bool(disciplina or ano or bncc or q)
    subtitle = f'{len(questoes)} de {total_geral} questão(ões)' if tem_filtro else f'{total_geral} questão(ões) cadastradas'

    toggle_js = """
    <script>
    function toggleQuestao(id) {
        const detalhes = document.getElementById('q-detalhes-' + id);
        const btn = document.getElementById('q-toggle-' + id);
        if (detalhes.style.display === 'none' || !detalhes.style.display) {
            detalhes.style.display = 'block';
            btn.textContent = 'Recolher ▴';
            if (window.MathJax) {
                // Limpa o estado anterior para forçar re-renderização
                if (MathJax.typesetClear) MathJax.typesetClear([detalhes]);
                if (MathJax.typesetPromise) MathJax.typesetPromise([detalhes]);
            }
        } else {
            detalhes.style.display = 'none';
            btn.textContent = 'Ver completa ▾';
        }
    }
    </script>
    """

    content = (
        f'<div class="page-header"><h1>Banco de questões</h1>'
        f'<p class="subtitle">{subtitle}</p>'
        f'<div class="page-actions"><a href="/questoes/nova" class="btn btn-primary">+ Nova questão</a></div></div>'
        f'{matriz_html}{filtros_html}{cards}{toggle_js}'
    )
    return render_page("Questões", content, active="questoes", head_extra=MATHJAX)


@app.get("/questoes/nova", response_class=HTMLResponse)
def form_nova_questao_passo1():
    """Passo 1: seleciona disciplina, ano e habilidades BNCC antes do cadastro completo."""
    conn = get_db()
    disciplinas = conn.execute("SELECT * FROM disciplinas ORDER BY nome").fetchall()
    habs_existentes = conn.execute("SELECT codigo FROM habilidades_bncc ORDER BY codigo").fetchall()
    conn.close()
    if not disciplinas:
        return render_page("Nova questão", '<div class="page-header"><h1>Nova questão</h1></div><div class="empty"><p>Você precisa cadastrar pelo menos uma disciplina antes de criar questões.</p><a href="/disciplinas/nova" class="btn btn-primary">Cadastrar disciplina</a></div>', active="questoes")

    options = "".join(f'<option value="{d["id"]}">{d["nome"]}</option>' for d in disciplinas)
    anos_options = '<option value="">— Não definido —</option>' + "".join(f'<option value="{a}">{a}</option>' for a in ANOS)

    total_habs = len(habs_existentes)
    link_catalogo = (
        f'<p class="muted-line" style="font-size:11px;">'
        f'💡 {total_habs} habilidade(s) cadastrada(s) no catálogo. '
        f'<a href="/habilidades" target="_blank" style="color:var(--text-muted);">Consultar lista completa</a>'
        f'</p>'
    ) if total_habs > 0 else '<p class="muted-line" style="font-size:11px;">Nenhuma habilidade cadastrada ainda. <a href="/habilidades/importar" target="_blank">Importar BNCC oficial</a>.</p>'

    js_preview = '\n    <script>\n    (function() {\n        var container = document.getElementById(\'bncc-container\');\n        var hiddenInput = document.getElementById(\'bncc-hidden\');\n        var searchInput = document.getElementById(\'bncc-search\');\n        var chipsDiv = document.getElementById(\'bncc-chips\');\n        var resultsDiv = document.getElementById(\'bncc-results\');\n        var discSel = document.querySelector(\'select[name="disciplina_id"]\');\n        if (!container || !hiddenInput || !searchInput) return;\n        var selecionados = [];\n        function renderChips() {\n            chipsDiv.innerHTML = \'\';\n            selecionados.forEach(function(cod) {\n                var chip = document.createElement(\'span\');\n                chip.style.cssText = \'display:inline-flex;align-items:center;gap:4px;background:var(--accent-bg);color:var(--accent);border:1px solid var(--accent-border);border-radius:4px;padding:2px 8px;font-size:12px;font-weight:600;\';\n                chip.innerHTML = cod + \' <button type="button" style="background:none;border:none;cursor:pointer;color:var(--accent);font-size:14px;padding:0;line-height:1;" title="Remover">\\xd7</button>\';\n                chip.querySelector(\'button\').addEventListener(\'click\', function() {\n                    selecionados = selecionados.filter(function(c){return c!==cod;});\n                    renderChips();\n                });\n                chipsDiv.appendChild(chip);\n            });\n            hiddenInput.value = selecionados.join(\', \');\n        }\n        function adicionar(cod) {\n            cod = cod.trim().toUpperCase();\n            if (!cod || selecionados.indexOf(cod) >= 0) return;\n            selecionados.push(cod); renderChips(); resultsDiv.innerHTML = \'\'; searchInput.value = \'\';\n        }\n        function buscar() {\n            var q = searchInput.value.trim();\n            if (q.length < 2) { resultsDiv.innerHTML = \'\'; return; }\n            var disc = discSel ? discSel.value : \'\';\n            var pareceCode = /^[A-Za-z]{2}\\d{2}[A-Za-z]{2}\\d{2}/.test(q);\n            var url = pareceCode ? \'/habilidades/buscar?codigos=\' + encodeURIComponent(q.toUpperCase())\n                : \'/habilidades/buscar?q=\' + encodeURIComponent(q) + (disc ? \'&disciplina_id=\' + disc : \'\');\n            fetch(url).then(function(r){return r.json();}).then(function(data) {\n                var results = [];\n                if (pareceCode) { Object.keys(data).forEach(function(k){if(k!==\'results\') results.push({codigo:k,descricao:data[k]});}); }\n                else { results = data.results || []; }\n                if (results.length === 0) {\n                    if (pareceCode) {\n                        resultsDiv.innerHTML = \'<div style="padding:6px 8px;font-size:12px;color:var(--text-muted);">Código não encontrado. <button type="button" style="background:none;border:none;color:var(--accent);cursor:pointer;font-size:12px;padding:0;text-decoration:underline;">Adicionar mesmo assim</button></div>\';\n                        resultsDiv.querySelector(\'button\').addEventListener(\'click\', function(){adicionar(q);});\n                    } else {\n                        resultsDiv.innerHTML = \'<div style="padding:6px 8px;font-size:12px;color:var(--text-muted);">Nenhum resultado.</div>\';\n                    }\n                    return;\n                }\n                var html = \'<div style="color:var(--text-muted);font-size:11px;padding:4px 2px;">\' + results.length + \' habilidade(s) \\u2014 clique para adicionar:</div>\';\n                results.forEach(function(r) {\n                    html += \'<div data-cod="\' + r.codigo + \'" style="padding:6px 8px;border:1px solid var(--border);border-radius:4px;margin-bottom:3px;cursor:pointer;background:var(--card);font-size:12px;" onmouseover="this.style.background=\\\'var(--accent-bg)\\\'" onmouseout="this.style.background=\\\'var(--card)\\\'"><strong style="color:var(--accent);">\' + r.codigo + \'</strong> \\xb7 \' + (r.descricao||\'\').replace(/</g,\'&lt;\') + \'</div>\';\n                });\n                resultsDiv.innerHTML = html;\n            }).catch(function(){resultsDiv.innerHTML=\'\';});\n        }\n        var _t;\n        searchInput.addEventListener(\'input\', function(){clearTimeout(_t); _t=setTimeout(buscar,350);});\n        searchInput.addEventListener(\'keydown\', function(e){if(e.key===\'Enter\'){e.preventDefault();buscar();}});\n        if (discSel) discSel.addEventListener(\'change\', buscar);\n        resultsDiv.addEventListener(\'click\', function(e){\n            var item = e.target.closest(\'[data-cod]\');\n            if (item) adicionar(item.dataset.cod);\n        });\n        var init = hiddenInput.value.trim();\n        if (init) {\n            init.split(/[,\\n]/).map(function(x){return x.trim().toUpperCase();}).filter(Boolean).forEach(function(c){\n                if(selecionados.indexOf(c)<0) selecionados.push(c);\n            });\n            renderChips();\n        }\n    })();\n    </script>\n'

    tipo_options = "".join(
        f'<option value="{k}">{v["icone"]} {v["label"]}</option>'
        for k, v in TIPOS_QUESTAO.items()
    )

    content = f"""
        <div class="page-header">
            <h1>Nova questão</h1>
            <p class="subtitle">Passo 1 de 2 — defina o tipo, disciplina, ano e habilidades.</p>
        </div>
        <form action="/questoes/nova/passo2" method="post">
            <label>Tipo de questão<select name="tipo" required>{tipo_options}</select></label>
            <div style="display:grid; grid-template-columns: 2fr 1fr; gap:12px;">
                <label>Disciplina<select name="disciplina_id" required>{options}</select></label>
                <label>Ano de escolaridade<select name="ano">{anos_options}</select></label>
            </div>
            <div id="bncc-container" style="margin:10px 0;">
                <label style="margin-bottom:6px;">Habilidades BNCC <span style="font-weight:400; color:var(--red); font-size:12px;">* obrigatório</span></label>
                <input type="hidden" name="habilidades_codigos" id="bncc-hidden">
                <div id="bncc-chips" style="display:flex; flex-wrap:wrap; gap:6px; min-height:24px; margin-bottom:8px;"></div>
                <input type="search" id="bncc-search" placeholder="Digite o código (EF09MA09) ou palavra-chave (fração, célula...)" style="margin:0;">
                <div id="bncc-results" style="margin-top:6px;"></div>
            </div>
            {link_catalogo}
            <div class="page-actions">
                <button type="submit" class="btn btn-primary" onclick="var h=document.getElementById(&quot;bncc-hidden&quot;); if(!h||!h.value.trim()){{alert(&quot;Selecione pelo menos uma Habilidade BNCC antes de avançar.&quot;);return false;}}">Próximo: cadastrar conteúdo →</button>
                <a href="/questoes" class="btn">Cancelar</a>
            </div>
        </form>
        {js_preview}
    """
    return render_page("Nova questão · Passo 1", content, active="questoes")


@app.post("/questoes/nova/passo2", response_class=HTMLResponse)
def form_nova_questao_passo2(
    disciplina_id: int = Form(...),
    ano: str = Form(""),
    habilidades_codigos: str = Form(""),
    tipo: str = Form("multipla_escolha"),
):
    """Passo 2: cadastramento do conteúdo da questão. Dados do Passo 1 ficam em hidden fields."""
    if tipo not in TIPOS_QUESTAO:
        tipo = "multipla_escolha"
    conn = get_db()
    disciplina = conn.execute("SELECT * FROM disciplinas WHERE id = ?", (disciplina_id,)).fetchone()
    conn.close()
    if not disciplina:
        return RedirectResponse("/questoes/nova", status_code=303)

    ano_label = ano if ano else "— não definido —"
    tipo_info = TIPOS_QUESTAO[tipo]

    # Badges informativas das habilidades digitadas no passo 1
    codigos_clean = [c.strip().upper() for c in habilidades_codigos.replace("\n", ",").split(",") if c.strip()]
    badges_bncc = "".join(f'<span class="badge">{c}</span>' for c in codigos_clean) if codigos_clean else '<span class="muted-line">— sem BNCC —</span>'

    # Bloco específico do tipo da questão
    fieldset_alternativas = ""
    enunciado_detecta_alts = (tipo == "multipla_escolha")
    if tipo == "multipla_escolha":
        alternativas_html = ""
        for letra in ["A", "B", "C", "D"]:
            required_radio = ' required' if letra == "A" else ''
            editor_alt = _editor_enunciado_html(
                name=f"alt_{letra.lower()}", valor_inicial="", required=True,
                label="", compact=True, min_height=42,
                placeholder=f"Texto da alternativa {letra}"
            )
            alternativas_html += (
                f'<div style="display:grid; grid-template-columns:auto 1fr; gap:12px; align-items:flex-start; margin-bottom:10px;">'
                f'<label style="margin:8px 0 0 0; display:flex; align-items:center; gap:8px; white-space:nowrap;">'
                f'<input type="radio" name="correta" value="{letra}"{required_radio} style="width:auto; margin:0;"> <strong>{letra})</strong>'
                f'</label>'
                f'<div style="margin:0;">{editor_alt}</div>'
                f'</div>'
            )
        fieldset_alternativas = f"""
            <fieldset>
                <legend>Alternativas — marque o radio da correta</legend>
                {alternativas_html}
            </fieldset>
        """
    elif tipo == "discursiva":
        # Discursiva: sem alternativas. Aviso visual.
        fieldset_alternativas = """
            <div style="background:var(--accent-bg); color:var(--accent); border:1px solid var(--accent); padding:14px 16px; border-radius:6px; margin:12px 0;">
                <strong>📝 Questão discursiva</strong><br>
                <span style="font-size:13px;">O aluno responderá em texto livre. No modo impresso, será reservado espaço para resposta manuscrita. A correção é manual — feita por você fora do sistema.</span>
            </div>
        """
    elif tipo == "vf":
        # V ou F: até 5 afirmações, cada uma com radio V/F
        afirms_html = ""
        for i in range(VF_MAX_AFIRMACOES):
            editor_afirm = _editor_enunciado_html(
                name=f"vf_afirm_{i}_texto", valor_inicial="", required=False,
                label="", compact=True, min_height=42,
                placeholder=f"Afirmação {i+1} (deixe em branco se não usar)"
            )
            afirms_html += (
                f'<div style="display:grid; grid-template-columns:1fr auto; gap:12px; align-items:flex-start; margin-bottom:10px;">'
                f'<div style="margin:0;"><strong style="font-size:13px;">Afirmação {i+1}</strong>{editor_afirm}</div>'
                f'<div style="display:flex; gap:10px; align-items:center; padding-top:24px; white-space:nowrap;">'
                f'<label style="margin:0; font-size:13px;"><input type="radio" name="vf_afirm_{i}_gabarito" value="V" style="width:auto; margin:0 4px 0 0;">V</label>'
                f'<label style="margin:0; font-size:13px;"><input type="radio" name="vf_afirm_{i}_gabarito" value="F" style="width:auto; margin:0 4px 0 0;">F</label>'
                f'</div></div>'
            )
        fieldset_alternativas = f"""
            <fieldset>
                <legend>Afirmações — marque V ou F para cada (até {VF_MAX_AFIRMACOES})</legend>
                <p class="muted-line" style="font-size:12px; margin:0 0 10px 0;">Deixe em branco as afirmações que não usar (mínimo 2 afirmações preenchidas).</p>
                {afirms_html}
            </fieldset>
        """
    elif tipo == "associacao":
        # Associação: 2 colunas de até 5 itens; coluna A tem texto + qual letra da B é a resposta correta
        col_a_html = ""
        for i in range(ASSOC_MAX_PARES):
            editor_a = _editor_enunciado_html(
                name=f"assoc_a_{i}_texto", valor_inicial="", required=False,
                label="", compact=True, min_height=42,
                placeholder=f"Item {i+1} da coluna A (em branco se não usar)"
            )
            # Select pra escolher qual letra da B é o gabarito
            letras_options = '<option value="">—</option>' + "".join(
                f'<option value="{chr(97+j)}">{chr(97+j)}</option>' for j in range(ASSOC_MAX_PARES)
            )
            col_a_html += (
                f'<div style="display:grid; grid-template-columns:auto 1fr auto; gap:12px; align-items:flex-start; margin-bottom:10px;">'
                f'<strong style="padding-top:20px;">{i+1}.</strong>'
                f'<div style="margin:0;">{editor_a}</div>'
                f'<label style="margin:0; padding-top:14px; font-size:12px; white-space:nowrap;">Resposta: '
                f'<select name="assoc_a_{i}_gabarito" style="width:auto; display:inline-block; margin-left:4px;">{letras_options}</select>'
                f'</label></div>'
            )
        col_b_html = ""
        for j in range(ASSOC_MAX_PARES):
            letra_b = chr(97+j)
            editor_b = _editor_enunciado_html(
                name=f"assoc_b_{letra_b}_texto", valor_inicial="", required=False,
                label="", compact=True, min_height=42,
                placeholder=f"Item ({letra_b}) da coluna B (em branco se não usar)"
            )
            col_b_html += (
                f'<div style="display:grid; grid-template-columns:auto 1fr; gap:12px; align-items:flex-start; margin-bottom:10px;">'
                f'<strong style="padding-top:20px;">({letra_b})</strong>'
                f'<div style="margin:0;">{editor_b}</div>'
                f'</div>'
            )
        fieldset_alternativas = f"""
            <fieldset>
                <legend>Coluna A — itens (1, 2, 3...) com gabarito da resposta</legend>
                <p class="muted-line" style="font-size:12px; margin:0 0 10px 0;">Para cada item da coluna A, indique qual letra da coluna B é a resposta correta. Mínimo 2 pares preenchidos.</p>
                {col_a_html}
            </fieldset>
            <fieldset>
                <legend>Coluna B — opções de associação (a, b, c...)</legend>
                {col_b_html}
            </fieldset>
        """

    # Hidden fields carregam dados do passo 1; valores escapados
    import html as _html
    h_disc = _html.escape(str(disciplina_id), quote=True)
    h_ano = _html.escape(ano, quote=True)
    h_habs = _html.escape(habilidades_codigos, quote=True)
    h_tipo = _html.escape(tipo, quote=True)

    content = f"""
        <div class="page-header">
            <h1>Nova questão</h1>
            <p class="subtitle">Passo 2 de 2 — conteúdo da questão.</p>
        </div>

        <div style="background:var(--bg-subtle); border:1px solid var(--border); border-radius:8px; padding:12px 16px; margin-bottom:18px; display:flex; align-items:center; gap:16px; flex-wrap:wrap;">
            <div><strong>Tipo:</strong> {tipo_info['icone']} {tipo_info['label']}</div>
            <div><strong>Disciplina:</strong> {disciplina["nome"]}</div>
            <div><strong>Ano:</strong> {ano_label}</div>
            <div><strong>BNCC:</strong> {badges_bncc}</div>
            <a href="/questoes/nova" style="margin-left:auto; font-size:13px; color:var(--text-muted);">← Voltar e alterar</a>
        </div>

        <div class="tip"><strong>Dica:</strong> use <code>$fórmula$</code> para fórmulas inline ou <code>$$fórmula$$</code> para centralizadas.</div>

        <form action="/questoes/criar" method="post" enctype="multipart/form-data">
            <input type="hidden" name="disciplina_id" value="{h_disc}">
            <input type="hidden" name="ano" value="{h_ano}">
            <input type="hidden" name="habilidades_codigos" value="{h_habs}">
            <input type="hidden" name="tipo" value="{h_tipo}">

            <style>
                .coll-sec{{border:1px solid var(--border);border-radius:8px;margin-bottom:12px;overflow:hidden;font-style:normal;}}
                .coll-hdr{{display:flex;align-items:center;justify-content:space-between;padding:10px 14px;background:var(--bg-subtle);cursor:pointer;user-select:none;font-size:12px;font-weight:600;color:var(--text-muted);letter-spacing:0.05em;text-transform:uppercase;font-style:normal;}}
                .coll-hdr:hover{{background:var(--border);}}
                .coll-arrow{{font-size:11px;transition:transform 0.2s;font-style:normal;}}
                .coll-body{{padding:14px;display:none;font-style:normal;}}
                .coll-sec.open .coll-body{{display:block;}}
                .coll-sec.open .coll-arrow{{transform:rotate(180deg);}}
            </style>
            <script>function toggleColl(el){{el.closest('.coll-sec').classList.toggle('open');}}</script>

            <div class="coll-sec">
                <div class="coll-hdr" onclick="toggleColl(this)">
                    <span>📝 Textos de apoio (opcionais)</span><span class="coll-arrow">▼</span>
                </div>
                <div class="coll-body">
                    {_editor_enunciado_html(name="texto1_conteudo", valor_inicial="", required=False, label="Texto 1 — conteúdo", min_height=80, placeholder="Cole ou digite aqui o texto de apoio (opcional)")}
                    <label>Texto 1 — fonte<input type="text" name="texto1_fonte" placeholder="Autor, obra, ano"></label>
                    {_editor_enunciado_html(name="texto2_conteudo", valor_inicial="", required=False, label="Texto 2 — conteúdo", min_height=80, placeholder="Segundo texto de apoio (opcional)")}
                    <label>Texto 2 — fonte<input type="text" name="texto2_fonte" placeholder="Autor, obra, ano"></label>
                </div>
            </div>

            <div class="coll-sec">
                <div class="coll-hdr" onclick="toggleColl(this)">
                    <span>🖼️ Imagens (opcionais)</span><span class="coll-arrow">▼</span>
                </div>
                <div class="coll-body">
                    <label>Imagem 1<input type="file" name="imagem1" accept="image/*"></label>
                    <label>Legenda da imagem 1<input type="text" name="imagem1_legenda"></label>
                    <label>Fonte da imagem 1<input type="text" name="imagem1_fonte"></label>
                    <label>Imagem 2<input type="file" name="imagem2" accept="image/*"></label>
                    <label>Legenda da imagem 2<input type="text" name="imagem2_legenda"></label>
                    <label>Fonte da imagem 2<input type="text" name="imagem2_fonte"></label>
                </div>
            </div>

            {_editor_enunciado_html(name="enunciado", valor_inicial="", required=True, label="Enunciado", placeholder="Digite o enunciado da questão. Use a barra abaixo para formatar.", detectar_alternativas=enunciado_detecta_alts)}

            {fieldset_alternativas}

            <div class="page-actions">
                <button type="submit" class="btn btn-primary">Cadastrar questão</button>
                <a href="/questoes/nova" class="btn">← Voltar</a>
            </div>
        </form>
    """
    return render_page("Nova questão · Passo 2", content, active="questoes", head_extra=MATHJAX_EDIT)


@app.post("/questoes/criar")
async def criar_questao(
    request: Request,
    disciplina_id: int = Form(...), enunciado: str = Form(...),
    tipo: str = Form("multipla_escolha"),
    alt_a: str = Form(""), alt_b: str = Form(""), alt_c: str = Form(""), alt_d: str = Form(""),
    correta: str = Form(""), habilidades_codigos: str = Form(""),
    ano: str = Form(""),
    texto1_conteudo: str = Form(""), texto1_fonte: str = Form(""),
    texto2_conteudo: str = Form(""), texto2_fonte: str = Form(""),
    imagem1: Optional[UploadFile] = File(None), imagem1_legenda: str = Form(""), imagem1_fonte: str = Form(""),
    imagem2: Optional[UploadFile] = File(None), imagem2_legenda: str = Form(""), imagem2_fonte: str = Form(""),
):
    if tipo not in TIPOS_QUESTAO:
        tipo = "multipla_escolha"
    # Form recebe dinamicamente os campos de V/F e Associação; pega tudo via request
    form_extra = await request.form()
    prof = get_current_professor(request)
    prof_id = prof["id"] if prof else None
    conn = get_db()
    cursor = conn.execute(
        "INSERT INTO questoes (disciplina_id, enunciado, ano, criada_por_professor_id, tipo) VALUES (?, ?, ?, ?, ?)",
        (disciplina_id, _sanitizar_html_enunciado(enunciado), ano.strip() or None, prof_id, tipo)
    )
    questao_id = cursor.lastrowid

    for ordem, (conteudo, fonte) in enumerate([(texto1_conteudo, texto1_fonte), (texto2_conteudo, texto2_fonte)]):
        conteudo_sanit = _sanitizar_html_enunciado(conteudo)
        if conteudo_sanit:
            conn.execute("INSERT INTO textos_apoio (questao_id, conteudo, fonte, ordem) VALUES (?, ?, ?, ?)", (questao_id, conteudo_sanit, fonte.strip() or None, ordem))

    for ordem, (img, legenda, fonte) in enumerate([(imagem1, imagem1_legenda, imagem1_fonte), (imagem2, imagem2_legenda, imagem2_fonte)]):
        if img and img.filename:
            content_bytes = await img.read()
            content_bytes = _redimensionar_imagem(content_bytes, max_width=800)
            unique_name = f"{uuid.uuid4().hex}.jpg"
            file_path = os.path.join(UPLOAD_DIR, unique_name)
            with open(file_path, "wb") as f:
                f.write(content_bytes)
            conn.execute("INSERT INTO imagens (questao_id, caminho, legenda, fonte, ordem) VALUES (?, ?, ?, ?, ?)", (questao_id, f"static/imagens/{unique_name}", legenda.strip() or None, fonte.strip() or None, ordem))

    # Conteúdo específico do tipo
    if tipo == "multipla_escolha":
        for letra, texto in [("A", alt_a), ("B", alt_b), ("C", alt_c), ("D", alt_d)]:
            conn.execute("INSERT INTO alternativas (questao_id, letra, texto, correta) VALUES (?, ?, ?, ?)", (questao_id, letra, _sanitizar_html_enunciado(texto), 1 if letra == correta else 0))
    elif tipo == "vf":
        ordem_real = 0
        for i in range(VF_MAX_AFIRMACOES):
            texto_afirm = _sanitizar_html_enunciado(str(form_extra.get(f"vf_afirm_{i}_texto", "")))
            gabarito = str(form_extra.get(f"vf_afirm_{i}_gabarito", "")).strip().upper()
            if texto_afirm and gabarito in ("V", "F"):
                conn.execute("INSERT INTO vf_afirmacoes (questao_id, ordem, texto, gabarito) VALUES (?, ?, ?, ?)",
                             (questao_id, ordem_real, texto_afirm, gabarito))
                ordem_real += 1
    elif tipo == "associacao":
        # Coluna A (com gabarito)
        ordem_real = 0
        for i in range(ASSOC_MAX_PARES):
            texto_a = _sanitizar_html_enunciado(str(form_extra.get(f"assoc_a_{i}_texto", "")))
            gabarito = str(form_extra.get(f"assoc_a_{i}_gabarito", "")).strip().lower()
            if texto_a and gabarito:
                conn.execute("INSERT INTO assoc_itens_a (questao_id, ordem, texto, gabarito_letra) VALUES (?, ?, ?, ?)",
                             (questao_id, ordem_real, texto_a, gabarito))
                ordem_real += 1
        # Coluna B (opções)
        for j in range(ASSOC_MAX_PARES):
            letra_b = chr(97+j)
            texto_b = _sanitizar_html_enunciado(str(form_extra.get(f"assoc_b_{letra_b}_texto", "")))
            if texto_b:
                conn.execute("INSERT INTO assoc_itens_b (questao_id, letra, texto) VALUES (?, ?, ?)",
                             (questao_id, letra_b, texto_b))

    for parte in habilidades_codigos.replace("\n", ",").split(","):
        codigo = parte.strip().upper()
        if not codigo: continue
        existing = conn.execute("SELECT id FROM habilidades_bncc WHERE codigo = ?", (codigo,)).fetchone()
        habilidade_id = existing["id"] if existing else conn.execute("INSERT INTO habilidades_bncc (codigo) VALUES (?)", (codigo,)).lastrowid
        try:
            conn.execute("INSERT INTO questao_habilidades (questao_id, habilidade_id) VALUES (?, ?)", (questao_id, habilidade_id))
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    conn.close()

    # Buscar provas do professor para o atalho "Adicionar a uma atividade"
    conn2 = get_db()
    provas_recentes = conn2.execute(
        "SELECT id, titulo FROM provas WHERE criada_por_professor_id = ? ORDER BY id DESC LIMIT 8",
        (prof["id"],)
    ).fetchall()
    conn2.close()

    opts_provas = "".join(
        f'<option value="{p["id"]}">{p["titulo"]}</option>'
        for p in provas_recentes
    )
    form_adicionar = f"""
        <form method="post" action="/questoes/{questao_id}/adicionar-a-prova" style="margin:0;">
            <div style="display:flex; gap:8px; align-items:flex-end; flex-wrap:wrap;">
                <label style="margin:0; flex:1; min-width:180px;">
                    Escolha a atividade
                    <select name="prova_id" required>
                        <option value="">— selecione —</option>
                        {opts_provas}
                    </select>
                </label>
                <button type="submit" class="btn btn-primary" style="margin:0;">Adicionar →</button>
            </div>
        </form>
    """ if provas_recentes else '<p style="color:var(--text-muted); font-size:13px; margin:0;">Você ainda não tem atividades criadas. <a href="/provas/nova">Criar uma agora →</a></p>'

    content_html = f"""
        <div style="max-width:560px; margin:60px auto; text-align:center; padding:0 20px;">
            <div style="font-size:52px; margin-bottom:12px;">✅</div>
            <h1 style="font-size:22px; margin-bottom:6px;">Questão salva!</h1>
            <p style="color:var(--text-muted); margin-bottom:32px;">O que deseja fazer agora?</p>
            <div style="display:flex; gap:10px; justify-content:center; flex-wrap:wrap; margin-bottom:32px;">
                <a href="/questoes/nova" class="btn btn-primary">✏️ Criar outra questão</a>
                <a href="/questoes" class="btn">📚 Ver banco de questões</a>
                <a href="/provas/nova" class="btn">📝 Criar nova atividade</a>
            </div>
            <div style="background:var(--bg-subtle); border:1px solid var(--border); border-radius:10px; padding:18px; text-align:left;">
                <p style="font-weight:600; font-size:13px; margin:0 0 10px 0;">➕ Adicionar esta questão a uma atividade existente:</p>
                {form_adicionar}
            </div>
        </div>
    """
    return HTMLResponse(render_page("Questão salva", content_html, active="questoes"))


@app.post("/questoes/{questao_id}/adicionar-a-prova", response_class=HTMLResponse)
def adicionar_questao_a_prova(questao_id: int, prova_id: int = Form(...)):
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)
    conn = get_db()
    # Verificar se questão já está na prova
    ja_existe = conn.execute(
        "SELECT id FROM prova_questoes WHERE prova_id = ? AND questao_id = ?",
        (prova_id, questao_id)
    ).fetchone()
    if not ja_existe:
        max_ordem = conn.execute(
            "SELECT COALESCE(MAX(ordem), -1) + 1 AS n FROM prova_questoes WHERE prova_id = ?",
            (prova_id,)
        ).fetchone()["n"]
        conn.execute(
            "INSERT INTO prova_questoes (prova_id, questao_id, ordem) VALUES (?, ?, ?)",
            (prova_id, questao_id, max_ordem)
        )
        conn.commit()
    conn.close()
    return RedirectResponse(f"/provas/{prova_id}", status_code=303)


# ==========================================
#  ROTAS DE PROVAS (ATUALIZADAS TAREFA A2)
# ==========================================

@app.get("/provas", response_class=HTMLResponse)
def listar_provas(request: Request, disciplina: Optional[str] = None, ano: Optional[str] = None, q: Optional[str] = None):
    disciplina_id: Optional[int] = int(disciplina) if (disciplina and disciplina.strip().isdigit()) else None
    prof = get_current_professor(request)
    is_admin = prof and prof["is_admin"]
    conn = get_db()

    # Filtros: admin vê tudo da escola; prof comum só as próprias
    where_extras = []
    params = []
    if not is_admin:
        where_extras.append("(p.criada_por_professor_id = ? OR p.criada_por_professor_id IS NULL)")
        params.append(prof["id"])
    if q and q.strip():
        where_extras.append("p.titulo LIKE ?")
        params.append(f"%{q.strip()}%")
    if disciplina_id:
        where_extras.append("""EXISTS (
            SELECT 1 FROM prova_questoes pq2
            JOIN questoes q2 ON q2.id = pq2.questao_id
            WHERE pq2.prova_id = p.id AND q2.disciplina_id = ?
        )""")
        params.append(disciplina_id)
    if ano and ano.strip():
        where_extras.append("""EXISTS (
            SELECT 1 FROM prova_questoes pq3
            JOIN questoes q3 ON q3.id = pq3.questao_id
            WHERE pq3.prova_id = p.id AND q3.ano = ?
        )""")
        params.append(ano)
    where_clause = " WHERE " + " AND ".join(where_extras) if where_extras else ""

    sql = f"""
        SELECT p.id, p.titulo, p.descricao, p.criada_por_professor_id,
               prof.nome AS criador_nome,
               (SELECT COUNT(*) FROM prova_questoes WHERE prova_id = p.id) AS qtd_questoes
        FROM provas p
        LEFT JOIN professores prof ON prof.id = p.criada_por_professor_id
        {where_clause}
        ORDER BY p.id DESC
    """
    provas = conn.execute(sql, params).fetchall()

    # Tags (disciplinas + anos) de cada prova — query única pra todas
    tags_map = {}
    if provas:
        prova_ids = [p["id"] for p in provas]
        placeholders = ",".join("?" * len(prova_ids))
        tags_rows = conn.execute(f"""
            SELECT pq.prova_id, d.nome AS disc_nome, q.ano
            FROM prova_questoes pq
            JOIN questoes q ON q.id = pq.questao_id
            JOIN disciplinas d ON d.id = q.disciplina_id
            WHERE pq.prova_id IN ({placeholders})
        """, prova_ids).fetchall()
        for r in tags_rows:
            tm = tags_map.setdefault(r["prova_id"], {"disciplinas": set(), "anos": set()})
            tm["disciplinas"].add(r["disc_nome"])
            if r["ano"]:
                tm["anos"].add(r["ano"])

    # Aplicações por prova (pra mostrar no card)
    apl_count = {row["prova_id"]: row["c"] for row in conn.execute(
        "SELECT prova_id, COUNT(*) AS c FROM aplicacoes GROUP BY prova_id"
    ).fetchall()}

    disciplinas_lista = conn.execute("SELECT * FROM disciplinas ORDER BY nome").fetchall()
    total_geral = conn.execute("SELECT COUNT(*) AS c FROM provas").fetchone()["c"]
    conn.close()

    # Filtros
    disciplinas_opts = '<option value="">Todas</option>' + "".join(
        f'<option value="{d["id"]}"{(" selected" if disciplina_id == d["id"] else "")}>{d["nome"]}</option>'
        for d in disciplinas_lista
    )
    anos_opts = '<option value="">Todos</option>' + "".join(
        f'<option value="{a}"{(" selected" if ano == a else "")}>{a}</option>'
        for a in ANOS
    )
    filtros_html = (
        f'<form action="/provas" method="get" '
        f'style="background:var(--bg-subtle); padding:14px 16px; border-radius:8px; margin-bottom:18px;">'
        f'<div style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">'
        f'<label style="margin:0; flex:1 1 160px;">Buscar por título<input type="text" name="q" placeholder="palavra do título" value="{q or ""}"></label>'
        f'<label style="margin:0; flex:1 1 130px;">Disciplina<select name="disciplina">{disciplinas_opts}</select></label>'
        f'<label style="margin:0; flex:1 1 100px;">Ano<select name="ano">{anos_opts}</select></label>'
        f'<button type="submit" class="btn btn-primary" style="margin:0; flex:0 0 auto;">Filtrar</button>'
        f'<a href="/provas" class="btn" style="margin:0; flex:0 0 auto;">Limpar</a>'
        f'</div></form>'
    )

    # Cards
    if provas:
        cards = ""
        for p in provas:
            tm = tags_map.get(p["id"], {"disciplinas": set(), "anos": set()})
            disc_tags = "".join(f'<span class="badge" style="background:var(--accent-bg); color:var(--accent);">{d}</span>' for d in sorted(tm["disciplinas"]))
            ano_tags = "".join(f'<span class="badge">{a}</span>' for a in sorted(tm["anos"]))
            desc = f'<div style="font-size:13px; color:var(--text-muted); margin-top:4px;">{p["descricao"]}</div>' if p["descricao"] else ""
            n_apl = apl_count.get(p["id"], 0)
            apl_badge = f'<span class="badge" style="background:var(--orange-bg); color:var(--orange);">{n_apl} aplicação{"" if n_apl == 1 else "ões"}</span>' if n_apl else ""

            # Badge "Por: <nome>" só pra admin (pra ele saber de quem é cada prova)
            autor_badge = ""
            if is_admin:
                nome_autor = p["criador_nome"] if p["criador_nome"] else "—"
                autor_badge = f'<span class="badge" style="background:var(--purple-bg); color:var(--purple);">Por: {nome_autor}</span>'

            cards += f"""
            <div style="background:var(--bg); border:1px solid var(--border); border-radius:8px; padding:14px 18px; margin-bottom:10px;">
                <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:14px;">
                    <div style="flex:1; min-width:0;">
                        <div style="font-weight:600; font-size:16px;">
                            <a href="/provas/{p["id"]}" style="color:inherit; text-decoration:none;">{p["titulo"]}</a>
                        </div>
                        {desc}
                        <div style="display:flex; gap:6px; flex-wrap:wrap; margin-top:8px; align-items:center;">
                            <span class="badge">{p["qtd_questoes"]} questões</span>
                            {disc_tags}{ano_tags}{apl_badge}{autor_badge}
                        </div>
                    </div>
                    <div style="display:flex; gap:6px; flex-shrink:0;">
                        <a href="/provas/{p["id"]}" class="btn" style="padding:4px 10px; font-size:12px;">Abrir</a>
                        <a href="/provas/{p["id"]}/editar" class="btn" style="padding:4px 10px; font-size:12px;">Editar</a>
                        <form action="/provas/{p["id"]}/deletar" method="post" style="margin:0;" onsubmit="return confirm('Excluir esta prova | tarefa? Se ela tiver aplicações, a exclusão será bloqueada.');">
                            <button type="submit" class="btn" style="padding:4px 10px; font-size:12px; background:var(--red); color:white; border-color:var(--red);">Excluir</button>
                        </form>
                    </div>
                </div>
            </div>
            """
    else:
        cards = '<div class="empty">Nenhuma prova | tarefa encontrada com esses filtros.</div>'

    tem_filtro = bool(disciplina or ano or q)
    subtitle = f'{len(provas)} de {total_geral} prova(s) | tarefa(s)' if tem_filtro else f'{total_geral} prova(s) | tarefa(s) cadastrada(s)'

    content = f"""
        <div class="page-header">
            <h1>Provas | Tarefas</h1>
            <p class="subtitle">{subtitle}</p>
            <div class="page-actions"><a href="/provas/nova" class="btn btn-primary">+ Nova Prova | Tarefa</a></div>
        </div>
        {filtros_html}
        {cards}
    """
    return render_page("Provas | Tarefas", content, active="provas")


def _render_picker_questoes(conn, selected_ids=None):
    """Widget de seleção de questões com filtros, duas colunas e reordenação.
    Usado tanto em criar quanto editar prova. JS serializa IDs em string CSV no campo 'questoes_serializadas'."""
    if selected_ids is None:
        selected_ids = []
    import json

    questoes_db = conn.execute("""
        SELECT q.id, q.enunciado, q.ano, q.tipo, d.nome AS disciplina_nome
        FROM questoes q JOIN disciplinas d ON d.id = q.disciplina_id
        ORDER BY d.nome, q.id
    """).fetchall()

    bncc_map = {}
    for row in conn.execute("""
        SELECT qh.questao_id, h.codigo
        FROM questao_habilidades qh JOIN habilidades_bncc h ON h.id = qh.habilidade_id
        ORDER BY h.codigo
    """).fetchall():
        bncc_map.setdefault(row["questao_id"], []).append(row["codigo"])

    disciplinas = conn.execute("SELECT * FROM disciplinas ORDER BY nome").fetchall()

    questoes_payload = [
        {
            "id": q["id"],
            "disciplina": q["disciplina_nome"],
            "ano": q["ano"] if q["ano"] else "",
            "enunciado": q["enunciado"],
            "preview": _preview_enunciado(q["enunciado"], max_chars=120),
            "bnccs": bncc_map.get(q["id"], []),
        }
        for q in questoes_db
    ]
    questoes_json = json.dumps(questoes_payload, ensure_ascii=False)
    selected_json = json.dumps(list(selected_ids))

    disciplinas_opts = '<option value="">Todas</option>' + "".join(
        f'<option value="{d["nome"]}">{d["nome"]}</option>' for d in disciplinas
    )
    anos_opts = '<option value="">Todos</option>' + "".join(f'<option value="{a}">{a}</option>' for a in ANOS)

    template = r'''
<input type="hidden" name="questoes_serializadas" id="questoes_serializadas" value="">

<div style="display:grid; grid-template-columns: 1.4fr 1fr; gap:20px; align-items:flex-start;">
    <div>
        <h3 style="margin-top:0;">Questões disponíveis</h3>
        <div style="background:var(--bg-subtle); padding:12px; border-radius:6px; margin-bottom:12px;">
            <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:8px;">
                <label style="margin:0;">Disciplina<select id="filtro-disciplina">__DISC_OPTS__</select></label>
                <label style="margin:0;">Ano<select id="filtro-ano">__ANOS_OPTS__</select></label>
            </div>
            <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:8px; margin-top:8px;">
                <label style="margin:0;">BNCC<input type="text" id="filtro-bncc" placeholder="EF06MA"></label>
                <label style="margin:0;">Buscar<input type="text" id="filtro-q" placeholder="palavra-chave no enunciado"></label>
            </div>
        </div>
        <div id="picker-disponiveis" style="max-height:600px; overflow-y:auto; border:1px solid var(--border); border-radius:6px; padding:8px;"></div>
    </div>
    <div style="position:sticky; top:20px;">
        <h3 style="margin-top:0;">Selecionadas (<span id="picker-counter">0</span>)</h3>
        <div id="picker-selecionadas" style="max-height:600px; overflow-y:auto; border:1px solid var(--border); border-radius:6px; padding:8px;"></div>
    </div>
</div>

<script>
const TODAS_QUESTOES = __QUESTOES_JSON__;
let selecionadas = __SELECTED_JSON__;

function escapeHtml(s) {
    return String(s).replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;').replace(/"/g, '&quot;');
}

function renderPicker() {
    const disc = document.getElementById('filtro-disciplina').value;
    const ano = document.getElementById('filtro-ano').value;
    const bncc = document.getElementById('filtro-bncc').value.trim().toUpperCase();
    const q = document.getElementById('filtro-q').value.trim().toLowerCase();

    const filtradas = TODAS_QUESTOES.filter(function(quest) {
        if (disc && quest.disciplina !== disc) return false;
        if (ano && quest.ano !== ano) return false;
        if (bncc && !quest.bnccs.some(function(c){ return c.includes(bncc); })) return false;
        if (q && !quest.enunciado.toLowerCase().includes(q)) return false;
        return true;
    });

    const dispDiv = document.getElementById('picker-disponiveis');
    if (filtradas.length === 0) {
        dispDiv.innerHTML = '<p style="color:var(--text-muted); padding:12px;">Nenhuma questão com esses filtros.</p>';
    } else {
        dispDiv.innerHTML = filtradas.map(function(quest) {
            const isSelected = selecionadas.includes(quest.id);
            const badgeAno = quest.ano ? ' · ' + escapeHtml(quest.ano) : '';
            const badgesBncc = quest.bnccs.map(function(c){ return '<span class="badge" style="font-size:10px;">' + escapeHtml(c) + '</span>'; }).join(' ');
            if (isSelected) {
                return '<div style="padding:8px 10px; margin-bottom:6px; background:var(--bg-muted); border-radius:4px; opacity:0.6;">' +
                    '<div style="font-size:11px; color:var(--text-muted);">Q' + quest.id + ' · ' + escapeHtml(quest.disciplina) + badgeAno + ' ' + badgesBncc + '</div>' +
                    '<div style="font-size:13px; margin-top:4px;">' + escapeHtml(quest.preview) + '</div>' +
                    '<div style="font-size:11px; color:var(--text-muted); margin-top:6px;">✓ Já adicionada</div>' +
                    '</div>';
            }
            return '<div style="padding:8px 10px; margin-bottom:6px; background:var(--bg); border:1px solid var(--border); border-radius:4px; display:flex; gap:8px; align-items:flex-start;">' +
                '<div style="flex:1; min-width:0;">' +
                '<div style="font-size:11px; color:var(--text-muted);">Q' + quest.id + ' · ' + escapeHtml(quest.disciplina) + badgeAno + ' ' + badgesBncc + '</div>' +
                '<div style="font-size:13px; margin-top:4px;">' + escapeHtml(quest.preview) + '</div>' +
                '</div>' +
                '<button type="button" onclick="adicionar(' + quest.id + ')" class="btn" style="padding:4px 10px; font-size:12px; white-space:nowrap;">+ Adicionar</button>' +
                '</div>';
        }).join('');
    }

    const selDiv = document.getElementById('picker-selecionadas');
    document.getElementById('picker-counter').textContent = selecionadas.length;
    if (selecionadas.length === 0) {
        selDiv.innerHTML = '<p style="color:var(--text-muted); padding:12px; font-size:13px;">Nenhuma questão selecionada ainda. Use o painel à esquerda para adicionar.</p>';
    } else {
        const byId = {};
        TODAS_QUESTOES.forEach(function(q){ byId[q.id] = q; });
        selDiv.innerHTML = selecionadas.map(function(qid, idx) {
            const quest = byId[qid];
            if (!quest) return '';
            const badgeAno = quest.ano ? ' · ' + escapeHtml(quest.ano) : '';
            const upDisabled = idx === 0 ? 'disabled style="opacity:0.3;"' : '';
            const downDisabled = idx === selecionadas.length - 1 ? 'disabled style="opacity:0.3;"' : '';
            return '<div style="padding:8px 10px; margin-bottom:6px; background:var(--bg); border:1px solid var(--border); border-radius:4px; display:flex; gap:8px; align-items:flex-start;">' +
                '<div style="flex:0 0 26px; font-weight:600;">' + (idx + 1) + '.</div>' +
                '<div style="flex:1; min-width:0;">' +
                '<div style="font-size:11px; color:var(--text-muted);">Q' + quest.id + ' · ' + escapeHtml(quest.disciplina) + badgeAno + '</div>' +
                '<div style="font-size:12px; margin-top:2px;">' + escapeHtml(quest.preview.slice(0, 80)) + (quest.preview.length > 80 ? '...' : '') + '</div>' +
                '</div>' +
                '<div style="display:flex; flex-direction:column; gap:2px;">' +
                '<button type="button" onclick="mover(' + idx + ', -1)" ' + upDisabled + ' class="btn" style="padding:0 6px; font-size:11px;">▴</button>' +
                '<button type="button" onclick="mover(' + idx + ', 1)" ' + downDisabled + ' class="btn" style="padding:0 6px; font-size:11px;">▾</button>' +
                '</div>' +
                '<button type="button" onclick="remover(' + quest.id + ')" class="btn" style="padding:4px 8px; font-size:11px; color:var(--red);">✕</button>' +
                '</div>';
        }).join('');
    }

    document.getElementById('questoes_serializadas').value = selecionadas.join(',');
}

function adicionar(id) { if (!selecionadas.includes(id)) selecionadas.push(id); renderPicker(); }
function remover(id) { selecionadas = selecionadas.filter(function(x){ return x !== id; }); renderPicker(); }
function mover(idx, delta) {
    const newIdx = idx + delta;
    if (newIdx < 0 || newIdx >= selecionadas.length) return;
    const tmp = selecionadas[idx];
    selecionadas[idx] = selecionadas[newIdx];
    selecionadas[newIdx] = tmp;
    renderPicker();
}

document.getElementById('filtro-disciplina').addEventListener('change', renderPicker);
document.getElementById('filtro-ano').addEventListener('change', renderPicker);
document.getElementById('filtro-bncc').addEventListener('input', renderPicker);
document.getElementById('filtro-q').addEventListener('input', renderPicker);

renderPicker();
</script>
'''
    return (template
        .replace("__QUESTOES_JSON__", questoes_json)
        .replace("__SELECTED_JSON__", selected_json)
        .replace("__DISC_OPTS__", disciplinas_opts)
        .replace("__ANOS_OPTS__", anos_opts))


@app.get("/provas/nova", response_class=HTMLResponse)
def form_nova_prova():
    conn = get_db()
    n_questoes = conn.execute("SELECT COUNT(*) AS c FROM questoes").fetchone()["c"]
    if n_questoes == 0:
        conn.close()
        return render_page("Nova prova", '<div class="page-header"><h1>Nova prova</h1></div><div class="empty"><p>Você precisa cadastrar questões antes de montar uma prova.</p><a href="/questoes/nova" class="btn btn-primary">Cadastrar questão</a></div>', active="provas")
    picker = _render_picker_questoes(conn, selected_ids=[])
    conn.close()
    content = (
        '<div class="page-header"><h1>Nova prova</h1></div>'
        '<form action="/provas/nova" method="post">'
        '<label>Título<input type="text" name="titulo" required placeholder="Ex: Prova de Matemática — 1º Bimestre — 9º Ano"></label>'
        '<label>Descrição (opcional)<textarea name="descricao" rows="2"></textarea></label>'
        f'{picker}'
        '<div class="page-actions"><button type="submit" class="btn btn-primary">Criar prova</button><a href="/provas" class="btn">Cancelar</a></div>'
        '</form>'
    )
    return render_page("Nova prova", content, active="provas", head_extra=MATHJAX_EDIT)


@app.post("/provas/nova")
def criar_prova(request: Request, titulo: str = Form(...), descricao: str = Form(""), questoes_serializadas: str = Form("")):
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    ids_str = [x.strip() for x in questoes_serializadas.split(",") if x.strip()]
    questao_ids = []
    for s in ids_str:
        try:
            questao_ids.append(int(s))
        except ValueError:
            pass
    if not questao_ids:
        return RedirectResponse("/provas/nova", status_code=303)
    conn = get_db()
    cursor = conn.execute("INSERT INTO provas (titulo, descricao, criada_por_professor_id) VALUES (?, ?, ?)",
                          (titulo.strip(), descricao.strip() or None, prof["id"]))
    prova_id = cursor.lastrowid
    for ordem, qid in enumerate(questao_ids):
        conn.execute("INSERT INTO prova_questoes (prova_id, questao_id, ordem) VALUES (?, ?, ?)", (prova_id, qid, ordem))
    conn.commit()
    conn.close()

    # Buscar turmas para o atalho "Aplicar agora"
    conn2 = get_db()
    turmas = conn2.execute("SELECT id, nome, ano_letivo FROM turmas ORDER BY ano_letivo DESC, nome").fetchall()
    conn2.close()

    if turmas:
        turmas_options = "".join(f'<option value="{t["id"]}">{t["nome"]} ({t["ano_letivo"]})</option>' for t in turmas)
        form_aplicar = f"""
            <form action="/aplicacoes/nova" method="post">
                <input type="hidden" name="prova_id" value="{prova_id}">
                <label style="margin:0 0 10px;">Turma
                    <select name="turma_id" required>{turmas_options}</select>
                </label>
                <label style="font-weight:normal; display:flex; align-items:center; gap:6px; margin-bottom:6px;">
                    <input type="radio" name="modo" value="online" required checked style="width:auto;"> Online
                </label>
                <label style="font-weight:normal; display:flex; align-items:center; gap:6px; margin-bottom:12px;">
                    <input type="radio" name="modo" value="impressa" style="width:auto;"> Impressa
                </label>
                <button type="submit" class="btn btn-primary" style="width:100%;">Criar aplicação</button>
            </form>
        """
    else:
        form_aplicar = '<p style="font-size:13px; color:var(--text-muted); margin:0;">Você ainda não tem nenhuma turma cadastrada. <a href="/turmas/nova">Criar turma</a></p>'

    content_html = f"""
        <div style="max-width:560px; margin:60px auto; text-align:center; padding:0 20px;">
            <div style="font-size:52px; margin-bottom:12px;">🎉</div>
            <h1 style="font-size:22px; margin-bottom:6px;">Atividade criada!</h1>
            <p style="color:var(--text-muted); margin-bottom:32px;">O que deseja fazer agora?</p>
            <div style="display:flex; gap:10px; justify-content:center; flex-wrap:wrap; margin-bottom:32px;">
                <a href="/provas/{prova_id}" class="btn btn-primary">👁️ Ver atividade</a>
                <a href="/provas/nova" class="btn">📝 Criar outra atividade</a>
                <a href="/questoes/nova" class="btn">✏️ Criar questão</a>
            </div>
            <div style="background:var(--bg-subtle); border:1px solid var(--border); border-radius:10px; padding:18px; text-align:left;">
                <p style="font-weight:600; font-size:13px; margin:0 0 10px 0;">🚀 Aplicar agora:</p>
                {form_aplicar}
            </div>
        </div>
    """
    return HTMLResponse(render_page("Atividade criada", content_html, active="provas"))


@app.get("/provas/{prova_id}", response_class=HTMLResponse)
def ver_prova(prova_id: int):
    conn = get_db()
    prova = conn.execute("SELECT * FROM provas WHERE id = ?", (prova_id,)).fetchone()
    if not prova:
        conn.close()
        return HTMLResponse(render_page("Não encontrada", '<h1>Prova | tarefa não encontrada</h1><p><a href="/provas">← Voltar</a></p>', active="provas"), status_code=404)
    questoes = conn.execute("SELECT q.id, q.enunciado, q.ano, d.nome AS disciplina_nome FROM prova_questoes pq JOIN questoes q ON q.id = pq.questao_id JOIN disciplinas d ON d.id = q.disciplina_id WHERE pq.prova_id = ? ORDER BY pq.ordem", (prova_id,)).fetchall()
    n_aplicacoes = conn.execute("SELECT COUNT(*) AS c FROM aplicacoes WHERE prova_id = ?", (prova_id,)).fetchone()["c"]
    questoes_html = "".join(render_questao_card(conn, q, numero=idx) for idx, q in enumerate(questoes, start=1))
    conn.close()
    desc_html = f'<p class="subtitle">{prova["descricao"]}</p>' if prova["descricao"] else ""
    comparativo_btn = ""
    if n_aplicacoes > 0:
        label = "Comparativo entre turmas" if n_aplicacoes >= 2 else "Ver análises pedagógicas"
        comparativo_btn = f'<a href="/provas/{prova_id}/comparativo" class="btn">📊 {label} ({n_aplicacoes} aplicação{"" if n_aplicacoes == 1 else "ões"})</a>'
    prof_ctx = _current_prof_ctx.get()
    status_rev = prova["status_revisao"] if "status_revisao" in prova.keys() else "rascunho"
    eh_dono = prof_ctx and (prova["criada_por_professor_id"] == prof_ctx["id"] or prof_ctx.get("is_admin"))
    eh_gestor_ou_admin = prof_ctx and (prof_ctx.get("is_admin") or prof_ctx.get("is_gestor"))
    status_badge_html = _status_badge_html(status_rev)
    obs_html = ""
    if "obs_gestao" in prova.keys() and prova["obs_gestao"]:
        obs_html = f'<div style="background:var(--orange-bg); border-left:3px solid var(--orange); padding:8px 12px; border-radius:6px; margin-top:8px; font-size:13px;"><strong>Obs. da gestão:</strong> {prova["obs_gestao"]}</div>'
    submeter_btn = ""
    if eh_dono and status_rev in ("rascunho", "devolvida"):
        submeter_btn = f'<form method="post" action="/provas/{prova_id}/submeter" style="margin:0;" onsubmit="return confirm(\'Submeter para revisão da gestão?\')"><button type="submit" class="btn btn-primary" style="background:var(--orange); border-color:var(--orange);">📤 Submeter para revisão</button></form>'
    imprimir_btn = f'<a href="/provas/{prova_id}/imprimir" class="btn btn-primary" target="_blank">🖨️ Imprimir prova</a>' if (status_rev == "aprovada" or eh_gestor_ou_admin) else ""
    acoes_html = f'<div class="page-actions" style="display:flex; gap:8px; flex-wrap:wrap; align-items:center;">{imprimir_btn}{comparativo_btn}{submeter_btn}{status_badge_html}{obs_html}</div>'
    content = f'<div class="page-header"><h1>{prova["titulo"]}</h1><p class="subtitle">{len(questoes)} questões</p>{desc_html}{acoes_html}</div><hr>{questoes_html}'
    return render_page(prova["titulo"], content, active="provas", head_extra=MATHJAX)


@app.get("/provas/{id}/editar", response_class=HTMLResponse)
def form_editar_prova(id: int):
    conn = get_db()
    prova = conn.execute("SELECT * FROM provas WHERE id = ?", (id,)).fetchone()
    if not prova:
        conn.close()
        return RedirectResponse("/provas", status_code=303)
    
    todas_questoes = conn.execute("""
        SELECT q.id, q.enunciado, q.ano, q.tipo, d.nome AS disciplina_nome 
        FROM questoes q 
        JOIN disciplinas d ON d.id = q.disciplina_id 
        ORDER BY d.nome, q.id
    """).fetchall()
    
    selecionadas = [r["questao_id"] for r in conn.execute("SELECT questao_id FROM prova_questoes WHERE prova_id = ?", (id,)).fetchall()]
    conn.close()
    
    questoes_html = ""
    for q in todas_questoes:
        checked = " checked" if q["id"] in selecionadas else ""
        resumo_enunciado = q["enunciado"][:110] + "..." if len(q["enunciado"]) > 110 else q["enunciado"]
        questoes_html += f"""
        <div style="margin-bottom:10px; display:flex; align-items:flex-start; gap:10px;">
            <input type="checkbox" name="questoes_ids" value="{q["id"]}"{checked} id="q_{q["id"]}" style="width:auto; margin-top:4px;">
            <label for="q_{q["id"]}" style="font-weight:normal; margin:0; cursor:pointer;">
                <span class="badge" style="margin-right:4px;">{q["disciplina_nome"]}</span> 
                <strong>(ID #{q["id"]})</strong> - {resumo_enunciado}
            </label>
        </div>
        """

    content = f"""
    <div class="page-header"><h1>Editar Prova: {prova["titulo"]}</h1></div>
    <form action="/provas/{id}/editar" method="post">
        <label>Título da Prova
            <input type="text" name="titulo" value="{prova["titulo"]}" required>
        </label>
        
        <fieldset style="margin-top:20px;">
            <legend>Selecione as Questões Integrantes</legend>
            {questoes_html if questoes_html else '<p class="empty">Nenhuma questão encontrada para vincular.</p>'}
        </fieldset>
        
        <div class="page-actions" style="margin-top:20px;">
            <button type="submit" class="btn btn-primary">Salvar Alterações</button>
            <a href="/provas" class="btn">Cancelar</a>
        </div>
    </form>
    """
    return render_page("Editar Prova", content, active="provas")


@app.post("/provas/{id}/editar")
def atualizar_prova(id: int, titulo: str = Form(...), questoes_ids: List[int] = Form([])):
    conn = get_db()
    conn.execute("UPDATE provas SET titulo = ? WHERE id = ?", (titulo.strip(), id))
    conn.execute("DELETE FROM prova_questoes WHERE prova_id = ?", (id,))
    
    for idx, q_id in enumerate(questoes_ids):
        conn.execute("INSERT INTO prova_questoes (prova_id, questao_id, ordem) VALUES (?, ?, ?)", (id, q_id, idx))
        
    conn.commit()
    conn.close()
    return RedirectResponse("/provas", status_code=303)


@app.post("/provas/{id}/deletar", response_class=HTMLResponse)
def deletar_prova(id: int):
    conn = get_db()
    uso_ativo = conn.execute("SELECT id FROM aplicacoes WHERE prova_id = ?", (id,)).fetchone()
    if uso_ativo:
        conn.close()
        content = """
        <div style="border: 1px solid var(--red); background: var(--red-bg); padding: 20px; border-radius: 6px; margin-top:20px; color:var(--red);">
            <h3 style="color:var(--red); margin-top:0;">Operação Impedida</h3>
            <p>Não é possível deletar esta prova | tarefa porque ela possui <strong>Aplicações</strong> em andamento ou histórico de notas associado a turmas.</p>
            <p>Se deseja realmente excluí-la, remova primeiro as respectivas aplicações na aba de "Aplicações".</p>
            <a href="/provas" class="btn" style="margin-top:10px;">Voltar para Provas</a>
        </div>
        """
        return render_page("Erro ao Deletar Prova", content, active="provas")
        
    conn.execute("DELETE FROM prova_questoes WHERE prova_id = ?", (id,))
    conn.execute("DELETE FROM provas WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/provas", status_code=303)


# ==========================================
#  ROTAS DE TURMAS
# ==========================================

@app.get("/turmas", response_class=HTMLResponse)
def listar_turmas(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    prof = get_current_professor(request)
    conn = get_db()
    turmas = conn.execute("SELECT t.id, t.nome, t.ano_letivo, COUNT(a.id) AS total_alunos FROM turmas t LEFT JOIN alunos a ON a.turma_id = t.id GROUP BY t.id ORDER BY t.ano_letivo DESC, t.nome").fetchall()
    conn.close()
    if turmas:
        cards = "".join(f'<a href="/turmas/{t["id"]}" class="card card-link"><div class="card-title">{t["nome"]}</div><div class="card-meta">Ano letivo {t["ano_letivo"]} · {t["total_alunos"]} alunos</div></a>' for t in turmas)
    else:
        cards = '<div class="empty">Nenhuma turma cadastrada ainda.</div>'
    botoes_admin = (
        '<div class="page-actions"><a href="/turmas/nova" class="btn btn-primary">+ Nova turma</a><a href="/turmas/importar" class="btn">Importar planilha</a></div>'
        if prof and prof["is_admin"] else
        '<p class="muted-line" style="font-size:13px;">As turmas são gerenciadas pelo administrador da escola.</p>'
    )
    content = f'<div class="page-header"><h1>Turmas</h1>{botoes_admin}</div>{cards}'
    return render_page("Turmas", content, active="turmas")


@app.get("/turmas/nova", response_class=HTMLResponse)
def form_nova_turma(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    content = '<div class="page-header"><h1>Nova turma</h1></div><form action="/turmas/nova" method="post"><label>Nome<input type="text" name="nome" required placeholder="Ex: 9º Ano A" autofocus></label><label>Ano letivo<input type="number" name="ano_letivo" required value="2026" min="2020" max="2099"></label><div class="page-actions"><button type="submit" class="btn btn-primary">Cadastrar</button><a href="/turmas" class="btn">Cancelar</a></div></form>'
    return render_page("Nova turma", content, active="turmas")


@app.post("/turmas/nova")
def criar_turma(request: Request, nome: str = Form(...), ano_letivo: int = Form(...)):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()
    cursor = conn.execute("INSERT INTO turmas (nome, ano_letivo) VALUES (?, ?)", (nome.strip(), ano_letivo))
    turma_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return RedirectResponse(f"/turmas/{turma_id}", status_code=303)

@app.get("/turmas/template")
def baixar_template_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"
    headers = ["Turma", "Ano Letivo", "Número", "Nome", "Raça", "E-mail", "Data de Nascimento"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.append(["9º Ano A", 2026, 1, "Maria da Silva", "Parda", "maria@escola.com", "2010-03-15"])
    ws.append(["9º Ano A", 2026, 2, "João Santos", "Branca", "joao@escola.com", "2010-07-22"])
    ws.append(["9º Ano B", 2026, 1, "Ana Pereira", "", "", ""])

    larguras = [16, 12, 10, 28, 12, 26, 22]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_alunos.xlsx"},
    )


@app.get("/turmas/importar", response_class=HTMLResponse)
def form_importar_excel(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    content = """
        <div class="page-header">
            <h1>Importar planilha</h1>
            <p class="subtitle">Cadastre várias turmas e alunos de uma vez subindo um arquivo Excel.</p>
        </div>
        <div class="tip">
            <strong>Como funciona:</strong> a planilha deve ter as colunas <code>Turma</code>, <code>Ano Letivo</code>, <code>Número</code>, <code>Nome</code>, <code>Raça</code>, <code>E-mail</code>, e <code>Data de Nascimento</code> (na primeira linha como cabeçalho). Cada linha seguinte é um aluno. Turmas que ainda não existem são criadas automaticamente. Alunos com nome já cadastrado na mesma turma são pulados (evita duplicação se você importar a planilha duas vezes).
        </div>

        <h2>1. Baixar template</h2>
        <p>Se você ainda não tem a planilha, baixa um modelo pronto com a estrutura certa:</p>
        <p><a href="/turmas/template" class="btn">Baixar template Excel</a></p>

        <h2>2. Subir planilha preenchida</h2>
        <form action="/turmas/importar" method="post" enctype="multipart/form-data">
            <label>
                Arquivo .xlsx
                <input type="file" name="arquivo" accept=".xlsx" required>
            </label>
            <div class="page-actions">
                <button type="submit" class="btn btn-primary">Importar</button>
                <a href="/turmas" class="btn">Cancelar</a>
            </div>
        </form>
    """
    return render_page("Importar planilha", content, active="turmas")


@app.post("/turmas/importar", response_class=HTMLResponse)
async def importar_excel(request: Request, arquivo: UploadFile = File(...)):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    if not arquivo.filename.lower().endswith(".xlsx"):
        content = """
            <div class="page-header"><h1>Erro na importação</h1></div>
            <div class="tip">O arquivo precisa ser .xlsx (Excel moderno).</div>
            <p><a href="/turmas/importar" class="btn">Voltar</a></p>
        """
        return HTMLResponse(render_page("Erro", content, active="turmas"))

    content_bytes = await arquivo.read()
    try:
        wb = load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        content = f"""
            <div class="page-header"><h1>Erro ao ler a planilha</h1></div>
            <div class="tip">{str(e)}</div>
            <p><a href="/turmas/importar" class="btn">Voltar</a></p>
        """
        return HTMLResponse(render_page("Erro", content, active="turmas"))

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        content = """
            <div class="page-header"><h1>Planilha vazia</h1></div>
            <p>A planilha não contém dados além do cabeçalho.</p>
            <p><a href="/turmas/importar" class="btn">Voltar</a></p>
        """
        return HTMLResponse(render_page("Vazia", content, active="turmas"))

    conn = get_db()
    turmas_criadas = 0
    alunos_criados = 0
    alunos_pulados = 0
    avisos = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        row = list(row) + [None] * (7 - len(row))
        turma_nome, ano_letivo_raw, numero_raw, nome, raca, email, data_nasc_raw = row[:7]

        if not turma_nome or not str(turma_nome).strip():
            avisos.append(f"Linha {row_num}: turma vazia, ignorada")
            continue
        if not nome or not str(nome).strip():
            avisos.append(f"Linha {row_num}: nome vazio, ignorada")
            continue

        try:
            ano_letivo = int(ano_letivo_raw) if ano_letivo_raw else 2026
        except (TypeError, ValueError):
            avisos.append(f"Linha {row_num}: ano letivo inválido ({ano_letivo_raw}), ignorada")
            continue

        turma_nome_clean = str(turma_nome).strip()
        nome_clean = str(nome).strip()

        turma = conn.execute(
            "SELECT id FROM turmas WHERE nome = ? AND ano_letivo = ?",
            (turma_nome_clean, ano_letivo),
        ).fetchone()

        if turma:
            turma_id = turma["id"]
        else:
            cursor = conn.execute(
                "INSERT INTO turmas (nome, ano_letivo) VALUES (?, ?)",
                (turma_nome_clean, ano_letivo),
            )
            turma_id = cursor.lastrowid
            turmas_criadas += 1

        existing = conn.execute(
            "SELECT id FROM alunos WHERE turma_id = ? AND LOWER(nome) = LOWER(?)",
            (turma_id, nome_clean),
        ).fetchone()
        if existing:
            alunos_pulados += 1
            continue

        try:
            numero = int(numero_raw) if numero_raw else None
        except (TypeError, ValueError):
            numero = None

        raca_clean = str(raca).strip() if raca else None
        email_clean = str(email).strip() if email else None

        data_nasc_str = None
        if data_nasc_raw:
            if isinstance(data_nasc_raw, (date, datetime)):
                data_nasc_str = data_nasc_raw.isoformat()[:10]
            else:
                s = str(data_nasc_raw).strip()
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
                    try:
                        data_nasc_str = datetime.strptime(s, fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
                if not data_nasc_str:
                    avisos.append(f"Linha {row_num}: data '{s}' não reconhecida, gravado vazio")

        codigo = gerar_codigo_aluno(conn)
        conn.execute(
            "INSERT INTO alunos (turma_id, nome, numero, codigo_unico, raca, email, data_nascimento) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (turma_id, nome_clean, numero, codigo, raca_clean, email_clean, data_nasc_str),
        )
        alunos_criados += 1

    conn.commit()
    conn.close()

    avisos_html = ""
    if avisos:
        items = "".join(f"<li>{a}</li>" for a in avisos)
        avisos_html = f'<h2>Avisos</h2><ul class="clean">{items}</ul>'

    content = f"""
        <div class="page-header">
            <h1>Importação concluída</h1>
            <p class="subtitle">Resumo do que foi processado.</p>
        </div>
        <div class="metric-grid">
            <div class="metric"><div class="metric-label">Turmas criadas</div><div class="metric-value">{turmas_criadas}</div></div>
            <div class="metric"><div class="metric-label">Alunos criados</div><div class="metric-value">{alunos_criados}</div></div>
            <div class="metric"><div class="metric-label">Alunos pulados</div><div class="metric-value">{alunos_pulados}</div></div>
            <div class="metric"><div class="metric-label">Avisos</div><div class="metric-value">{len(avisos)}</div></div>
        </div>
        {avisos_html}
        <div class="page-actions">
            <a href="/turmas" class="btn btn-primary">Ver turmas</a>
            <a href="/turmas/importar" class="btn">Importar outra</a>
        </div>
    """
    return HTMLResponse(render_page("Importação concluída", content, active="turmas"))


# ==========================================
#  BOLETIM / CONSELHO DE CLASSE
# ==========================================

BOLETIM_SUBJECT_MAP = {
    'educacao fisica': 'Ed. Física', 'ed fisica': 'Ed. Física', 'ed. fisica': 'Ed. Física',
    'ed digital': 'Educação Digital', 'educacao digital': 'Educação Digital', 'ed. digital': 'Educação Digital',
    'ciencias': 'Ciências', 'historia': 'História', 'geografia': 'Geografia',
    'ingles': 'Inglês', 'portugues': 'Português', 'matematica': 'Matemática', 'arte': 'Arte',
}
BOLETIM_CANONICAL_SUBJECTS = ['Português', 'Matemática', 'Ciências', 'História', 'Geografia',
                              'Inglês', 'Arte', 'Ed. Física', 'Educação Digital', 'Geral']

# Mapeamento das siglas do export oficial do e-cidade pra disciplina canônica.
# Posições fixas no CSV (24/08/2026): colunas exportadas sempre nesta ordem.
ECIDADE_COLUNAS_DISCIPLINA = [
    ("LPOR", "Português"), ("EF", "Ed. Física"), ("A", "Arte"), ("MAT", "Matemática"),
    ("CIE", "Ciências"), ("HIS", "História"), ("GEO", "Geografia"), ("LI", "Inglês"),
    ("ED", "Educação Digital"),
]
ECIDADE_CAMPOS_MINIMOS = 4 + 3 + len(ECIDADE_COLUNAS_DISCIPLINA) * 2 + 1  # Nº,Nome,S,Código + 3 par + 9×(nota,falta) + TF


def _ecidade_normalizar(s):
    """Normalização pra comparação de PREFIXO de nome (o e-cidade trunca nomes em ~20
    caracteres, e às vezes o corte cai no meio de uma palavra deixando um ponto solto,
    tipo 'DA S.' em vez de 'DA SILVA' — removemos os pontos pra não travar esse caso)."""
    return _boletim_normalizar(s).replace(".", "").strip()


ECIDADE_CONECTIVOS_NOME = {"de", "da", "do", "dos", "das", "e"}


def _ecidade_nome_reduzido(nome_completo):
    """Reduz um nome completo (como cadastrado no sistema) pro MESMO formato que o
    e-cidade usa na exportação: primeiro e último nome por extenso, conectivos
    (de/da/do/dos/das/e) por extenso, e qualquer outro nome do meio vira só a inicial.
    'Alice Bastos de Oliveira' -> 'alice b de oliveira'
    'Analice de Jesus da Silva Machado' -> 'analice de j da s machado'
    Descoberto em 24/08/2026 comparando a exportação real com o cadastro real: o
    casamento por prefixo simples falhava porque o e-cidade abrevia nomes do meio pra
    uma letra e o sistema guarda o nome completo — sem essa redução, quase nada batia."""
    tokens = _ecidade_normalizar(nome_completo).split()
    if len(tokens) <= 2:
        return " ".join(tokens)
    meio_reduzido = [t if t in ECIDADE_CONECTIVOS_NOME else t[0] for t in tokens[1:-1]]
    return " ".join([tokens[0]] + meio_reduzido + [tokens[-1]])


def _ecidade_nome_bate(nome_sistema, nome_csv):
    """True se o nome do sistema (completo) é compatível com o nome truncado/abreviado
    do e-cidade — comparando a forma REDUZIDA do nome do sistema com o nome do CSV,
    tolerando o corte de ~20 caracteres do e-cidade (prefixo em qualquer direção como
    reforço, pro caso de nomes que não seguem exatamente o padrão)."""
    reduzido = _ecidade_nome_reduzido(nome_sistema)
    csv_norm = _ecidade_normalizar(nome_csv)
    if reduzido.startswith(csv_norm) or csv_norm.startswith(reduzido):
        return True
    bruto = _ecidade_normalizar(nome_sistema)
    return bruto.startswith(csv_norm) or csv_norm.startswith(bruto)


def _ecidade_parse_csv(content_bytes):
    """Faz o parse do CSV exportado do Conselho de Classe do e-cidade (24/08/2026).
    Retorna dict: {turma_detectada, alunos: [{nº, nome, codigo_rede, notas: {disciplina: valor_ou_None},
    faltas: {disciplina: int_ou_None}, tf, pd: bool}], avisos: [str]}.

    Formato: ';'-separado, Latin-1. 4 linhas de metadado, 1 linha de cabeçalho, depois 1 linha
    por aluno com campos fixos (Nº;Nome;S;Código;par;par;par;LPOR nota;falta;EF nota;falta;...;TF).

    ARMADILHA CONHECIDA: alunos com parecer descritivo (currículo adaptado / 'PD') têm o parecer
    inteiro despejado SEM ESCAPE no meio do CSV, ocupando várias linhas e quebrando a contagem de
    campos. Detectamos isso pela linha ter menos campos que o esperado e tratamos tudo até a
    próxima linha válida como parecer (descartado, só um aviso é gerado — decisão de 24/08/2026)."""
    import re as _re_ecid

    try:
        texto = content_bytes.decode("latin-1")
    except Exception:
        texto = content_bytes.decode("utf-8", errors="replace")

    linhas = texto.split("\n")
    avisos = []

    # Turma vem de uma linha de metadado tipo "...;;;;;Turma: EF 601"
    turma_detectada = None
    for l in linhas[:6]:
        m = _re_ecid.search(r"Turma:\s*(.+?)\s*(?:;|$)", l)
        if m:
            bruto = m.group(1).strip()
            # "EF 601" -> pega o último grupo de dígitos (o código real da turma)
            m2 = _re_ecid.search(r"(\d+)\s*$", bruto)
            turma_detectada = m2.group(1) if m2 else bruto
            break

    alunos = []
    i = 0
    while i < len(linhas):
        linha = linhas[i].rstrip("\r")
        i += 1
        if not linha.strip():
            continue
        if linha.strip() in ("<script>", "</script>") or "window.close()" in linha:
            continue
        campos = linha.split(";")
        # Linha de cabeçalho ou metadado institucional — pula
        if not campos[0].strip().isdigit():
            continue

        if len(campos) >= ECIDADE_CAMPOS_MINIMOS:
            # Linha completa e bem-formada
            num_aluno = campos[0].strip()
            nome = campos[1].strip()
            codigo_rede = campos[3].strip()
            notas = {}
            faltas = {}
            idx = 7  # primeira coluna de nota é a posição 7 (0-indexed)
            for sigla, canon in ECIDADE_COLUNAS_DISCIPLINA:
                nota_raw = campos[idx].strip() if idx < len(campos) else ""
                falta_raw = campos[idx + 1].strip() if idx + 1 < len(campos) else ""
                if nota_raw:
                    try:
                        notas[canon] = float(nota_raw.replace(",", "."))
                    except ValueError:
                        notas[canon] = nota_raw  # categórico (PS/PA/PI/PD solto)
                else:
                    notas[canon] = None
                try:
                    faltas[canon] = int(falta_raw) if falta_raw else 0
                except ValueError:
                    faltas[canon] = None
                idx += 2
            tf_raw = campos[idx].strip() if idx < len(campos) else ""
            try:
                tf = int(tf_raw) if tf_raw else None
            except ValueError:
                tf = None
            alunos.append({
                "num": num_aluno, "nome": nome, "codigo_rede": codigo_rede,
                "notas": notas, "faltas": faltas, "tf": tf, "pd": False,
            })
        else:
            # Linha malformada — início de um bloco de PARECER (aluno PD). Recupera o que dá
            # (Nº, Nome, Código, se vieram antes do texto livre começar) e consome as linhas
            # seguintes até achar a próxima linha válida, sem tentar interpretá-las.
            num_aluno = campos[0].strip()
            nome = campos[1].strip() if len(campos) > 1 else "(nome não recuperado)"
            codigo_rede = campos[3].strip() if len(campos) > 3 else ""
            n_linhas_parecer = 1
            while i < len(linhas):
                prox = linhas[i].rstrip("\r")
                prox_campos = prox.split(";")
                if prox_campos[0].strip().isdigit() and len(prox_campos) >= ECIDADE_CAMPOS_MINIMOS:
                    break  # achou a próxima linha de aluno válida — não consome
                i += 1
                n_linhas_parecer += 1
            avisos.append(
                f"Aluno {num_aluno} ({nome}) tem parecer descritivo no lugar das notas — "
                f"gravado como 'PD' em todas as disciplinas; o texto do parecer ({n_linhas_parecer} "
                f"linha(s)) NÃO foi importado (não há campo pra isso no sistema hoje)."
            )
            notas = {canon: "PD" for _, canon in ECIDADE_COLUNAS_DISCIPLINA}
            faltas = {canon: None for _, canon in ECIDADE_COLUNAS_DISCIPLINA}
            alunos.append({
                "num": num_aluno, "nome": nome, "codigo_rede": codigo_rede,
                "notas": notas, "faltas": faltas, "tf": None, "pd": True,
            })

    return {"turma_detectada": turma_detectada, "alunos": alunos, "avisos": avisos}


def _boletim_normalizar(s):
    """Remove acentos, baixa caixa, colapsa espaços — mesma ideia do clean() do
    Google Apps Script original, pra casar nomes mesmo com pequenas diferenças
    de acentuação/espaçamento entre a planilha e o banco."""
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def _boletim_normalizar_disciplina(nome):
    return BOLETIM_SUBJECT_MAP.get(_boletim_normalizar(nome), str(nome or "").strip())


@app.get("/boletim/importar-ecidade", response_class=HTMLResponse)
def form_importar_ecidade(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()
    turmas = conn.execute("SELECT id, nome, ano_letivo FROM turmas ORDER BY ano_letivo DESC, nome").fetchall()
    conn.close()
    opts_turmas = "".join(f'<option value="{t["id"]}">{t["nome"]} ({t["ano_letivo"]})</option>' for t in turmas)
    content = f"""
        <div class="page-header">
            <h1>📥 Importar CSV do e-cidade (Conselho de Classe)</h1>
            <p class="subtitle">Sobe o CSV exportado direto da plataforma oficial da rede (e-cidade) — o mesmo arquivo que vem com as notas lançadas pelos professores.</p>
        </div>
        <div class="tip">
            <strong>Como funciona:</strong> o arquivo já traz a turma no cabeçalho — o sistema tenta detectar
            automaticamente, mas confirme abaixo antes de importar. Os alunos são casados com os que
            <strong>já existem</strong> no sistema — nenhum aluno novo é criado aqui. Na primeira importação de
            cada turma, o casamento é feito por nome (o e-cidade trunca nomes longos, então usamos
            correspondência por início do nome); o código oficial do aluno é gravado automaticamente pra que as
            próximas importações dessa turma casem direto por ele, sem depender do nome de novo. Alunos com
            parecer descritivo (currículo adaptado) entram como "PD" em todas as disciplinas — o texto do
            parecer não é importado. Rodar de novo pra mesma turma/trimestre <strong>atualiza</strong> os
            valores (não duplica).
        </div>
        <form action="/boletim/importar-ecidade" method="post" enctype="multipart/form-data">
            <div style="display:flex; flex-wrap:wrap; gap:14px; align-items:flex-end;">
                <label style="margin:0; flex:1 1 160px;">Trimestre
                    <select name="trimestre" required>
                        <option value="1">1º Trimestre</option>
                        <option value="2">2º Trimestre</option>
                        <option value="3">3º Trimestre</option>
                    </select>
                </label>
                <label style="margin:0; flex:1 1 120px;">Ano
                    <input type="number" name="ano" value="2026" required>
                </label>
                <label style="margin:0; flex:1 1 220px;">Turma
                    <select name="turma_id" required>
                        <option value="">— selecione —</option>
                        {opts_turmas}
                    </select>
                </label>
                <label style="margin:0; flex:1 1 220px;">Arquivo CSV do e-cidade
                    <input type="file" name="arquivo" accept=".csv" required>
                </label>
            </div>
            <div class="page-actions">
                <button type="submit" class="btn btn-primary">Importar</button>
                <a href="/painel-gestao" class="btn">Cancelar</a>
            </div>
        </form>
    """
    return render_page("Importar e-cidade", content, active="boletim-importar")


@app.post("/boletim/importar-ecidade", response_class=HTMLResponse)
async def importar_ecidade(request: Request, trimestre: int = Form(...), ano: int = Form(...), turma_id: int = Form(...), arquivo: UploadFile = File(...)):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    if not arquivo.filename.lower().endswith(".csv"):
        content = '<div class="page-header"><h1>Erro</h1></div><div class="tip">O arquivo precisa ser .csv (o export direto do e-cidade).</div><p><a href="/boletim/importar-ecidade" class="btn">Voltar</a></p>'
        return HTMLResponse(render_page("Erro", content, active=""))

    content_bytes = await arquivo.read()
    resultado = _ecidade_parse_csv(content_bytes)
    avisos = list(resultado["avisos"])

    conn = get_db()
    turma = conn.execute("SELECT * FROM turmas WHERE id = ?", (turma_id,)).fetchone()
    if not turma:
        conn.close()
        return RedirectResponse("/boletim/importar-ecidade", status_code=303)

    if resultado["turma_detectada"] and str(resultado["turma_detectada"]) != str(turma["nome"]):
        avisos.append(
            f"O arquivo indica turma '{resultado['turma_detectada']}', mas você selecionou "
            f"'{turma['nome']}'. Conferido? Segui com a turma que você selecionou."
        )

    # Garante que as 9 disciplinas canônicas existem
    disc_ids = {}
    for s in BOLETIM_CANONICAL_SUBJECTS:
        row = conn.execute("SELECT id FROM disciplinas WHERE nome = ?", (s,)).fetchone()
        if row:
            disc_ids[s] = row["id"]
        else:
            cur = conn.execute("INSERT INTO disciplinas (nome) VALUES (?)", (s,))
            disc_ids[s] = cur.lastrowid

    # Alunos já cadastrados NESSA turma — casamento por código da rede primeiro, nome-reduzido depois
    alunos_turma = conn.execute("SELECT id, nome, codigo_rede FROM alunos WHERE turma_id = ?", (turma_id,)).fetchall()
    por_codigo_rede = {a["codigo_rede"]: a["id"] for a in alunos_turma if a["codigo_rede"]}
    candidatos_nome = [(a["id"], a["nome"]) for a in alunos_turma]

    n_medias = n_faltas = n_pd = n_codigo_gravado = 0
    nao_casados = []

    for aluno_csv in resultado["alunos"]:
        aid = por_codigo_rede.get(aluno_csv["codigo_rede"])
        if not aid:
            # Casamento por nome reduzido (o e-cidade abrevia nomes do meio pra uma letra
            # e ainda trunca a string toda em ~20 caracteres — ver _ecidade_nome_bate)
            matches = [cid for cid, nome_sistema in candidatos_nome if _ecidade_nome_bate(nome_sistema, aluno_csv["nome"])]
            if len(matches) == 1:
                aid = matches[0]
                # Grava o código da rede pra próxima importação casar direto
                if aluno_csv["codigo_rede"]:
                    conn.execute("UPDATE alunos SET codigo_rede = ? WHERE id = ?", (aluno_csv["codigo_rede"], aid))
                    n_codigo_gravado += 1
            elif len(matches) > 1:
                nao_casados.append(f"{aluno_csv['nome']} (nº{aluno_csv['num']}) — {len(matches)} alunos da turma batem com esse nome, ambíguo, pulado")
                continue
            else:
                nao_casados.append(f"{aluno_csv['nome']} (nº{aluno_csv['num']}) — nenhum aluno da turma encontrado com esse nome, pulado")
                continue

        if aluno_csv["pd"]:
            n_pd += 1

        for disciplina, valor in aluno_csv["notas"].items():
            did = disc_ids.get(disciplina)
            if not did or valor is None:
                continue
            if isinstance(valor, (int, float)):
                conn.execute("""INSERT INTO boletim_medias (aluno_id, disciplina_id, trimestre, ano, nota, nota_texto)
                                 VALUES (?,?,?,?,?,NULL)
                                 ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET nota = excluded.nota, nota_texto = NULL""",
                             (aid, did, trimestre, ano, float(valor)))
            else:
                conn.execute("""INSERT INTO boletim_medias (aluno_id, disciplina_id, trimestre, ano, nota, nota_texto)
                                 VALUES (?,?,?,?,NULL,?)
                                 ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET nota = NULL, nota_texto = excluded.nota_texto""",
                             (aid, did, trimestre, ano, str(valor).strip()))
            n_medias += 1

        for disciplina, faltas_val in aluno_csv["faltas"].items():
            did = disc_ids.get(disciplina)
            if not did or faltas_val is None:
                continue
            conn.execute("""INSERT INTO boletim_faltas (aluno_id, disciplina_id, trimestre, ano, faltas)
                             VALUES (?,?,?,?,?)
                             ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET faltas = excluded.faltas""",
                         (aid, did, trimestre, ano, faltas_val))
            n_faltas += 1

    conn.commit()
    conn.close()

    if nao_casados:
        avisos.append(f"{len(nao_casados)} aluno(s) do arquivo não foram casados com ninguém da turma:")
        avisos.extend(nao_casados)

    resumo_html = f"""
        <li>{len(resultado["alunos"])} aluno(s) lidos do arquivo (turma detectada no arquivo: {resultado["turma_detectada"] or "não identificada"})</li>
        <li>{n_medias} notas gravadas/atualizadas</li>
        <li>{n_faltas} registros de falta gravados/atualizados</li>
        <li>{n_codigo_gravado} aluno(s) tiveram o código da rede gravado agora (próximas importações casam direto por ele)</li>
        <li>{n_pd} aluno(s) com parecer descritivo (PD) — gravados como PD em todas as disciplinas, texto do parecer não importado</li>
    """
    avisos_html = ""
    if avisos:
        itens = "".join(f'<li>{a}</li>' for a in avisos)
        avisos_html = f'<div class="tip" style="background:var(--orange-bg); border-color:var(--orange); margin-top:14px;"><strong>Avisos:</strong><ul style="margin:8px 0 0 18px;">{itens}</ul></div>'

    content = f"""
        <div class="page-header">
            <h1>✅ Importação e-cidade concluída — {trimestre}º Trimestre {ano} · Turma {turma["nome"]}</h1>
        </div>
        <ul style="line-height:1.9;">{resumo_html}</ul>
        {avisos_html}
        <div class="page-actions" style="margin-top:18px;">
            <a href="/boletim/importar-ecidade" class="btn">Importar outro arquivo</a>
            <a href="/painel-gestao" class="btn btn-primary">Voltar ao painel</a>
        </div>
    """
    return HTMLResponse(render_page("Importação concluída", content, active="boletim-importar"))


@app.get("/boletim/importar", response_class=HTMLResponse)
def form_importar_boletim(request: Request):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    content = """
        <div class="page-header">
            <h1>📊 Importar dados do Boletim / Conselho de Classe</h1>
            <p class="subtitle">Sobe a planilha exportada do Conselho de Classe (Google Sheets) e grava as notas, faltas, raça/etnia e observações dos professores pro trimestre selecionado.</p>
        </div>
        <div class="tip">
            <strong>Como funciona:</strong> a planilha precisa ter as abas <code>Estudantes</code>, <code>RacaEtnia</code>,
            <code>Faltas</code>, <code>Medias</code>, <code>Analise</code> (os nomes exatos, sem acento). Os alunos são
            casados com os que <strong>já existem</strong> no sistema (por turma + nome, ignorando acentuação) — nenhum
            aluno novo é criado aqui. Se algum não casar, ele aparece na lista de avisos no final, sem interromper o
            restante da importação. Rodar a importação de novo com a mesma planilha/trimestre <strong>atualiza</strong>
            os valores anteriores (não duplica).
        </div>
        <form action="/boletim/importar" method="post" enctype="multipart/form-data">
            <div style="display:flex; flex-wrap:wrap; gap:14px; align-items:flex-end;">
                <label style="margin:0; flex:1 1 160px;">Trimestre
                    <select name="trimestre" required>
                        <option value="1">1º Trimestre</option>
                        <option value="2">2º Trimestre</option>
                        <option value="3">3º Trimestre</option>
                    </select>
                </label>
                <label style="margin:0; flex:1 1 120px;">Ano
                    <input type="number" name="ano" value="2026" required>
                </label>
                <label style="margin:0; flex:1 1 220px;">Planilha (.xlsx)
                    <input type="file" name="arquivo" accept=".xlsx" required>
                </label>
            </div>
            <div class="page-actions">
                <button type="submit" class="btn btn-primary">Importar</button>
                <a href="/painel-gestao" class="btn">Cancelar</a>
            </div>
        </form>
    """
    return render_page("Importar Boletim", content, active="boletim-importar")


@app.post("/boletim/importar", response_class=HTMLResponse)
async def importar_boletim(request: Request, trimestre: int = Form(...), ano: int = Form(...), arquivo: UploadFile = File(...)):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    if not arquivo.filename.lower().endswith(".xlsx"):
        content = '<div class="page-header"><h1>Erro</h1></div><div class="tip">O arquivo precisa ser .xlsx.</div><p><a href="/boletim/importar" class="btn">Voltar</a></p>'
        return HTMLResponse(render_page("Erro", content, active=""))

    content_bytes = await arquivo.read()
    try:
        wb = load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
    except Exception as e:
        content = f'<div class="page-header"><h1>Erro ao ler a planilha</h1></div><div class="tip">{str(e)}</div><p><a href="/boletim/importar" class="btn">Voltar</a></p>'
        return HTMLResponse(render_page("Erro", content, active=""))

    abas_esperadas = ["Estudantes", "RacaEtnia", "Faltas", "Medias", "Analise"]
    faltando_abas = [a for a in abas_esperadas if a not in wb.sheetnames]
    if faltando_abas:
        content = f'<div class="page-header"><h1>Planilha incompleta</h1></div><div class="tip">Faltam as abas: {", ".join(faltando_abas)}. Abas encontradas: {", ".join(wb.sheetnames)}</div><p><a href="/boletim/importar" class="btn">Voltar</a></p>'
        return HTMLResponse(render_page("Erro", content, active=""))

    conn = get_db()

    # Garante que as 9 disciplinas canônicas existem
    disc_ids = {}
    for s in BOLETIM_CANONICAL_SUBJECTS:
        row = conn.execute("SELECT id FROM disciplinas WHERE nome = ?", (s,)).fetchone()
        if row:
            disc_ids[s] = row["id"]
        else:
            cur = conn.execute("INSERT INTO disciplinas (nome) VALUES (?)", (s,))
            disc_ids[s] = cur.lastrowid

    def disciplina_id_por_nome(nome):
        return disc_ids.get(_boletim_normalizar_disciplina(nome))

    # Monta o índice (turma, nome_normalizado) -> aluno_id a partir do que JÁ EXISTE
    # no banco (não cria aluno novo — só casa com o cadastro atual).
    alunos_existentes = conn.execute("""
        SELECT a.id, a.nome, t.id AS turma_id, t.nome AS turma_nome FROM alunos a JOIN turmas t ON t.id = a.turma_id
        WHERE t.ano_letivo = ?
    """, (ano,)).fetchall()
    indice_alunos = {}
    ambiguos_nome_only = {}
    turma_id_por_aluno = {}
    for a in alunos_existentes:
        chave = (a["turma_nome"], _boletim_normalizar(a["nome"]))
        indice_alunos[chave] = a["id"]
        ambiguos_nome_only.setdefault(_boletim_normalizar(a["nome"]), []).append(a["id"])
        turma_id_por_aluno[a["id"]] = a["turma_id"]

    def buscar_aluno_id(nome, turma_raw):
        try:
            turma_str = str(int(float(turma_raw)))
        except (TypeError, ValueError):
            turma_str = str(turma_raw).strip()
        return indice_alunos.get((turma_str, _boletim_normalizar(nome)))

    avisos = []
    resumo = {}

    # ---------- ESTUDANTES: só usado pra relatar quantos batem, não grava nada ----------
    ws_e = wb["Estudantes"]
    total_planilha = matched_e = n_sexo = 0
    for row in ws_e.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        total_planilha += 1
        aid_e = buscar_aluno_id(row[0], row[1])
        if aid_e:
            matched_e += 1
            sexo_raw = row[2] if len(row) > 2 else None
            if sexo_raw:
                sx = _boletim_normalizar(sexo_raw)
                sexo_norm = "M" if sx in ("m", "masculino", "masc") else ("F" if sx in ("f", "feminino", "fem") else None)
                if sexo_norm:
                    conn.execute("UPDATE alunos SET sexo = ? WHERE id = ?", (sexo_norm, aid_e))
                    n_sexo += 1
    resumo["estudantes"] = f"{matched_e}/{total_planilha} alunos da planilha já cadastrados no sistema ({n_sexo} com sexo atualizado)"
    if matched_e < total_planilha:
        avisos.append(f"{total_planilha - matched_e} aluno(s) da planilha não foram encontrados no cadastro atual (turma+nome) — não tiveram nenhum dado importado.")

    # ---------- MEDIAS ----------
    ws_m = wb["Medias"]
    header_m = [str(c or "").strip() for c in next(ws_m.iter_rows(max_row=1, values_only=True))]
    n_medias = 0
    for row in ws_m.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        aid = buscar_aluno_id(row[0], row[1])
        if not aid:
            continue
        for col_idx in range(2, min(len(header_m), len(row))):
            valor = row[col_idx]
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                continue  # célula vazia
            did = disciplina_id_por_nome(header_m[col_idx])
            if not did:
                continue
            if isinstance(valor, (int, float)):
                conn.execute("""INSERT INTO boletim_medias (aluno_id, disciplina_id, trimestre, ano, nota, nota_texto)
                                 VALUES (?,?,?,?,?,NULL)
                                 ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET nota = excluded.nota, nota_texto = NULL""",
                             (aid, did, trimestre, ano, float(valor)))
            else:
                # Nota categórica (ex: Educação Digital = PA/PS/PI) — não é número, mas
                # ainda precisa aparecer no boletim impresso.
                conn.execute("""INSERT INTO boletim_medias (aluno_id, disciplina_id, trimestre, ano, nota, nota_texto)
                                 VALUES (?,?,?,?,NULL,?)
                                 ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET nota = NULL, nota_texto = excluded.nota_texto""",
                             (aid, did, trimestre, ano, str(valor).strip()))
            n_medias += 1
    resumo["medias"] = f"{n_medias} notas gravadas/atualizadas"

    # ---------- FALTAS ----------
    ws_f = wb["Faltas"]
    header_f = [str(c or "").strip() for c in next(ws_f.iter_rows(max_row=1, values_only=True))]
    n_faltas = 0
    for row in ws_f.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        aid = buscar_aluno_id(row[0], row[1])
        if not aid:
            continue
        for col_idx in range(2, min(len(header_f), len(row))):
            valor = row[col_idx]
            if valor is None:
                continue
            try:
                faltas_int = int(valor)
            except (TypeError, ValueError):
                continue
            did = disciplina_id_por_nome(header_f[col_idx])
            if not did:
                continue
            conn.execute("""INSERT INTO boletim_faltas (aluno_id, disciplina_id, trimestre, ano, faltas)
                             VALUES (?,?,?,?,?)
                             ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET faltas = excluded.faltas""",
                         (aid, did, trimestre, ano, faltas_int))
            n_faltas += 1
    resumo["faltas"] = f"{n_faltas} registros de falta gravados/atualizados"

    # ---------- RACA/ETNIA (atualiza alunos.raca — só quando o nome bate único) ----------
    ws_r = wb["RacaEtnia"]
    n_raca = 0
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nome = row[0]
        raca = row[1] if len(row) > 1 else None
        if not raca:
            continue
        candidatos = ambiguos_nome_only.get(_boletim_normalizar(nome), [])
        if len(candidatos) == 1:
            conn.execute("UPDATE alunos SET raca = ? WHERE id = ?", (str(raca).strip(), candidatos[0]))
            n_raca += 1
    resumo["raca"] = f"{n_raca} alunos com raça/etnia atualizada"

    # ---------- ANALISE (observações dos professores) ----------
    ws_a = wb["Analise"]
    n_analise = 0
    profs_criados = 0
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        row = list(row) + [None] * (12 - len(row))
        _sid, nome, turma, disc_raw, professor_nome, email, emocional, apoio, alfab, faltoso_json, obs, _ts = row[:12]
        if not nome or not turma:
            continue
        aid = buscar_aluno_id(nome, turma)
        if not aid:
            continue
        if disc_raw and str(disc_raw).strip() not in ("_geral_", "__geral__", ""):
            did = disciplina_id_por_nome(disc_raw)
        else:
            did = disc_ids["Geral"]  # sentinela real em vez de NULL — NULL nunca "colide" numa
            # constraint UNIQUE no SQLite, então cada observação "geral" viraria uma linha nova
            # a cada reimportação em vez de atualizar a existente (bug encontrado e corrigido
            # em teste antes de publicar).

        prof_id = None
        if email:
            email_clean = str(email).strip().lower()
            prof_row = conn.execute("SELECT id FROM professores WHERE email = ?", (email_clean,)).fetchone()
            if prof_row:
                prof_id = prof_row["id"]
            else:
                cur = conn.execute(
                    "INSERT INTO professores (email, nome, status) VALUES (?, ?, 'pendente')",
                    (email_clean, str(professor_nome or email_clean).strip())
                )
                prof_id = cur.lastrowid
                profs_criados += 1

        # Deriva o mapeamento professor↔turma↔disciplina a partir dessa mesma linha
        # (a aba "Professores" da planilha não tem e-mail, só primeiro nome — arriscado
        # de casar; aqui reaproveitamos o e-mail já validado da própria aba Análise).
        if prof_id and did:
            aluno_turma_id = turma_id_por_aluno.get(aid)
            if aluno_turma_id:
                conn.execute("""INSERT OR IGNORE INTO boletim_professor_turma (professor_id, turma_id, disciplina_id)
                                 VALUES (?,?,?)""", (prof_id, aluno_turma_id, did))

        conn.execute("""
            INSERT INTO boletim_analise
                (aluno_id, disciplina_id, professor_id, trimestre, ano, emocional, apoio, alfabetizacao, faltoso, faltoso_json, observacao, atualizado_em)
            VALUES (?,?,?,?,?,?,?,?,0,?,?, CURRENT_TIMESTAMP)
            ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET
                professor_id = excluded.professor_id, emocional = excluded.emocional,
                apoio = excluded.apoio, alfabetizacao = excluded.alfabetizacao,
                faltoso_json = excluded.faltoso_json, observacao = excluded.observacao,
                atualizado_em = CURRENT_TIMESTAMP
        """, (aid, did, prof_id, trimestre, ano, emocional, bool(apoio), bool(alfab), faltoso_json, obs))
        # Nota: "faltoso" (o novo campo simples por disciplina) NÃO é sobrescrito aqui de
        # propósito — o faltoso_json legado da planilha lista as OUTRAS disciplinas em que
        # o aluno falta muito (não a desta própria linha), então não dá pra derivar com
        # segurança o valor certo pra "faltoso nesta disciplina" a partir dele. Fica 0 em
        # linhas novas, e uma reimportação nunca apaga o que um professor já marcou pela
        # tela de Análise.
        n_analise += 1
    resumo["analise"] = f"{n_analise} observações de professores gravadas/atualizadas"
    if profs_criados:
        avisos.append(f"{profs_criados} professor(es) novo(s) criado(s) como 'pendente' (email da planilha não estava cadastrado) — aprove em Usuários se for o caso.")

    conn.commit()
    conn.close()

    resumo_html = "".join(f'<li>{v}</li>' for v in resumo.values())
    avisos_html = ""
    if avisos:
        itens = "".join(f'<li>{a}</li>' for a in avisos)
        avisos_html = f'<div class="tip" style="background:var(--orange-bg); border-color:var(--orange); margin-top:14px;"><strong>Avisos:</strong><ul style="margin:8px 0 0 18px;">{itens}</ul></div>'

    content = f"""
        <div class="page-header">
            <h1>✅ Importação concluída — {trimestre}º Trimestre {ano}</h1>
        </div>
        <ul style="line-height:1.9;">{resumo_html}</ul>
        {avisos_html}
        <div class="page-actions" style="margin-top:18px;">
            <a href="/boletim/importar" class="btn">Importar outra planilha</a>
            <a href="/painel-gestao" class="btn btn-primary">Voltar ao painel</a>
        </div>
    """
    return render_page("Importação concluída", content, active="boletim-importar")


@app.get("/boletim", response_class=HTMLResponse)
def boletim_hub(request: Request, trimestre: Optional[int] = None, ano: Optional[int] = None):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()

    combinacoes = conn.execute("""
        SELECT trimestre, ano, COUNT(DISTINCT aluno_id) AS n_alunos, COUNT(*) AS n_notas
        FROM boletim_medias GROUP BY trimestre, ano ORDER BY ano DESC, trimestre DESC
    """).fetchall()

    if not combinacoes:
        content = """
            <div class="page-header"><h1>📊 Boletim / Conselho de Classe</h1></div>
            <div class="empty">Nenhum dado importado ainda. <a href="/boletim/importar">Importar planilha</a></div>
        """
        return render_page("Boletim", content, active="boletim-dashboard")

    if trimestre is None or ano is None:
        trimestre, ano = combinacoes[0]["trimestre"], combinacoes[0]["ano"]

    seletor_opts = "".join(
        f'<option value="{c["trimestre"]}:{c["ano"]}"{" selected" if c["trimestre"]==trimestre and c["ano"]==ano else ""}>'
        f'{c["trimestre"]}º Trimestre {c["ano"]} — {c["n_alunos"]} aluno(s), {c["n_notas"]} nota(s)</option>'
        for c in combinacoes
    )

    resumo = conn.execute("""
        SELECT
            (SELECT COUNT(*) FROM boletim_medias WHERE trimestre=? AND ano=?) AS n_medias,
            (SELECT COUNT(DISTINCT aluno_id) FROM boletim_medias WHERE trimestre=? AND ano=?) AS n_alunos_medias,
            (SELECT COUNT(*) FROM boletim_faltas WHERE trimestre=? AND ano=?) AS n_faltas,
            (SELECT COUNT(*) FROM boletim_analise WHERE trimestre=? AND ano=?) AS n_analise
    """, (trimestre, ano, trimestre, ano, trimestre, ano, trimestre, ano)).fetchone()

    turmas = conn.execute("""
        SELECT t.id, t.nome, COUNT(DISTINCT bm.aluno_id) AS n_alunos_com_nota
        FROM turmas t
        JOIN alunos a ON a.turma_id = t.id
        JOIN boletim_medias bm ON bm.aluno_id = a.id AND bm.trimestre = ? AND bm.ano = ?
        GROUP BY t.id ORDER BY t.nome
    """, (trimestre, ano)).fetchall()
    conn.close()

    turmas_html = "".join(
        f'<a href="/boletim/turma?trimestre={trimestre}&ano={ano}&turma_id={t["id"]}" class="card card-link">'
        f'<div class="card-title">Turma {t["nome"]}</div>'
        f'<div class="card-meta">{t["n_alunos_com_nota"]} aluno(s) com nota lançada</div>'
        f'</a>'
        for t in turmas
    ) or '<div class="empty">Nenhuma turma com dados nesse trimestre.</div>'

    content = f"""
        <div class="page-header">
            <h1>📊 Boletim / Conselho de Classe</h1>
            <p class="subtitle">Conferência dos dados importados — clique numa turma pra ver o detalhe.</p>
        </div>
        <form method="get" action="/boletim" style="background:var(--bg-subtle); padding:12px 16px; border-radius:8px; margin-bottom:18px;">
            <label style="margin:0; max-width:420px;">Trimestre
                <select name="trimestre_ano" onchange="var v=this.value.split(':'); window.location='/boletim?trimestre='+v[0]+'&ano='+v[1];">
                    {seletor_opts}
                </select>
            </label>
        </form>
        <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(140px, 1fr)); gap:10px; margin-bottom:20px;">
            <div class="metric"><div class="metric-label">Alunos com nota</div><div class="metric-value">{resumo["n_alunos_medias"]}</div></div>
            <div class="metric"><div class="metric-label">Notas lançadas</div><div class="metric-value">{resumo["n_medias"]}</div></div>
            <div class="metric"><div class="metric-label">Registros de falta</div><div class="metric-value">{resumo["n_faltas"]}</div></div>
            <div class="metric"><div class="metric-label">Observações de professores</div><div class="metric-value">{resumo["n_analise"]}</div></div>
        </div>
        <h3>Turmas</h3>
        <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(200px, 1fr)); gap:12px;">
            {turmas_html}
        </div>
    """
    return render_page("Boletim", content, active="boletim-dashboard")


@app.get("/boletim/turma", response_class=HTMLResponse)
def boletim_ver_turma(request: Request, trimestre: int, ano: int, turma_id: int):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()
    turma = conn.execute("SELECT * FROM turmas WHERE id = ?", (turma_id,)).fetchone()
    if not turma:
        conn.close()
        return RedirectResponse("/boletim", status_code=303)

    disciplinas = conn.execute("""
        SELECT DISTINCT d.id, d.nome FROM disciplinas d
        JOIN boletim_medias bm ON bm.disciplina_id = d.id
        WHERE bm.trimestre = ? AND bm.ano = ?
        ORDER BY CASE d.nome
            WHEN 'Português' THEN 1 WHEN 'Matemática' THEN 2 WHEN 'Ciências' THEN 3
            WHEN 'História' THEN 4 WHEN 'Geografia' THEN 5 WHEN 'Inglês' THEN 6
            WHEN 'Arte' THEN 7 WHEN 'Ed. Física' THEN 8 WHEN 'Educação Digital' THEN 9
            ELSE 10 END
    """, (trimestre, ano)).fetchall()

    alunos = conn.execute("""
        SELECT id, nome, numero FROM alunos WHERE turma_id = ? ORDER BY numero, nome
    """, (turma_id,)).fetchall()

    notas_map = {}
    for row in conn.execute("""
        SELECT aluno_id, disciplina_id, nota FROM boletim_medias
        WHERE trimestre = ? AND ano = ? AND aluno_id IN (SELECT id FROM alunos WHERE turma_id = ?)
    """, (trimestre, ano, turma_id)).fetchall():
        notas_map[(row["aluno_id"], row["disciplina_id"])] = row["nota"]

    faltas_map = {}
    for row in conn.execute("""
        SELECT aluno_id, SUM(faltas) AS total FROM boletim_faltas
        WHERE trimestre = ? AND ano = ? AND aluno_id IN (SELECT id FROM alunos WHERE turma_id = ?)
        GROUP BY aluno_id
    """, (trimestre, ano, turma_id)).fetchall():
        faltas_map[row["aluno_id"]] = row["total"]

    analise_count = {}
    for row in conn.execute("""
        SELECT aluno_id, COUNT(*) AS n FROM boletim_analise
        WHERE trimestre = ? AND ano = ? AND aluno_id IN (SELECT id FROM alunos WHERE turma_id = ?)
        GROUP BY aluno_id
    """, (trimestre, ano, turma_id)).fetchall():
        analise_count[row["aluno_id"]] = row["n"]
    conn.close()

    cabecalho_disc = "".join(f'<th style="padding:6px 8px; text-align:center;">{d["nome"]}</th>' for d in disciplinas)
    linhas = ""
    for a in alunos:
        celulas = ""
        for d in disciplinas:
            nota = notas_map.get((a["id"], d["id"]))
            celulas += f'<td style="padding:6px 8px; text-align:center;">{nota if nota is not None else "—"}</td>'
        faltas_total = faltas_map.get(a["id"], 0)
        n_obs = analise_count.get(a["id"], 0)
        linhas += f"""<tr>
            <td style="padding:6px 8px;">{a["numero"] or "—"}</td>
            <td style="padding:6px 8px;">{a["nome"]}</td>
            {celulas}
            <td style="padding:6px 8px; text-align:center;">{faltas_total}</td>
            <td style="padding:6px 8px; text-align:center;">{n_obs if n_obs else "—"}</td>
        </tr>"""

    if not disciplinas:
        body = '<div class="empty">Nenhuma nota lançada pra essa turma nesse trimestre.</div>'
    else:
        body = f"""
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <thead><tr style="background:var(--bg-subtle);">
                <th style="padding:6px 8px; text-align:left;">Nº</th>
                <th style="padding:6px 8px; text-align:left;">Aluno</th>
                {cabecalho_disc}
                <th style="padding:6px 8px;">Faltas (total)</th>
                <th style="padding:6px 8px;">Observações</th>
            </tr></thead>
            <tbody>{linhas}</tbody>
        </table>
        """

    content = f"""
        <div class="page-header">
            <h1>Turma {turma["nome"]} — {trimestre}º Trimestre {ano}</h1>
            <p class="subtitle">{len(alunos)} aluno(s) na turma</p>
        </div>
        {body}
        <div style="margin-top:16px; display:flex; gap:8px; flex-wrap:wrap;">
            <a href="/boletim?trimestre={trimestre}&ano={ano}" class="btn">← Voltar</a>
            <a href="/boletim/relatorio-turma?ano={ano}&turma_id={turma_id}" class="btn btn-primary" target="_blank">🖨️ Gerar Boletim (todos os trimestres)</a>
        </div>
    """
    return render_page(f"Turma {turma['nome']}", content, active="")


BOLETIM_DISC_NUMERICAS = ['Português', 'Matemática', 'Ciências', 'História', 'Geografia', 'Inglês', 'Arte', 'Ed. Física']
BOLETIM_PRIORIDADE_EMOCIONAL = {"fragilizado": 3, "oscilando": 2, "bem": 1}


def _boletim_ano_da_turma(turma_nome):
    t = str(turma_nome or "")
    return {"6": "6°", "7": "7°", "8": "8°", "9": "9°"}.get(t[:1], "Outro")


def _boletim_saeb_nivel(media):
    """Escala SAEB padronizada em todo o sistema (24/08/2026) — mesmos cortes e linguagem
    do Boletim Individual oficial: N1 Insuficiente 0-4,9 | N2 Básico 5,0-6,9 |
    N3 Adequado 7,0-8,4 | N4 Avançado 8,5-10. Antes dashboard/comparativo usavam uma escala
    diferente (5,0/6,5/8,0, sem numeração N1-N4) — unificado a pedido, pra não ter duas
    linguagens de proficiência diferentes dentro do mesmo sistema."""
    if media is None:
        return None
    if media < 5:
        return {"key": "n1", "label": "N1 — Insuficiente", "color": "#dc2626"}
    if media < 7:
        return {"key": "n2", "label": "N2 — Básico", "color": "#ea580c"}
    if media < 8.5:
        return {"key": "n3", "label": "N3 — Adequado", "color": "#16a34a"}
    return {"key": "n4", "label": "N4 — Avançado", "color": "#6366f1"}


def _home_alunos_atencao(conn, prof):
    """Monta o card 'Alunos que precisam de atenção' da tela inicial (24/08/2026):
    combina risco de repetência (mesma regra do dashboard: qualquer disciplina com média
    T1+T2 < 5,0) e maiores faltas. Admin/gestão vê a escola toda; docente só vê o que
    está vinculado a ele em boletim_professor_turma (disciplina+turma específicas).
    Retorna dict {risco: [...], faltas: [...], ano: int} ou None se não há dado ainda."""
    combinacoes = conn.execute("SELECT ano FROM boletim_medias GROUP BY ano ORDER BY ano DESC LIMIT 1").fetchone()
    if not combinacoes:
        return None
    ano = combinacoes["ano"]

    eh_docente_restrito = not prof.get("is_admin") and prof.get("papel") == "docente"
    turma_ids_permitidas = None
    pares_permitidos = None  # set de (turma_nome, disciplina_nome) — só pra docente
    if eh_docente_restrito:
        vinculos = conn.execute("""
            SELECT t.id AS turma_id, t.nome AS turma_nome, d.nome AS disciplina_nome
            FROM boletim_professor_turma bpt
            JOIN turmas t ON t.id = bpt.turma_id
            JOIN disciplinas d ON d.id = bpt.disciplina_id
            WHERE bpt.professor_id = ?
        """, (prof["id"],)).fetchall()
        if not vinculos:
            return {"risco": [], "faltas": [], "ano": ano}
        turma_ids_permitidas = {v["turma_id"] for v in vinculos}
        pares_permitidos = {(v["turma_nome"], v["disciplina_nome"]) for v in vinculos}

    # --- Risco de repetência ---
    risco_bruto = _boletim_possiveis_repetentes(conn, ano, turma_id=None)
    if eh_docente_restrito:
        risco_filtrado = []
        for aluno in risco_bruto:
            discs_ok = [d for d in aluno["disciplinas"] if (aluno["turma_nome"], d["nome"]) in pares_permitidos]
            if discs_ok:
                risco_filtrado.append({**aluno, "disciplinas": discs_ok})
        risco_bruto = risco_filtrado
    risco = risco_bruto[:5]

    # --- Maiores faltas ---
    if eh_docente_restrito:
        placeholders = ",".join("?" * len(turma_ids_permitidas))
        faltas_rows = conn.execute(f"""
            SELECT a.id, a.nome, t.nome AS turma_nome, SUM(bf.faltas) AS total
            FROM boletim_faltas bf
            JOIN alunos a ON a.id = bf.aluno_id
            JOIN turmas t ON t.id = a.turma_id
            JOIN disciplinas d ON d.id = bf.disciplina_id
            WHERE bf.ano = ? AND bf.aluno_id IN (
                SELECT id FROM alunos WHERE turma_id IN ({placeholders})
            )
            GROUP BY a.id ORDER BY total DESC LIMIT 5
        """, [ano] + list(turma_ids_permitidas)).fetchall()
    else:
        faltas_rows = conn.execute("""
            SELECT a.id, a.nome, t.nome AS turma_nome, SUM(bf.faltas) AS total
            FROM boletim_faltas bf
            JOIN alunos a ON a.id = bf.aluno_id
            JOIN turmas t ON t.id = a.turma_id
            WHERE bf.ano = ?
            GROUP BY a.id ORDER BY total DESC LIMIT 5
        """, (ano,)).fetchall()
    faltas = [dict(r) for r in faltas_rows if r["total"]]

    return {"risco": risco, "faltas": faltas, "ano": ano}


def _boletim_possiveis_repetentes(conn, ano, turma_id=None):
    """Lista alunos com risco de repetência: QUALQUER disciplina com média (T1+T2)/2 < 5,0.
    Só considera disciplinas com nota NUMÉRICA nos dois trimestres (Educação Digital é
    descritiva e nunca entra; alunos 'PD'/parecer também ficam de fora dessa disciplina).
    Retorna lista de dicts: {aluno_id, nome, turma_nome, disciplinas: [{nome, t1, t2, media}]},
    ordenada por nome. (24/08/2026)"""
    where_turma = "AND a.turma_id = ?" if turma_id else ""
    params = [ano, ano, ano] + ([turma_id] if turma_id else [])

    rows = conn.execute(f"""
        SELECT a.id AS aluno_id, a.nome AS aluno_nome, t.nome AS turma_nome, d.nome AS disciplina_nome,
               m1.nota AS nota_t1, m2.nota AS nota_t2
        FROM alunos a
        JOIN turmas t ON t.id = a.turma_id AND t.ano_letivo = ?
        JOIN boletim_medias m1 ON m1.aluno_id = a.id AND m1.trimestre = 1 AND m1.ano = ? AND m1.nota IS NOT NULL
        JOIN boletim_medias m2 ON m2.aluno_id = a.id AND m2.trimestre = 2 AND m2.ano = ? AND m2.nota IS NOT NULL
            AND m2.disciplina_id = m1.disciplina_id
        JOIN disciplinas d ON d.id = m1.disciplina_id
        WHERE d.nome != 'Educação Digital' {where_turma}
        ORDER BY a.nome, d.nome
    """, params).fetchall()

    por_aluno = {}
    for r in rows:
        media = (r["nota_t1"] + r["nota_t2"]) / 2
        if media >= 5.0:
            continue
        chave = r["aluno_id"]
        if chave not in por_aluno:
            por_aluno[chave] = {"aluno_id": r["aluno_id"], "nome": r["aluno_nome"], "turma_nome": r["turma_nome"], "disciplinas": []}
        por_aluno[chave]["disciplinas"].append({"nome": r["disciplina_nome"], "t1": r["nota_t1"], "t2": r["nota_t2"], "media": round(media, 1)})

    return sorted(por_aluno.values(), key=lambda x: x["nome"])


def _boletim_enriquecer_alunos(conn, trimestre, ano, turma_id=None):
    """Monta, por aluno, a média (das 8 disciplinas numéricas — Educação Digital é
    categórica PA/PS/PI e não entra na média), nível SAEB, risco de retenção (4+
    disciplinas abaixo de 5,0), se precisa de apoio/tem dificuldade de alfabetização,
    e o estado emocional mais grave relatado entre as disciplinas."""
    sql = """
        SELECT a.id, a.nome, a.numero, a.raca, a.sexo, a.data_nascimento, t.nome AS turma_nome
        FROM alunos a JOIN turmas t ON t.id = a.turma_id
        WHERE t.ano_letivo = ?
    """
    params = [ano]
    if turma_id:
        sql += " AND t.id = ?"
        params.append(turma_id)
    alunos = conn.execute(sql, params).fetchall()

    notas_rows = conn.execute("""
        SELECT bm.aluno_id, d.nome AS disc_nome, bm.nota, bm.nota_texto FROM boletim_medias bm
        JOIN disciplinas d ON d.id = bm.disciplina_id
        WHERE bm.trimestre = ? AND bm.ano = ?
    """, (trimestre, ano)).fetchall()
    notas_por_aluno = {}
    notas_texto_por_aluno = {}
    for r in notas_rows:
        notas_por_aluno.setdefault(r["aluno_id"], {})[r["disc_nome"]] = r["nota"]
        if r["nota_texto"]:
            notas_texto_por_aluno.setdefault(r["aluno_id"], {})[r["disc_nome"]] = r["nota_texto"]

    analise_rows = conn.execute("""
        SELECT aluno_id, emocional, apoio, alfabetizacao, faltoso, observacao FROM boletim_analise
        WHERE trimestre = ? AND ano = ?
    """, (trimestre, ano)).fetchall()
    analise_por_aluno = {}
    for r in analise_rows:
        analise_por_aluno.setdefault(r["aluno_id"], []).append(r)

    faltas_rows = conn.execute("""
        SELECT aluno_id, SUM(faltas) AS total FROM boletim_faltas
        WHERE trimestre = ? AND ano = ? GROUP BY aluno_id
    """, (trimestre, ano)).fetchall()
    faltas_por_aluno = {r["aluno_id"]: r["total"] for r in faltas_rows}

    resultado = []
    for a in alunos:
        notas = notas_por_aluno.get(a["id"], {})
        vals = [notas[d] for d in BOLETIM_DISC_NUMERICAS if notas.get(d) is not None]
        media = sum(vals) / len(vals) if vals else None
        red_count = sum(1 for v in vals if v < 5)

        analises = analise_por_aluno.get(a["id"], [])
        emocional = None
        for r in analises:
            if r["emocional"] and BOLETIM_PRIORIDADE_EMOCIONAL.get(r["emocional"], 0) > BOLETIM_PRIORIDADE_EMOCIONAL.get(emocional, 0):
                emocional = r["emocional"]
        apoio = any(r["apoio"] for r in analises)
        alfab = any(r["alfabetizacao"] for r in analises)
        faltoso_disc = any(r["faltoso"] for r in analises)
        observacoes = " | ".join(r["observacao"] for r in analises if r["observacao"])

        resultado.append({
            "id": a["id"], "nome": a["nome"], "numero": a["numero"], "turma": a["turma_nome"], "raca": a["raca"], "sexo": a["sexo"],
            "data_nascimento": a["data_nascimento"],
            "notas": notas, "notas_texto": notas_texto_por_aluno.get(a["id"], {}),
            "media": media, "saeb": _boletim_saeb_nivel(media),
            "risco_retencao": red_count >= 4,
            "apoio": apoio, "alfab": alfab, "faltoso": faltoso_disc, "emocional": emocional, "observacoes": observacoes,
            "faltas_total": faltas_por_aluno.get(a["id"], 0),
        })
    return resultado


@app.get("/boletim/dashboard", response_class=HTMLResponse)
def boletim_dashboard(request: Request, trimestre: Optional[int] = None, ano: Optional[int] = None,
                       turma_id: Optional[str] = None, ano_esc: Optional[str] = None):
    # Liberado pra todos os docentes (antes só admin) — 25/08/2026
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    # turma_id chega como "" quando o filtro é "todas as turmas" — Optional[int] quebrava
    # aqui com erro 422 (25/08/2026). Aceita como string e converte manualmente.
    turma_id = int(turma_id) if turma_id and turma_id.strip().isdigit() else None
    conn = get_db()

    combinacoes = conn.execute("""
        SELECT trimestre, ano FROM boletim_medias GROUP BY trimestre, ano ORDER BY ano DESC, trimestre DESC
    """).fetchall()
    if not combinacoes:
        conn.close()
        content = '<div class="page-header"><h1>📈 Dashboard Pedagógico</h1></div><div class="empty">Nenhum dado importado ainda. <a href="/boletim/importar">Importar planilha</a></div>'
        return render_page("Dashboard Pedagógico", content, active="boletim-dashboard")
    if trimestre is None or ano is None:
        trimestre, ano = combinacoes[0]["trimestre"], combinacoes[0]["ano"]

    turmas = conn.execute("SELECT id, nome FROM turmas WHERE ano_letivo = ? ORDER BY nome", (ano,)).fetchall()

    enriquecidos = _boletim_enriquecer_alunos(conn, trimestre, ano, turma_id=turma_id)
    conn.close()

    if ano_esc and not turma_id:
        enriquecidos = [e for e in enriquecidos if _boletim_ano_da_turma(e["turma"]) == ano_esc]

    total = len(enriquecidos)
    medias_validas = [e["media"] for e in enriquecidos if e["media"] is not None]
    media_geral = sum(medias_validas) / len(medias_validas) if medias_validas else None
    saeb_geral = _boletim_saeb_nivel(media_geral)
    n_retencao = sum(1 for e in enriquecidos if e["risco_retencao"])
    n_apoio = sum(1 for e in enriquecidos if e["apoio"])
    n_alfab = sum(1 for e in enriquecidos if e["alfab"])

    emo_count = {"bem": 0, "oscilando": 0, "fragilizado": 0}
    for e in enriquecidos:
        if e["emocional"]:
            emo_count[e["emocional"]] += 1

    # --- seletor de trimestre/turma/ano de escolaridade ---
    trimestre_opts = "".join(
        f'<option value="{c["trimestre"]}:{c["ano"]}"{" selected" if c["trimestre"]==trimestre and c["ano"]==ano else ""}>{c["trimestre"]}º Trimestre {c["ano"]}</option>'
        for c in combinacoes
    )
    turma_opts = '<option value="">Escola toda</option>' + "".join(
        f'<option value="{t["id"]}"{" selected" if turma_id==t["id"] else ""}>Turma {t["nome"]}</option>' for t in turmas
    )
    ano_esc_opts = '<option value="">Todos os anos</option>' + "".join(
        f'<option value="{a}"{" selected" if ano_esc==a else ""}>{a} Ano</option>' for a in ["6°", "7°", "8°", "9°"]
    )

    # --- cards de estatística ---
    stat_cards = [
        ("👥", "Total de Estudantes", total, "var(--accent)", None),
        ("🎯", "Média Geral", f"{media_geral:.1f}" if media_geral is not None else "—", "var(--green)", saeb_geral["label"] if saeb_geral else None),
        ("⚠️", "Risco de Retenção (4+ notas vermelhas)", n_retencao, "var(--red)", None),
        ("🤝", "Precisam de Apoio", n_apoio, "var(--orange)", None),
        ("📖", "Dif. de Alfabetização", n_alfab, "var(--purple)", None),
    ]
    stat_cards_html = "".join(
        f'<div class="metric"><div class="metric-label">{lbl}</div><div class="metric-value" style="color:{cor};">{val}</div>'
        + (f'<div style="font-size:11px; color:{cor}; font-weight:600; margin-top:2px;">{sub}</div>' if sub else '')
        + '</div>'
        for ico, lbl, val, cor, sub in stat_cards
    )

    # --- panorama por ano de escolaridade (só quando "escola toda") ---
    panorama_html = ""
    if not turma_id:
        from collections import defaultdict
        por_ano = defaultdict(list)
        for e in enriquecidos:
            por_ano[_boletim_ano_da_turma(e["turma"])].append(e)
        linhas_panorama = ""
        for label in ["6°", "7°", "8°", "9°"]:
            grupo = por_ano.get(label, [])
            if not grupo:
                continue
            medias_g = [e["media"] for e in grupo if e["media"] is not None]
            media_g = sum(medias_g) / len(medias_g) if medias_g else None
            saeb_g = _boletim_saeb_nivel(media_g)
            n_adeq_avanc = sum(1 for e in grupo if e["saeb"] and e["saeb"]["key"] in ("n3", "n4"))
            pct = round(n_adeq_avanc / len(grupo) * 100) if grupo else 0
            n_ret_g = sum(1 for e in grupo if e["risco_retencao"])
            n_apoio_g = sum(1 for e in grupo if e["apoio"])
            cor_g = saeb_g["color"] if saeb_g else "var(--text-muted)"
            linhas_panorama += f"""<tr>
                <td style="padding:8px 10px;"><strong>{label} Ano</strong> <span style="font-size:11px; color:var(--text-muted);">{len(grupo)} alunos</span></td>
                <td style="padding:8px 10px; text-align:center; font-weight:700; color:{cor_g};">{(f"{media_g:.1f}" if media_g is not None else "—")}</td>
                <td style="padding:8px 10px; text-align:center;">{f'<span class="badge" style="background:{cor_g}22; color:{cor_g};">{saeb_g["label"]}</span>' if saeb_g else "—"}</td>
                <td style="padding:8px 10px;">
                    <div style="display:flex; align-items:center; gap:6px;">
                        <div style="flex:1; background:var(--border); border-radius:3px; height:10px; overflow:hidden;"><div style="width:{pct}%; background:{cor_g}; height:100%;"></div></div>
                        <span style="font-size:11px; color:{cor_g}; font-weight:700; width:34px;">{pct}%</span>
                    </div>
                </td>
                <td style="padding:8px 10px; text-align:center; color:var(--red); font-weight:700;">{n_ret_g}</td>
                <td style="padding:8px 10px; text-align:center; color:var(--orange); font-weight:700;">{n_apoio_g}</td>
            </tr>"""
        panorama_html = f"""
        <div class="card" style="margin-bottom:18px; padding:0; overflow:hidden;">
            <div style="padding:14px 16px; border-bottom:1px solid var(--border); font-weight:700; font-size:14px;">📊 Panorama por Ano de Escolaridade</div>
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <thead><tr style="background:var(--bg-subtle);">
                    <th style="padding:8px 10px; text-align:left;">Ano</th><th style="padding:8px 10px;">Média</th>
                    <th style="padding:8px 10px;">SAEB</th><th style="padding:8px 10px;">% Adequado+Avançado</th>
                    <th style="padding:8px 10px;">Risco Retenção</th><th style="padding:8px 10px;">Apoio</th>
                </tr></thead>
                <tbody>{linhas_panorama}</tbody>
            </table>
        </div>"""

    # --- Possíveis repetentes: qualquer disciplina com média (T1+T2)/2 < 5,0 ---
    # Só faz sentido ver isso a partir do 2º trimestre (24/08/2026).
    repetentes_html = ""
    if trimestre == 2:
        conn2 = get_db()
        repetentes = _boletim_possiveis_repetentes(conn2, ano, turma_id=turma_id)
        conn2.close()
        if repetentes:
            linhas_rep = ""
            for r in repetentes:
                discs_html = "".join(
                    f'<span class="badge" style="background:var(--red-bg); color:var(--red); margin-right:4px;" '
                    f'title="T1: {d["t1"]:.1f} · T2: {d["t2"]:.1f}">{d["nome"]} ({d["media"]:.1f})</span>'
                    for d in r["disciplinas"]
                )
                linhas_rep += f"""<tr>
                    <td style="padding:8px 10px;"><strong>{r["nome"]}</strong></td>
                    <td style="padding:8px 10px; color:var(--text-muted);">{r["turma_nome"]}</td>
                    <td style="padding:8px 10px;">{discs_html}</td>
                </tr>"""
            repetentes_html = f"""
            <div class="card" style="margin-bottom:18px; padding:0; overflow:hidden; border-color:var(--red);">
                <div style="padding:14px 16px; border-bottom:1px solid var(--border); font-weight:700; font-size:14px; color:var(--red);">
                    ⚠️ Possíveis Repetentes — {len(repetentes)} aluno(s)
                    <span style="font-weight:400; font-size:12px; color:var(--text-muted); display:block; margin-top:2px;">
                        Média (1º + 2º trimestre) ÷ 2 abaixo de 5,0 em pelo menos uma disciplina
                    </span>
                </div>
                <table style="width:100%; border-collapse:collapse; font-size:13px;">
                    <thead><tr style="background:var(--bg-subtle);">
                        <th style="padding:8px 10px; text-align:left;">Aluno</th>
                        <th style="padding:8px 10px; text-align:left;">Turma</th>
                        <th style="padding:8px 10px; text-align:left;">Disciplina(s) abaixo de 5,0 (média)</th>
                    </tr></thead>
                    <tbody>{linhas_rep}</tbody>
                </table>
            </div>"""
        else:
            repetentes_html = """
            <div class="card" style="margin-bottom:18px; padding:14px 16px; border-color:var(--green);">
                <strong style="color:var(--green);">✅ Nenhum possível repetente</strong>
                <span style="font-size:12px; color:var(--text-muted);"> — nenhum aluno com média (T1+T2)/2 abaixo de 5,0 em alguma disciplina, considerando os dados já importados.</span>
            </div>"""

    # --- distribuição SAEB por disciplina (Português e Matemática) ---
    def dist_saeb_disciplina(disc):
        dist = {"n1": 0, "n2": 0, "n3": 0, "n4": 0}
        abaixo_lista = []
        for e in enriquecidos:
            v = e["notas"].get(disc)
            lvl = _boletim_saeb_nivel(v)
            if lvl:
                dist[lvl["key"]] += 1
                if lvl["key"] == "n1":
                    abaixo_lista.append((e["nome"], e["turma"], v))
        abaixo_lista.sort(key=lambda x: x[2])
        return dist, abaixo_lista

    dist_pt, abaixo_pt = dist_saeb_disciplina("Português")
    dist_mt, abaixo_mt = dist_saeb_disciplina("Matemática")

    def abaixo_html(lista):
        if not lista:
            return '<p style="font-size:12px; color:var(--text-muted);">Nenhum aluno abaixo do básico. 🎉</p>'
        itens = "".join(
            f'<div style="display:flex; justify-content:space-between; padding:3px 8px; font-size:12px; background:var(--red-bg); border-radius:4px; margin-bottom:3px;">'
            f'<span>{nome} <span style="color:var(--text-muted);">· {turma}</span></span><strong style="color:var(--red);">{v:.1f}</strong></div>'
            for nome, turma, v in lista[:15]
        )
        extra = f'<p style="font-size:11px; color:var(--text-muted); margin-top:4px;">+{len(lista)-15} outro(s)</p>' if len(lista) > 15 else ""
        return itens + extra

    # --- Desempenho por Disciplina (ranking de todas as disciplinas numéricas) ---
    ranking_disc = []
    for d in BOLETIM_DISC_NUMERICAS:
        vals = [e["notas"].get(d) for e in enriquecidos if e["notas"].get(d) is not None]
        if vals:
            media_d = sum(vals) / len(vals)
            ranking_disc.append((d, media_d, _boletim_saeb_nivel(media_d)))
    ranking_disc.sort(key=lambda x: x[1], reverse=True)
    disc_labels_js = "[" + ",".join(f'"{d[:4]}."' for d, _, _ in ranking_disc) + "]"
    disc_valores_js = "[" + ",".join(f"{m:.2f}" for _, m, _ in ranking_disc) + "]"
    disc_cores_js = "[" + ",".join(f'"{lvl["color"]}"' if lvl else '"#94a3b8"' for _, _, lvl in ranking_disc) + "]"
    ranking_disc_html = "".join(
        f'<div style="display:flex; justify-content:space-between; align-items:center; padding:4px 8px; background:var(--bg-subtle); border-radius:5px; margin-bottom:3px;">'
        f'<span style="font-size:12px; font-weight:600;">{d}</span>'
        f'<span><strong style="color:{lvl["color"] if lvl else "var(--text-muted)"};">{m:.1f}</strong>'
        + (f' <span class="badge" style="background:{lvl["color"]}22; color:{lvl["color"]};">{lvl["label"]}</span>' if lvl else '')
        + '</span></div>'
        for d, m, lvl in ranking_disc
    )

    # --- Painéis de Faltosos, Alfabetização e Raça/Etnia ---
    def lista_alunos_html(criterio):
        alunos_filtrados = [e for e in enriquecidos if criterio(e)]
        if not alunos_filtrados:
            return '<p style="font-size:12px; color:var(--text-muted);">Nenhum aluno.</p>', 0
        itens = "".join(
            f'<div style="padding:3px 8px; font-size:12px; background:var(--bg-subtle); border-radius:4px; margin-bottom:3px;">{e["nome"]} <span style="color:var(--text-muted);">· {e["turma"]}</span></div>'
            for e in alunos_filtrados[:20]
        )
        extra = f'<p style="font-size:11px; color:var(--text-muted); margin-top:4px;">+{len(alunos_filtrados)-20} outro(s)</p>' if len(alunos_filtrados) > 20 else ""
        return itens + extra, len(alunos_filtrados)

    faltosos_html, n_faltosos_lista = lista_alunos_html(lambda e: e["faltoso"])
    alfab_html, n_alfab_lista = lista_alunos_html(lambda e: e["alfab"])

    raca_count = {}
    for e in enriquecidos:
        r = e["raca"] or "Não informado"
        raca_count[r] = raca_count.get(r, 0) + 1
    raca_ordenado = sorted(raca_count.items(), key=lambda x: -x[1])
    raca_html = "".join(
        f'<div style="display:flex; justify-content:space-between; padding:3px 8px; font-size:12px; background:var(--bg-subtle); border-radius:4px; margin-bottom:3px;">'
        f'<span>{r}</span><strong>{c}</strong></div>'
        for r, c in raca_ordenado
    ) or '<p style="font-size:12px; color:var(--text-muted);">Sem dados de raça/etnia.</p>'

    # --- Estudantes que Precisam de Atenção ---
    alertas_alunos = []
    for e in enriquecidos:
        motivos = []
        if e["risco_retencao"]:
            motivos.append(f'<span class="badge" style="background:var(--red-bg); color:var(--red);">⚠️ {sum(1 for v in e["notas"].values() if isinstance(v,(int,float)) and v<5)} vermelhas</span>')
        if e["apoio"]:
            motivos.append('<span class="badge" style="background:var(--orange-bg); color:var(--orange);">🤝 Apoio</span>')
        if e["alfab"]:
            motivos.append('<span class="badge" style="background:var(--purple-bg); color:var(--purple);">📖 Alfabetização</span>')
        if e["emocional"] == "fragilizado":
            motivos.append('<span class="badge" style="background:var(--red-bg); color:var(--red);">😟 Fragilizado</span>')
        if e["faltoso"]:
            motivos.append('<span class="badge" style="background:var(--orange-bg); color:var(--orange);">🚫 Faltoso</span>')
        if motivos:
            alertas_alunos.append((e, motivos))
    alertas_alunos.sort(key=lambda x: -len(x[1]))

    alertas_html = "".join(
        f'<div style="background:var(--card); border:1px solid var(--border); border-radius:8px; padding:10px 12px;">'
        f'<div style="font-weight:600; font-size:13px;">{e["nome"]}</div>'
        f'<div style="font-size:11px; color:var(--text-muted); margin-bottom:6px;">{e["turma"]} · Média: {f"{e['media']:.1f}" if e["media"] is not None else "—"}</div>'
        f'<div style="display:flex; flex-wrap:wrap; gap:4px;">{"".join(motivos)}</div>'
        f'</div>'
        for e, motivos in alertas_alunos[:60]
    ) if alertas_alunos else '<p style="color:var(--text-muted);">Nenhum estudante com alertas nesse recorte. 🎉</p>'
    extra_alertas = f'<p style="font-size:12px; color:var(--text-muted); margin-top:8px;">+{len(alertas_alunos)-60} outro(s) — refine o filtro pra ver todos.</p>' if len(alertas_alunos) > 60 else ""

    # --- Diferença de Médias por Disciplina — Negro × Branco ---
    def _media_disc_grupo(grupo, disc):
        vs = [e["notas"].get(disc) for e in grupo if e["notas"].get(disc) is not None]
        return (sum(vs) / len(vs)) if vs else None

    negro = [e for e in enriquecidos if e.get("raca") in ("Preta", "Parda")]
    branco = [e for e in enriquecidos if e.get("raca") == "Branca"]
    gap_racial_html = ""
    gap_racial_labels_js = gap_racial_negro_js = gap_racial_branco_js = "[]"
    tem_gap_racial = len(negro) >= 3 and len(branco) >= 3
    if tem_gap_racial:
        labels_gr, negro_avgs, branco_avgs, diffs_gr = [], [], [], []
        for d in BOLETIM_DISC_NUMERICAS:
            nA, bA = _media_disc_grupo(negro, d), _media_disc_grupo(branco, d)
            if nA is None and bA is None:
                continue
            labels_gr.append(d)
            negro_avgs.append(nA)
            branco_avgs.append(bA)
            diffs_gr.append(((nA - bA) / bA * 100) if (nA is not None and bA is not None and bA) else None)
        gap_racial_labels_js = "[" + ",".join(f'"{d[:4]}."' for d in labels_gr) + "]"
        gap_racial_negro_js = "[" + ",".join(f"{v:.2f}" if v is not None else "null" for v in negro_avgs) + "]"
        gap_racial_branco_js = "[" + ",".join(f"{v:.2f}" if v is not None else "null" for v in branco_avgs) + "]"
        diffs_validos = [d for d in diffs_gr if d is not None]
        media_diff = sum(diffs_validos) / len(diffs_validos) if diffs_validos else 0
        cor_diff = "var(--red)" if media_diff < -5 else ("var(--orange)" if media_diff < 0 else "var(--green)")
        gap_racial_html = f"""
        <div class="card" style="margin-bottom:18px;">
            <h3 style="margin-top:0;">⚖️ Diferença de Médias por Disciplina — Negro × Branco</h3>
            <p style="font-size:11px; color:var(--text-muted); margin-top:-6px;">Negro = Preta + Parda ({len(negro)} alunos) · Branco ({len(branco)} alunos)</p>
            <div style="height:200px; position:relative; margin-bottom:10px;"><canvas id="ch-gap-racial"></canvas></div>
            <div style="background:var(--bg-subtle); border-radius:8px; padding:8px 14px; display:inline-block;">
                <div style="font-size:11px; color:var(--text-muted);">Diferença média geral</div>
                <div style="font-size:18px; font-weight:800; color:{cor_diff};">{"+" if media_diff>=0 else ""}{media_diff:.1f}%</div>
                <div style="font-size:10px; color:var(--text-muted);">Negativo = estudantes negros com média inferior</div>
            </div>
        </div>"""

    # --- Diferença de Médias por Disciplina — Meninos × Meninas ---
    masc = [e for e in enriquecidos if e.get("sexo") == "M"]
    fem = [e for e in enriquecidos if e.get("sexo") == "F"]
    gap_genero_html = ""
    gap_genero_labels_js = gap_genero_m_js = gap_genero_f_js = "[]"
    tem_gap_genero = len(masc) >= 3 and len(fem) >= 3
    if tem_gap_genero:
        labels_gg, m_avgs, f_avgs = [], [], []
        for d in BOLETIM_DISC_NUMERICAS:
            mA, fA = _media_disc_grupo(masc, d), _media_disc_grupo(fem, d)
            if mA is None and fA is None:
                continue
            labels_gg.append(d)
            m_avgs.append(mA)
            f_avgs.append(fA)
        gap_genero_labels_js = "[" + ",".join(f'"{d[:4]}."' for d in labels_gg) + "]"
        gap_genero_m_js = "[" + ",".join(f"{v:.2f}" if v is not None else "null" for v in m_avgs) + "]"
        gap_genero_f_js = "[" + ",".join(f"{v:.2f}" if v is not None else "null" for v in f_avgs) + "]"
        gap_genero_html = f"""
        <div class="card" style="margin-bottom:18px;">
            <h3 style="margin-top:0;">⚥ Diferença de Médias por Disciplina — Meninos × Meninas</h3>
            <p style="font-size:11px; color:var(--text-muted); margin-top:-6px;">Meninos ({len(masc)}) · Meninas ({len(fem)})</p>
            <div style="height:200px; position:relative;"><canvas id="ch-gap-genero"></canvas></div>
        </div>"""

    # --- Diferença Negro × Branco por Ano de Escolaridade (só escola toda) ---
    gap_anos_html = ""
    gap_anos_charts_js = ""
    if not turma_id and tem_gap_racial:
        from collections import defaultdict as _dd
        blocos_anos = ""
        for ano_esc in ["6", "7", "8", "9"]:
            suf = f"{ano_esc}°"
            neg_grp = [e for e in negro if _boletim_ano_da_turma(e["turma"]) == suf]
            bra_grp = [e for e in branco if _boletim_ano_da_turma(e["turma"]) == suf]
            if len(neg_grp) < 2 or len(bra_grp) < 2:
                continue
            lbs, nAs, bAs = [], [], []
            for d in BOLETIM_DISC_NUMERICAS:
                nA, bA = _media_disc_grupo(neg_grp, d), _media_disc_grupo(bra_grp, d)
                if nA is None and bA is None:
                    continue
                lbs.append(d)
                nAs.append(nA)
                bAs.append(bA)
            canvas_id = f"ch-gap-ano-{ano_esc}"
            blocos_anos += f"""
            <div class="card">
                <h4 style="margin:0 0 8px;">{ano_esc}º Ano <span style="font-size:11px; color:var(--text-muted); font-weight:400;">({len(neg_grp)} negros, {len(bra_grp)} brancos)</span></h4>
                <div style="height:160px; position:relative;"><canvas id="{canvas_id}"></canvas></div>
            </div>"""
            lbs_js = "[" + ",".join(f'"{d[:4]}."' for d in lbs) + "]"
            nAs_js = "[" + ",".join(f"{v:.2f}" if v is not None else "null" for v in nAs) + "]"
            bAs_js = "[" + ",".join(f"{v:.2f}" if v is not None else "null" for v in bAs) + "]"
            gap_anos_charts_js += f"""
            boletimChart('{canvas_id}', 'bar', {lbs_js}, [
                {{ label:'Negro', data:{nAs_js}, backgroundColor:'#a78bfa' }},
                {{ label:'Branco', data:{bAs_js}, backgroundColor:'#38bdf8' }}
            ], {{scales:{{y:{{max:10}}}}}});"""
        if blocos_anos:
            gap_anos_html = f"""
            <div class="card" style="margin-bottom:18px;">
                <h3 style="margin-top:0;">⚖️ Diferença Negro × Branco por Ano de Escolaridade</h3>
                <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(240px, 1fr)); gap:12px;">
                    {blocos_anos}
                </div>
            </div>"""

    content = f"""
        <div class="page-header" style="display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:10px;">
            <div>
                <h1>📈 Dashboard Pedagógico</h1>
                <p class="subtitle">{total} estudante(s) · {trimestre}º Trimestre {ano}</p>
            </div>
            <a href="/boletim/relatorio-geral?trimestre={trimestre}&ano={ano}{f'&turma_id={turma_id}' if turma_id else ''}{f'&ano_esc={ano_esc}' if ano_esc else ''}" class="btn" target="_blank">🖨️ Relatório Geral de Gestão</a>
        </div>
        <form method="get" action="/boletim/dashboard" style="background:var(--bg-subtle); padding:12px 16px; border-radius:8px; margin-bottom:18px;">
            <div style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">
                <label style="margin:0; flex:1 1 180px;">Trimestre
                    <select onchange="var v=this.value.split(':'); document.getElementById('f-trimestre').value=v[0]; document.getElementById('f-ano').value=v[1]; this.form.submit();">
                        {trimestre_opts}
                    </select>
                    <input type="hidden" id="f-trimestre" name="trimestre" value="{trimestre}">
                    <input type="hidden" id="f-ano" name="ano" value="{ano}">
                </label>
                <label style="margin:0; flex:1 1 160px;">Turma
                    <select name="turma_id" onchange="this.form.submit();">{turma_opts}</select>
                </label>
                <label style="margin:0; flex:1 1 160px;">Ano de escolaridade
                    <select name="ano_esc" onchange="this.form.submit();">{ano_esc_opts}</select>
                </label>
            </div>
        </form>

        <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(160px, 1fr)); gap:10px; margin-bottom:18px;">
            {stat_cards_html}
        </div>

        {panorama_html}

        {repetentes_html}

        <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(280px, 1fr)); gap:16px; margin-bottom:18px;">
            <div class="card">
                <h3 style="margin-top:0;">📖 SAEB — Português</h3>
                <div style="height:180px; position:relative;"><canvas id="ch-saeb-pt"></canvas></div>
                <div style="margin-top:10px;">{abaixo_html(abaixo_pt)}</div>
            </div>
            <div class="card">
                <h3 style="margin-top:0;">🔢 SAEB — Matemática</h3>
                <div style="height:180px; position:relative;"><canvas id="ch-saeb-mt"></canvas></div>
                <div style="margin-top:10px;">{abaixo_html(abaixo_mt)}</div>
            </div>
            <div class="card">
                <h3 style="margin-top:0;">😊 Estado Emocional</h3>
                <div style="max-width:260px; height:180px; margin:0 auto; position:relative;"><canvas id="ch-emo"></canvas></div>
            </div>
        </div>

        <div class="card" style="margin-bottom:18px;">
            <h3 style="margin-top:0;">📐 Desempenho por Disciplina</h3>
            <p style="font-size:12px; color:var(--text-muted); margin-top:-6px;">Maior → menor · cor = nível SAEB</p>
            <div style="height:220px; position:relative; margin-bottom:12px;"><canvas id="ch-disc"></canvas></div>
            {ranking_disc_html}
        </div>

        <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(240px, 1fr)); gap:16px; margin-bottom:18px;">
            <div class="card">
                <h3 style="margin-top:0;">🚫 Faltosos ({n_faltosos_lista})</h3>
                {faltosos_html}
            </div>
            <div class="card">
                <h3 style="margin-top:0;">📖 Dificuldade de Alfabetização ({n_alfab_lista})</h3>
                {alfab_html}
            </div>
            <div class="card">
                <h3 style="margin-top:0;">🧑🏾 Raça/Etnia</h3>
                {raca_html}
            </div>
        </div>

        <div class="card" style="margin-bottom:18px;">
            <h3 style="margin-top:0;">⚠️ Estudantes que Precisam de Atenção ({len(alertas_alunos)})</h3>
            <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:10px;">
                {alertas_html}
            </div>
            {extra_alertas}
        </div>

        {gap_racial_html}
        {gap_genero_html}
        {gap_anos_html}

        <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
        <script>
        function _boletimCssVar(nome) {{
            return getComputedStyle(document.documentElement).getPropertyValue(nome).trim() || '#94a3b8';
        }}
        function boletimChart(canvasId, tipo, labels, datasets, opcoesExtra) {{
            var el = document.getElementById(canvasId);
            if (!el) return null;
            var corMuted = _boletimCssVar('--text-muted');
            var corBorda = _boletimCssVar('--border');
            datasets.forEach(function(ds) {{
                if (ds.borderRadius === undefined) ds.borderRadius = 6;
                if (ds.borderWidth === undefined) ds.borderWidth = 0;
                if (tipo === 'doughnut' && ds.borderWidth === 0) ds.borderWidth = 3;
                if (tipo === 'doughnut' && ds.borderColor === undefined) ds.borderColor = _boletimCssVar('--card') || '#fff';
            }});
            var opcoesBase = {{
                responsive: true, maintainAspectRatio: false,
                font: {{ family: 'Sora' }},
                plugins: {{
                    legend: {{
                        display: tipo === 'doughnut' || (datasets.length > 1),
                        position: tipo === 'doughnut' ? 'bottom' : 'top',
                        labels: {{ color: corMuted, font: {{ family: 'Sora', size: 11 }}, padding: 10, usePointStyle: true, pointStyle: 'circle' }}
                    }},
                    tooltip: {{
                        backgroundColor: '#0a1220', borderColor: '#1e3050', borderWidth: 1,
                        titleColor: '#e2eaf5', bodyColor: '#8ca3c4', padding: 10, cornerRadius: 8,
                        titleFont: {{ family: 'Sora', weight: '700' }}, bodyFont: {{ family: 'Sora' }},
                        usePointStyle: true
                    }}
                }}
            }};
            if (tipo === 'doughnut') {{
                opcoesBase.cutout = '68%';
            }} else {{
                opcoesBase.scales = {{
                    x: {{ ticks: {{ color: corMuted, font: {{ family: 'Sora', size: 10 }} }}, grid: {{ display: false }}, border: {{ display: false }} }},
                    y: {{ ticks: {{ color: corMuted, font: {{ family: 'Sora', size: 10 }} }}, grid: {{ color: corBorda }}, border: {{ display: false }}, beginAtZero: true }}
                }};
            }}
            var opcoesFinal = Object.assign({{}}, opcoesBase, opcoesExtra || {{}});
            if (opcoesExtra && opcoesExtra.scales && opcoesBase.scales) {{
                opcoesFinal.scales = Object.assign({{}}, opcoesBase.scales, opcoesExtra.scales);
            }}
            return new Chart(el, {{ type: tipo, data: {{ labels: labels, datasets: datasets }}, options: opcoesFinal }});
        }}

            {"boletimChart('ch-gap-racial','bar'," + gap_racial_labels_js + ",[{label:'Negro (Preta+Parda)',data:" + gap_racial_negro_js + ",backgroundColor:'#a78bfa'},{label:'Branco',data:" + gap_racial_branco_js + ",backgroundColor:'#38bdf8'}],{scales:{y:{max:10}}});" if tem_gap_racial else ""}
            {"boletimChart('ch-gap-genero','bar'," + gap_genero_labels_js + ",[{label:'Meninos',data:" + gap_genero_m_js + ",backgroundColor:'#38bdf8'},{label:'Meninas',data:" + gap_genero_f_js + ",backgroundColor:'#f472b6'}],{scales:{y:{max:10}}});" if tem_gap_genero else ""}
            {gap_anos_charts_js}
            boletimChart('ch-disc', 'bar', {disc_labels_js}, [{{ data: {disc_valores_js}, backgroundColor: {disc_cores_js} }}], {{scales:{{y:{{max:10}}}}}});
            boletimChart('ch-saeb-pt', 'bar', ['N1 Insuf.', 'N2 Básico', 'N3 Adequado', 'N4 Avançado'],
                [{{ data: [{dist_pt["n1"]}, {dist_pt["n2"]}, {dist_pt["n3"]}, {dist_pt["n4"]}],
                   backgroundColor: ['#dc2626','#ea580c','#16a34a','#6366f1'] }}]);
            boletimChart('ch-saeb-mt', 'bar', ['N1 Insuf.', 'N2 Básico', 'N3 Adequado', 'N4 Avançado'],
                [{{ data: [{dist_mt["n1"]}, {dist_mt["n2"]}, {dist_mt["n3"]}, {dist_mt["n4"]}],
                   backgroundColor: ['#dc2626','#ea580c','#16a34a','#6366f1'] }}]);
            boletimChart('ch-emo', 'doughnut', ['Bem','Oscilando','Fragilizado'],
                [{{ data: [{emo_count["bem"]}, {emo_count["oscilando"]}, {emo_count["fragilizado"]}],
                   backgroundColor: ['#16a34a','#ea580c','#dc2626'] }}]);
        </script>
    """
    return render_page("Dashboard Pedagógico", content, active="boletim-dashboard")


BOLETIM_EMOJI_EMOCIONAL = {"bem": "😊", "oscilando": "😐", "fragilizado": "😟"}
BOLETIM_LABEL_EMOCIONAL = {"bem": "Bem", "oscilando": "Oscilando", "fragilizado": "Fragilizado"}
BOLETIM_ORDEM_DISCIPLINAS = ['Português', 'Matemática', 'Ciências', 'História', 'Geografia',
                             'Inglês', 'Arte', 'Ed. Física', 'Educação Digital']


def _boletim_dados_turma_multitrimestre(conn, ano, turma_id):
    """Monta, pra cada aluno de uma turma, o quadro completo dos 3 trimestres:
    notas por disciplina em cada trimestre, faltas por trimestre, e um resumo (média
    final, situação SAEB, alertas, observações) — base pro boletim impresso, que
    mostra os 3 trimestres lado a lado mesmo que só o 1º já tenha dado."""
    alunos = conn.execute("""
        SELECT id, nome, numero FROM alunos WHERE turma_id = ? ORDER BY numero, nome
    """, (turma_id,)).fetchall()
    aluno_ids = [a["id"] for a in alunos]
    if not aluno_ids:
        return []
    placeholders = ",".join("?" * len(aluno_ids))

    notas_rows = conn.execute(f"""
        SELECT bm.aluno_id, d.nome AS disc, bm.trimestre, bm.nota, bm.nota_texto
        FROM boletim_medias bm JOIN disciplinas d ON d.id = bm.disciplina_id
        WHERE bm.ano = ? AND bm.aluno_id IN ({placeholders})
    """, [ano] + aluno_ids).fetchall()
    notas_por_aluno = {}
    for r in notas_rows:
        notas_por_aluno.setdefault(r["aluno_id"], {}).setdefault(r["disc"], {})[r["trimestre"]] = (r["nota"], r["nota_texto"])

    faltas_rows = conn.execute(f"""
        SELECT aluno_id, trimestre, SUM(faltas) AS total FROM boletim_faltas
        WHERE ano = ? AND aluno_id IN ({placeholders}) GROUP BY aluno_id, trimestre
    """, [ano] + aluno_ids).fetchall()
    faltas_por_aluno = {}
    for r in faltas_rows:
        faltas_por_aluno.setdefault(r["aluno_id"], {})[r["trimestre"]] = r["total"]

    analise_rows = conn.execute(f"""
        SELECT aluno_id, trimestre, emocional, apoio, alfabetizacao, faltoso, observacao
        FROM boletim_analise WHERE ano = ? AND aluno_id IN ({placeholders})
        ORDER BY trimestre
    """, [ano] + aluno_ids).fetchall()
    analise_por_aluno = {}
    for r in analise_rows:
        analise_por_aluno.setdefault(r["aluno_id"], []).append(r)

    resultado = []
    for a in alunos:
        notas = notas_por_aluno.get(a["id"], {})
        medias_finais = {}
        for disc in BOLETIM_DISC_NUMERICAS:
            vals = [v[0] for t, v in notas.get(disc, {}).items() if v[0] is not None]
            medias_finais[disc] = sum(vals) / len(vals) if vals else None
        todas_vals = [v for v in medias_finais.values() if v is not None]
        media_geral = sum(todas_vals) / len(todas_vals) if todas_vals else None

        analises = analise_por_aluno.get(a["id"], [])
        emocional = None
        for r in analises:
            if r["emocional"] and BOLETIM_PRIORIDADE_EMOCIONAL.get(r["emocional"], 0) > BOLETIM_PRIORIDADE_EMOCIONAL.get(emocional, 0):
                emocional = r["emocional"]
        apoio = any(r["apoio"] for r in analises)
        alfab = any(r["alfabetizacao"] for r in analises)
        faltoso = any(r["faltoso"] for r in analises)
        observacoes = " | ".join(f"[{r['trimestre']}º Tri] {r['observacao']}" for r in analises if r["observacao"])

        resultado.append({
            "id": a["id"], "nome": a["nome"], "numero": a["numero"],
            "notas": notas, "medias_finais": medias_finais, "media_geral": media_geral,
            "saeb_geral": _boletim_saeb_nivel(media_geral),
            "faltas_por_trim": faltas_por_aluno.get(a["id"], {}),
            "emocional": emocional, "apoio": apoio, "alfab": alfab, "faltoso": faltoso,
            "observacoes": observacoes,
        })
    return resultado


def _boletim_turma_medias_disciplina(dados_turma):
    """A partir do resultado de _boletim_dados_turma_multitrimestre, calcula a média da
    TURMA em cada disciplina (média das médias finais dos alunos que têm nota numérica
    naquela disciplina) — usado no comparativo Aluno × Turma do boletim individual (24/08/2026)."""
    medias_turma = {}
    for disc in BOLETIM_DISC_NUMERICAS:
        vals = [e["medias_finais"].get(disc) for e in dados_turma if e["medias_finais"].get(disc) is not None]
        medias_turma[disc] = sum(vals) / len(vals) if vals else None
    return medias_turma


def _boletim_ranking_faltas(dados_turma):
    """Retorna {aluno_id: (posicao, total_turma)} rankeando por total de faltas acumulado
    (menos faltas = melhor posição/nº menor), só considerando quem tem algum registro de
    falta. Empates dividem a mesma posição (24/08/2026)."""
    com_faltas = [(e["id"], sum(v for v in e["faltas_por_trim"].values() if v)) for e in dados_turma]
    com_faltas.sort(key=lambda x: x[1])
    ranking = {}
    total = len(com_faltas)
    pos_atual = 0
    valor_anterior = None
    for i, (aid, total_faltas) in enumerate(com_faltas):
        if total_faltas != valor_anterior:
            pos_atual = i + 1
            valor_anterior = total_faltas
        ranking[aid] = (pos_atual, total)
    return ranking


@app.get("/boletim/boletim-individual", response_class=HTMLResponse)
def form_boletim_individual(request: Request):
    # Liberado pra todos os docentes (antes só admin/gestão) — 25/08/2026
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)

    conn = get_db()
    turmas = conn.execute("SELECT id, nome, ano_letivo FROM turmas ORDER BY ano_letivo DESC, nome").fetchall()
    alunos_todos = conn.execute("SELECT id, turma_id, nome FROM alunos ORDER BY nome").fetchall()
    conn.close()
    opts_turmas = "".join(f'<option value="{t["id"]}" data-ano="{t["ano_letivo"]}">{t["nome"]} ({t["ano_letivo"]})</option>' for t in turmas)

    # Mapa turma_id -> [alunos] embutido como JSON pro <select> de aluno se popular
    # sozinho no navegador, sem precisar recarregar a página (26/08/2026, a pedido).
    alunos_por_turma = {}
    for a in alunos_todos:
        alunos_por_turma.setdefault(a["turma_id"], []).append({"id": a["id"], "nome": a["nome"]})
    alunos_por_turma_json = json.dumps(alunos_por_turma, ensure_ascii=False)

    content = f"""
        <div class="page-header">
            <h1>🧾 Gerar Boletim Individual</h1>
            <p class="subtitle">Gera o boletim de cada aluno com as notas acumuladas de todos os trimestres já importados (1º, 2º e/ou 3º), no modelo oficial da escola.</p>
        </div>
        <div class="tip">Escolha a turma — dá pra gerar o boletim de <strong>toda a turma</strong> (um aluno por página) ou escolher um aluno específico depois de carregar a turma. Use "Imprimir" no navegador e escolha "Salvar como PDF" pra baixar.</div>
        <form action="/boletim/boletim-individual/gerar" method="get" target="_blank">
            <div style="display:flex; flex-wrap:wrap; gap:14px; align-items:flex-end;">
                <label style="margin:0; flex:1 1 220px;">Turma
                    <select name="turma_id" id="sel-turma-boletim" required onchange="_atualizarAlunosBoletim(this)">
                        <option value="">— selecione —</option>
                        {opts_turmas}
                    </select>
                </label>
                <label style="margin:0; flex:1 1 220px;">Aluno (opcional)
                    <select name="aluno_id" id="sel-aluno-boletim">
                        <option value="">— toda a turma —</option>
                    </select>
                </label>
                <label style="margin:0; flex:1 1 120px;">Ano
                    <input type="number" name="ano" id="inp-ano-boletim" value="2026" required>
                </label>
            </div>
            <div class="page-actions">
                <button type="submit" class="btn btn-primary">Gerar boletim</button>
            </div>
        </form>
        <script>
        const _alunosPorTurmaBoletim = {alunos_por_turma_json};
        function _atualizarAlunosBoletim(selTurma) {{
            const opt = selTurma.selectedOptions[0];
            document.getElementById('inp-ano-boletim').value = opt.dataset.ano || '';
            const selAluno = document.getElementById('sel-aluno-boletim');
            const lista = _alunosPorTurmaBoletim[selTurma.value] || [];
            lista.sort((a, b) => a.nome.localeCompare(b.nome, 'pt-BR'));
            selAluno.innerHTML = '<option value="">— toda a turma —</option>' +
                lista.map(a => `<option value="${{a.id}}">${{a.nome}}</option>`).join('');
        }}
        </script>
    """
    return render_page("Boletim Individual", content, active="boletim-individual")


@app.get("/boletim/boletim-individual/gerar", response_class=HTMLResponse)
def gerar_boletim_individual(ano: int, turma_id: str, aluno_id: Optional[str] = None):
    # Liberado pra todos os docentes (antes só admin/gestão) — 25/08/2026
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)
    # Mesma proteção contra string vazia do formulário (25/08/2026)
    if not turma_id or not turma_id.strip().isdigit():
        return RedirectResponse("/boletim/boletim-individual", status_code=303)
    turma_id = int(turma_id)
    aluno_id = int(aluno_id) if aluno_id and aluno_id.strip().isdigit() else None

    conn = get_db()
    turma = conn.execute("SELECT * FROM turmas WHERE id = ?", (turma_id,)).fetchone()
    if not turma:
        conn.close()
        return RedirectResponse("/boletim/boletim-individual", status_code=303)

    dados_turma_completo = _boletim_dados_turma_multitrimestre(conn, ano, turma_id)

    # Faltas ACUMULADAS por disciplina (a função acima só soma por trimestre, perdendo
    # a quebra por disciplina que o modelo do boletim precisa mostrar) — 24/08/2026.
    aluno_ids_turma = [e["id"] for e in dados_turma_completo]
    faltas_disc_por_aluno = {}
    if aluno_ids_turma:
        placeholders_f = ",".join("?" * len(aluno_ids_turma))
        faltas_disc_rows = conn.execute(f"""
            SELECT bf.aluno_id, d.nome AS disc, SUM(bf.faltas) AS total
            FROM boletim_faltas bf JOIN disciplinas d ON d.id = bf.disciplina_id
            WHERE bf.ano = ? AND bf.aluno_id IN ({placeholders_f})
            GROUP BY bf.aluno_id, d.nome
        """, [ano] + aluno_ids_turma).fetchall()
        for r in faltas_disc_rows:
            faltas_disc_por_aluno.setdefault(r["aluno_id"], {})[r["disc"]] = r["total"]
    conn.close()

    # Médias da turma e ranking de faltas usam sempre a turma INTEIRA, mesmo que a página
    # gerada seja de um único aluno (comparativo precisa da turma toda como referência).
    medias_turma = _boletim_turma_medias_disciplina(dados_turma_completo)
    ranking_faltas = _boletim_ranking_faltas(dados_turma_completo)

    dados_turma = dados_turma_completo
    if aluno_id:
        dados_turma = [e for e in dados_turma_completo if e["id"] == aluno_id]
        if not dados_turma:
            return HTMLResponse(_pagina_simples("Erro", "<p>Aluno não encontrado nessa turma.</p>"))

    # Rótulo do período: reflete os trimestres que de fato têm dado importado (acumulado)
    trimestres_com_dado = sorted({t for e in dados_turma for disc in e["notas"].values() for t in disc.keys()})
    if not trimestres_com_dado:
        periodo_label = f"{ano}"
    elif len(trimestres_com_dado) == 1:
        periodo_label = f"{trimestres_com_dado[0]}º Trimestre {ano}"
    else:
        periodo_label = f"{trimestres_com_dado[0]}º ao {trimestres_com_dado[-1]}º Trimestre {ano} (acumulado)"

    paginas_html = ""
    for i, e in enumerate(dados_turma):
        saeb_geral = _boletim_saeb_nivel(e["media_geral"])
        situacao_html = f'<span style="color:{saeb_geral["color"]}; font-weight:700;">{saeb_geral["label"]}</span>' if saeb_geral else "—"

        # Colunas de trimestre dinâmicas — mostra T1/T2/(T3) separados, não só a média
        # acumulada, a pedido (24/08/2026). Usa os mesmos trimestres_com_dado da página toda,
        # pra manter as colunas iguais em todos os alunos do PDF.
        colunas_trim_header = "".join(f'<th style="padding:6px 8px; text-align:center;">{t}º Tri.</th>' for t in trimestres_com_dado)

        # Alerta de possível repetente: qualquer disciplina com média (T1+T2)/2 < 5,0
        # (mesma regra do dashboard) — 24/08/2026.
        disciplinas_risco = []

        linhas_disc = ""
        for disc in BOLETIM_ORDEM_DISCIPLINAS:
            trims = e["notas"].get(disc, {})
            e_numerica = disc in BOLETIM_DISC_NUMERICAS

            colunas_trim_html = ""
            for t in trimestres_com_dado:
                valor_trim = trims.get(t)
                if valor_trim is None:
                    colunas_trim_html += '<td style="padding:6px 8px; border-bottom:1px solid #eee; text-align:center; color:#bbb;">—</td>'
                elif valor_trim[0] is not None:
                    colunas_trim_html += f'<td style="padding:6px 8px; border-bottom:1px solid #eee; text-align:center;">{valor_trim[0]:.1f}</td>'
                else:
                    colunas_trim_html += f'<td style="padding:6px 8px; border-bottom:1px solid #eee; text-align:center;">{valor_trim[1] or "—"}</td>'

            if e_numerica:
                media_disc = e["medias_finais"].get(disc)
                nivel = _boletim_saeb_nivel(media_disc)
                nivel_html = f'<span style="color:{nivel["color"]}; font-weight:600;">{nivel["label"]}</span>' if nivel else "—"
                nota_str = f"{media_disc:.1f}" if media_disc is not None else "—"

                if 1 in trims and 2 in trims and trims[1][0] is not None and trims[2][0] is not None:
                    media_t1t2 = (trims[1][0] + trims[2][0]) / 2
                    if media_t1t2 < 5.0:
                        disciplinas_risco.append((disc, media_t1t2))
            else:
                # Educação Digital: categórica (PS/PA/PI) ou PD — pega o valor mais recente
                nota_texto_disc = None
                for t in sorted(trims.keys(), reverse=True):
                    if trims[t][1]:
                        nota_texto_disc = trims[t][1]
                        break
                nota_str = nota_texto_disc or "—"
                nivel_html = '<span style="color:#16a34a; font-weight:600;">Proficiência Adequada</span>' if nota_texto_disc == "PS" else (
                    '<span style="color:#ea580c; font-weight:600;">Proficiência Parcial</span>' if nota_texto_disc == "PA" else "—"
                )
            faltas_disc_str = "—" if not e_numerica else str(faltas_disc_por_aluno.get(e["id"], {}).get(disc, 0))
            linhas_disc += f"""<tr>
                <td style="padding:6px 8px; border-bottom:1px solid #eee;">{disc}</td>
                {colunas_trim_html}
                <td style="padding:6px 8px; border-bottom:1px solid #eee; text-align:center; font-weight:700;">{nota_str}</td>
                <td style="padding:6px 8px; border-bottom:1px solid #eee; text-align:center;">{nivel_html}</td>
                <td style="padding:6px 8px; border-bottom:1px solid #eee; text-align:center;">{faltas_disc_str}</td>
            </tr>"""

        alerta_repetente_html = ""
        if disciplinas_risco:
            itens_risco = ", ".join(f'{disc} ({media:.1f})' for disc, media in disciplinas_risco)
            alerta_repetente_html = f"""
            <div style="background:#fee2e2; border:1px solid #dc2626; border-radius:6px; padding:5px 10px; margin-bottom:8px; font-size:11px;">
                <strong style="color:#dc2626;">⚠ Possível repetente</strong>
                <span style="color:#7f1d1d;"> — média 1º+2º tri. abaixo de 5,0: {itens_risco}</span>
            </div>"""

        # Comparativo Aluno × Turma (só disciplinas numéricas) — em grid de 2 colunas
        # pra economizar espaço vertical (antes um gráfico embaixo do outro estourava a
        # página A4) — 26/08/2026, a pedido.
        comparativo_itens = ""
        for disc in BOLETIM_DISC_NUMERICAS:
            nota_aluno = e["medias_finais"].get(disc)
            nota_turma = medias_turma.get(disc)
            if nota_aluno is None:
                continue
            pct_aluno = min(100, (nota_aluno / 10) * 100)
            pct_turma = min(100, (nota_turma / 10) * 100) if nota_turma is not None else 0
            turma_str = f"{nota_turma:.1f}" if nota_turma is not None else "—"
            comparativo_itens += f"""
            <div>
                <div style="font-size:10px; font-weight:600; margin-bottom:2px;">{disc}</div>
                <div style="display:flex; align-items:center; gap:4px; margin-bottom:1px;">
                    <span style="font-size:9px; width:32px; color:#555;">Aluno</span>
                    <div style="flex:1; background:#eee; border-radius:3px; height:7px; position:relative;">
                        <div style="width:{pct_aluno}%; background:#2563eb; height:100%; border-radius:3px;"></div>
                    </div>
                    <span style="font-size:9px; width:22px; text-align:right;">{nota_aluno:.1f}</span>
                </div>
                <div style="display:flex; align-items:center; gap:4px;">
                    <span style="font-size:9px; width:32px; color:#555;">Turma</span>
                    <div style="flex:1; background:#eee; border-radius:3px; height:7px; position:relative;">
                        <div style="width:{pct_turma}%; background:#94a3b8; height:100%; border-radius:3px;"></div>
                    </div>
                    <span style="font-size:9px; width:22px; text-align:right;">{turma_str}</span>
                </div>
            </div>"""
        comparativo_html = f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:4px 20px;">{comparativo_itens}</div>'

        faltas_total = sum(v for v in e["faltas_por_trim"].values() if v)
        pos, total_turma = ranking_faltas.get(e["id"], (None, None))
        ranking_str = f"{pos}º na turma entre {total_turma}" if pos else "—"

        emoji = BOLETIM_EMOJI_EMOCIONAL.get(e["emocional"], "❔")
        label_emo = BOLETIM_LABEL_EMOCIONAL.get(e["emocional"], "Não informado")

        media_geral_str = f'{e["media_geral"]:.1f}' if e["media_geral"] is not None else "—"

        quebra = "page-break-after: always;" if i < len(dados_turma) - 1 else ""
        paginas_html += f"""
        <div style="{quebra} padding:14px 20px; font-family: Helvetica, Arial, sans-serif; max-width:760px; margin:0 auto; font-size:12px;">
            <div style="display:flex; align-items:center; justify-content:center; gap:12px; margin-bottom:4px;">
                <img src="/static/imagens/logo_walmir.png" alt="" style="height:38px; width:auto;">
                <div style="text-align:center;">
                    <div style="font-weight:700; font-size:14px;">E.M Walmir de Freitas Monteiro</div>
                    <div style="font-size:11px; color:#555;">Volta Redonda — RJ</div>
                </div>
            </div>
            <h2 style="text-align:center; margin:8px 0 2px 0; font-size:16px;">BOLETIM INDIVIDUAL</h2>
            <div style="text-align:center; font-size:12px; color:#555; margin-bottom:10px;">{periodo_label}</div>

            <table style="width:100%; margin-bottom:6px; font-size:12px;">
                <tr><td style="width:80px; color:#555;">Nome</td><td style="font-weight:700;">{e["nome"]}</td></tr>
                <tr><td style="color:#555;">Turma</td><td>{turma["nome"]}</td></tr>
            </table>
            <div style="margin-bottom:10px; font-size:12px;"><strong>Situação Geral:</strong> {situacao_html}</div>

            {alerta_repetente_html}

            <h3 style="font-size:12px; text-transform:uppercase; letter-spacing:0.5px; color:#555; border-bottom:1px solid #ddd; padding-bottom:3px; margin:10px 0 6px 0;">Desempenho por Disciplina — Escala SAEB</h3>
            <table style="width:100%; border-collapse:collapse; font-size:11px; margin-bottom:10px;">
                <thead><tr style="background:#f5f5f5;">
                    <th style="padding:4px 6px; text-align:left;">Disciplina</th>
                    {colunas_trim_header}
                    <th style="padding:4px 6px; text-align:center;">Média</th>
                    <th style="padding:4px 6px; text-align:center;">Nível SAEB</th>
                    <th style="padding:4px 6px; text-align:center;">Faltas</th>
                </tr></thead>
                <tbody>{linhas_disc}</tbody>
            </table>

            <h3 style="font-size:12px; text-transform:uppercase; letter-spacing:0.5px; color:#555; border-bottom:1px solid #ddd; padding-bottom:3px; margin:10px 0 6px 0;">📊 Comparativo: Nota do Aluno × Média da Turma</h3>
            <div style="margin-bottom:10px;">{comparativo_html}</div>

            <div style="display:flex; gap:10px; margin-bottom:10px;">
                <div style="flex:1; text-align:center; padding:8px; background:#f5f5f5; border-radius:6px;">
                    <div style="font-size:18px; font-weight:700;">{media_geral_str}</div>
                    <div style="font-size:10px; color:#555;">Média geral · {saeb_geral["label"] if saeb_geral else "—"}</div>
                </div>
                <div style="flex:1; text-align:center; padding:8px; background:#f5f5f5; border-radius:6px;">
                    <div style="font-size:18px; font-weight:700;">{faltas_total}</div>
                    <div style="font-size:10px; color:#555;">Total de faltas · {ranking_str}</div>
                </div>
                <div style="flex:1; text-align:center; padding:8px; background:#f5f5f5; border-radius:6px;">
                    <div style="font-size:18px;">{emoji}</div>
                    <div style="font-size:10px; color:#555;">Estado emocional · {label_emo}</div>
                </div>
            </div>

            <h3 style="font-size:11px; text-transform:uppercase; letter-spacing:0.5px; color:#555; border-bottom:1px solid #ddd; padding-bottom:3px; margin:8px 0 4px 0;">📏 Escala SAEB</h3>
            <table style="width:100%; border-collapse:collapse; font-size:9px; margin-bottom:6px;">
                <thead><tr style="background:#f5f5f5;">
                    <th style="padding:3px 5px; text-align:left;">Nível</th>
                    <th style="padding:3px 5px; text-align:center;">Faixa</th>
                    <th style="padding:3px 5px; text-align:left;">O que significa</th>
                </tr></thead>
                <tbody>
                    <tr><td style="padding:3px 5px; color:#dc2626; font-weight:600;">N1 — Insuficiente</td><td style="padding:3px 5px; text-align:center;">0,0–4,9</td><td style="padding:3px 5px;">Não demonstra os conhecimentos mínimos esperados; necessita intervenção urgente.</td></tr>
                    <tr><td style="padding:3px 5px; color:#ea580c; font-weight:600;">N2 — Básico</td><td style="padding:3px 5px; text-align:center;">5,0–6,9</td><td style="padding:3px 5px;">Conhecimentos elementares; ainda há lacunas importantes a superar.</td></tr>
                    <tr><td style="padding:3px 5px; color:#16a34a; font-weight:600;">N3 — Adequado</td><td style="padding:3px 5px; text-align:center;">7,0–8,4</td><td style="padding:3px 5px;">Atingiu o nível de aprendizagem esperado para o ano.</td></tr>
                    <tr><td style="padding:3px 5px; color:#6366f1; font-weight:600;">N4 — Avançado</td><td style="padding:3px 5px; text-align:center;">8,5–10,0</td><td style="padding:3px 5px;">Supera os conhecimentos esperados; excelente domínio das habilidades.</td></tr>
                </tbody>
            </table>

            <div style="font-size:9px; color:#888; border-top:1px solid #ddd; padding-top:5px; margin-top:8px; text-align:center;">
                Gerado automaticamente — E.M Walmir de Freitas Monteiro · {datetime.now().strftime("%d de %B de %Y")}
            </div>
        </div>
        """

    html_completo = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<title>Boletim Individual · {turma["nome"]}</title>
<style>
    @media print {{ @page {{ size: A4; margin: 10mm; }} body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}
    body {{ margin:0; background:#fff; }}
</style>
</head><body>{paginas_html}
<script>window.onload = () => window.print();</script>
</body></html>"""
    return HTMLResponse(html_completo)


@app.get("/boletim/relatorio-turma", response_class=HTMLResponse)
def boletim_relatorio_turma(ano: Optional[int] = None, turma_id: Optional[str] = None):
    # Liberado pra todos os docentes (antes só admin/gestão) — 25/08/2026
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)
    turma_id = int(turma_id) if turma_id and turma_id.strip().isdigit() else None

    conn = get_db()
    if not turma_id or not ano:
        # Sem turma selecionada (ex: clicou direto no menu) — mostra uma telinha de
        # seleção em vez de dar erro pedindo os campos (24/08/2026).
        turmas = conn.execute("SELECT id, nome, ano_letivo FROM turmas ORDER BY ano_letivo DESC, nome").fetchall()
        conn.close()
        opts_turmas = "".join(f'<option value="{t["id"]}" data-ano="{t["ano_letivo"]}">{t["nome"]} ({t["ano_letivo"]})</option>' for t in turmas)
        content = f"""
            <div class="page-header"><h1>📄 Relatório por Turma</h1></div>
            <form action="/boletim/relatorio-turma" method="get">
                <div style="display:flex; flex-wrap:wrap; gap:14px; align-items:flex-end;">
                    <label style="margin:0; flex:1 1 220px;">Turma
                        <select name="turma_id" id="sel-turma-rel" required onchange="document.getElementById('inp-ano-rel').value = this.selectedOptions[0].dataset.ano;">
                            <option value="">— selecione —</option>
                            {opts_turmas}
                        </select>
                    </label>
                    <label style="margin:0; flex:1 1 120px;">Ano
                        <input type="number" name="ano" id="inp-ano-rel" value="2026" required>
                    </label>
                </div>
                <div class="page-actions"><button type="submit" class="btn btn-primary">Ver relatório</button></div>
            </form>
        """
        return HTMLResponse(render_page("Relatório por Turma", content, active="boletim-relatorio-turma"))

    turma = conn.execute("SELECT * FROM turmas WHERE id = ?", (turma_id,)).fetchone()
    if not turma:
        conn.close()
        return RedirectResponse("/boletim", status_code=303)
    lista = _boletim_dados_turma_multitrimestre(conn, ano, turma_id)
    conn.close()

    paginas_html = ""
    for i, e in enumerate(lista):
        linhas_disc = ""
        for d in BOLETIM_ORDEM_DISCIPLINAS:
            trims = e["notas"].get(d, {})
            celulas_trim = ""
            for t in [1, 2, 3]:
                nota, nota_texto = trims.get(t, (None, None))
                if nota_texto:
                    celulas_trim += f'<td style="text-align:center;">{nota_texto}</td>'
                elif nota is not None:
                    cor = "color:#dc2626; font-weight:600;" if nota < 5 else ""
                    celulas_trim += f'<td style="text-align:center; {cor}">{nota:.1f}</td>'
                else:
                    celulas_trim += '<td style="text-align:center; color:#bbb;">—</td>'
            media_disc = e["medias_finais"].get(d)
            media_disc_str = f'{media_disc:.1f}' if media_disc is not None else "—"
            linhas_disc += f"""<tr>
                <td style="padding:5px 8px;">{d}</td>
                {celulas_trim}
                <td style="text-align:center; font-weight:700;">{media_disc_str}</td>
            </tr>"""

        saeb = e["saeb_geral"]
        saeb_html = f'<span style="background:{saeb["color"]}22; color:{saeb["color"]}; padding:2px 8px; border-radius:4px; font-weight:600;">{saeb["label"]}</span>' if saeb else "—"
        media_geral_str = f'{e["media_geral"]:.1f}' if e["media_geral"] is not None else "—"

        faltas_cols = "".join(
            f'<span style="margin-right:14px;"><strong>{t}º Tri:</strong> {e["faltas_por_trim"].get(t, "—")}</span>'
            for t in [1, 2, 3]
        )
        faltas_total = sum(v for v in e["faltas_por_trim"].values() if v)

        alertas = []
        if e["apoio"]:
            alertas.append("🤝 Precisa de apoio pedagógico")
        if e["alfab"]:
            alertas.append("📖 Dificuldade de alfabetização")
        if e["faltoso"]:
            alertas.append("🚫 Frequência preocupante")
        emo_str = f'{BOLETIM_EMOJI_EMOCIONAL.get(e["emocional"], "")} {BOLETIM_LABEL_EMOCIONAL.get(e["emocional"], "Não informado")}'
        alertas_html = "".join(f'<div style="margin-top:2px;">{al}</div>' for al in alertas)

        quebra_pagina = "page-break-after: always;" if i < len(lista) - 1 else ""
        paginas_html += f"""
        <div style="{quebra_pagina} padding-bottom:20px;">
            <div class="header">
                <img src="/static/imagens/logo_walmir.png" style="max-height:55px;" alt="Walmir">
                <div>
                    <h2 style="margin:0;">Boletim Escolar</h2>
                    <div style="color:#555; font-size:13px;">E.M. Walmir de Freitas Monteiro · {ano}</div>
                </div>
            </div>
            <table style="margin-bottom:12px;">
                <tr><td style="padding:4px 8px; width:25%;"><strong>Aluno:</strong></td><td style="padding:4px 8px;">{e["nome"]}</td>
                    <td style="padding:4px 8px; width:15%;"><strong>Nº:</strong></td><td style="padding:4px 8px;">{e["numero"] if e["numero"] is not None else "—"}</td></tr>
                <tr><td style="padding:4px 8px;"><strong>Turma:</strong></td><td style="padding:4px 8px;">{turma["nome"]}</td>
                    <td style="padding:4px 8px;"><strong>Ano letivo:</strong></td><td style="padding:4px 8px;">{ano}</td></tr>
            </table>
            <table>
                <thead><tr style="background:#f0f0f0;">
                    <th style="padding:5px 8px; text-align:left;">Disciplina</th>
                    <th style="padding:5px;">1º Tri</th><th style="padding:5px;">2º Tri</th><th style="padding:5px;">3º Tri</th>
                    <th style="padding:5px;">Média Final</th>
                </tr></thead>
                <tbody>{linhas_disc}</tbody>
            </table>
            <table style="margin-top:14px;">
                <tr><td style="padding:6px 8px; width:25%;"><strong>Média Geral</strong></td>
                    <td style="padding:6px 8px;"><span style="font-size:16px; font-weight:800;">{media_geral_str}</span> {saeb_html}</td></tr>
                <tr><td style="padding:6px 8px;"><strong>Faltas</strong></td>
                    <td style="padding:6px 8px;">{faltas_cols}<strong>Total: {faltas_total}</strong></td></tr>
                <tr><td style="padding:6px 8px;"><strong>Estado Emocional</strong></td><td style="padding:6px 8px;">{emo_str}</td></tr>
                {f'<tr><td style="padding:6px 8px; vertical-align:top;"><strong>Alertas</strong></td><td style="padding:6px 8px;">{alertas_html}</td></tr>' if alertas else ''}
                {f'<tr><td style="padding:6px 8px; vertical-align:top;"><strong>Observações</strong></td><td style="padding:6px 8px; font-size:12px;">{e["observacoes"]}</td></tr>' if e["observacoes"] else ''}
            </table>
        </div>"""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Boletim — Turma {turma['nome']}</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 24px; color:#222; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th, td {{ border:1px solid #ccc; }}
  .header {{ display:flex; align-items:center; gap:16px; margin-bottom:16px; }}
  @media print {{ .no-print {{ display:none; }} }}
</style></head>
<body>
  <div class="no-print" style="margin-bottom:16px;">
    <button onclick="window.print()">🖨️ Imprimir / Salvar PDF</button>
    <a href="/boletim/turma?trimestre=1&ano={ano}&turma_id={turma_id}">← Voltar</a>
    <span style="color:#777; margin-left:10px;">{len(lista)} boletim(ns) — um por página</span>
  </div>
  {paginas_html}
</body></html>"""
    return HTMLResponse(content=html)


@app.get("/boletim/estudantes", response_class=HTMLResponse)
def boletim_estudantes(request: Request, trimestre: Optional[int] = None, ano: Optional[int] = None,
                        turma_id: Optional[str] = None, ano_esc: Optional[str] = None, q: Optional[str] = None):
    # Liberado pra todos os docentes (antes só admin) — 25/08/2026
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    turma_id = int(turma_id) if turma_id and turma_id.strip().isdigit() else None
    conn = get_db()

    combinacoes = conn.execute("""
        SELECT trimestre, ano FROM boletim_medias GROUP BY trimestre, ano ORDER BY ano DESC, trimestre DESC
    """).fetchall()
    if not combinacoes:
        conn.close()
        content = '<div class="page-header"><h1>👥 Estudantes</h1></div><div class="empty">Nenhum dado importado ainda.</div>'
        return render_page("Estudantes", content, active="")
    if trimestre is None or ano is None:
        trimestre, ano = combinacoes[0]["trimestre"], combinacoes[0]["ano"]

    turmas = conn.execute("SELECT id, nome FROM turmas WHERE ano_letivo = ? ORDER BY nome", (ano,)).fetchall()
    lista = _boletim_enriquecer_alunos(conn, trimestre, ano, turma_id=turma_id)
    conn.close()

    if ano_esc and not turma_id:
        lista = [e for e in lista if _boletim_ano_da_turma(e["turma"]) == ano_esc]

    if q and q.strip():
        ql = q.strip().lower()
        lista = [e for e in lista if ql in e["nome"].lower()]

    lista.sort(key=lambda e: (e["turma"], e["nome"]))

    turma_opts = '<option value="">Todas as turmas</option>' + "".join(
        f'<option value="{t["id"]}"{" selected" if turma_id==t["id"] else ""}>Turma {t["nome"]}</option>' for t in turmas
    )
    ano_esc_opts = '<option value="">Todos os anos</option>' + "".join(
        f'<option value="{a}"{" selected" if ano_esc==a else ""}>{a} Ano</option>' for a in ["6°", "7°", "8°", "9°"]
    )
    trimestre_opts = "".join(
        f'<option value="{c["trimestre"]}:{c["ano"]}"{" selected" if c["trimestre"]==trimestre and c["ano"]==ano else ""}>{c["trimestre"]}º Trimestre {c["ano"]}</option>'
        for c in combinacoes
    )

    cabecalho_disc = "".join(f'<th style="padding:6px;">{d[:4]}.</th>' for d in BOLETIM_ORDEM_DISCIPLINAS)

    linhas = ""
    for e in lista:
        celulas = ""
        for d in BOLETIM_ORDEM_DISCIPLINAS:
            v = e["notas"].get(d)
            vt = e["notas_texto"].get(d)
            if vt:
                celulas += f'<td style="padding:6px; text-align:center;">{vt}</td>'
            elif v is not None:
                cor = "color:var(--red); font-weight:600;" if v < 5 else ""
                celulas += f'<td style="padding:6px; text-align:center; {cor}">{v:.1f}</td>'
            else:
                celulas += '<td style="padding:6px; text-align:center; color:var(--text-muted);">—</td>'
        saeb = e["saeb"]
        saeb_html = f'<span class="badge" style="background:{saeb["color"]}22; color:{saeb["color"]};">{saeb["label"]}</span>' if saeb else "—"
        emoji_emo = BOLETIM_EMOJI_EMOCIONAL.get(e["emocional"], "—")
        media_str = f'{e["media"]:.1f}' if e["media"] is not None else "—"
        marcas = []
        if e["risco_retencao"]:
            marcas.append('⚠️')
        if e["apoio"]:
            marcas.append('🤝')
        if e["alfab"]:
            marcas.append('📖')
        if e["faltoso"]:
            marcas.append('🚫')
        linhas += f"""<tr>
            <td style="padding:6px; white-space:nowrap;"><strong>{e["nome"]}</strong></td>
            <td style="padding:6px;">{e["turma"]}</td>
            {celulas}
            <td style="padding:6px; text-align:center; font-weight:700;">{media_str}</td>
            <td style="padding:6px; text-align:center;">{saeb_html}</td>
            <td style="padding:6px; text-align:center;">{emoji_emo}</td>
            <td style="padding:6px; text-align:center;">{" ".join(marcas) or "—"}</td>
            <td style="padding:6px; font-size:11px; max-width:220px;">{e["observacoes"] or ""}</td>
        </tr>"""

    content = f"""
        <div class="page-header">
            <h1>👥 Estudantes</h1>
            <p class="subtitle">{len(lista)} estudante(s) · {trimestre}º Trimestre {ano}</p>
        </div>
        <form method="get" action="/boletim/estudantes" style="background:var(--bg-subtle); padding:12px 16px; border-radius:8px; margin-bottom:18px;">
            <div style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">
                <label style="margin:0; flex:1 1 180px;">Trimestre
                    <select onchange="var v=this.value.split(':'); document.getElementById('f-trimestre').value=v[0]; document.getElementById('f-ano').value=v[1]; this.form.submit();">
                        {trimestre_opts}
                    </select>
                    <input type="hidden" id="f-trimestre" name="trimestre" value="{trimestre}">
                    <input type="hidden" id="f-ano" name="ano" value="{ano}">
                </label>
                <label style="margin:0; flex:1 1 160px;">Turma
                    <select name="turma_id" onchange="this.form.submit();">{turma_opts}</select>
                </label>
                <label style="margin:0; flex:1 1 160px;">Ano de escolaridade
                    <select name="ano_esc" onchange="this.form.submit();">{ano_esc_opts}</select>
                </label>
                <label style="margin:0; flex:1 1 200px;">Buscar por nome
                    <input type="text" name="q" value="{q or ''}" placeholder="nome do aluno">
                </label>
                <button type="submit" class="btn btn-primary">Filtrar</button>
            </div>
        </form>
        <div style="overflow-x:auto;">
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <thead><tr style="background:var(--bg-subtle);">
                <th style="padding:6px; text-align:left;">Aluno</th><th style="padding:6px; text-align:left;">Turma</th>
                {cabecalho_disc}
                <th style="padding:6px;">Média</th><th style="padding:6px;">SAEB</th>
                <th style="padding:6px;">Emoc.</th><th style="padding:6px;">Alertas</th><th style="padding:6px; text-align:left;">Observações</th>
            </tr></thead>
            <tbody>{linhas}</tbody>
        </table>
        </div>
    """
    return render_page("Estudantes", content, active="boletim-estudantes")


@app.get("/boletim/relatorio-geral", response_class=HTMLResponse)
def boletim_relatorio_geral(trimestre: Optional[int] = None, ano: Optional[int] = None, turma_id: Optional[str] = None, ano_esc: Optional[str] = None):
    # Liberado pra todos os docentes (antes só admin/gestão) — 25/08/2026
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)
    turma_id = int(turma_id) if turma_id and turma_id.strip().isdigit() else None

    conn = get_db()
    if trimestre is None or ano is None:
        # Sem parâmetro na URL (ex: clicou direto no menu) — usa o trimestre/ano mais
        # recente com dado importado, em vez de dar erro pedindo os campos (24/08/2026).
        combinacoes = conn.execute("SELECT trimestre, ano FROM boletim_medias GROUP BY trimestre, ano ORDER BY ano DESC, trimestre DESC").fetchall()
        if not combinacoes:
            conn.close()
            content = '<div class="page-header"><h1>📄 Relatório Geral</h1></div><div class="empty">Nenhum dado importado ainda. <a href="/boletim/importar-ecidade">Importar notas</a></div>'
            return render_page("Relatório Geral", content, active="boletim-relatorio-geral")
        trimestre, ano = combinacoes[0]["trimestre"], combinacoes[0]["ano"]

    turma_nome_filtro = None
    if turma_id:
        row_t = conn.execute("SELECT nome FROM turmas WHERE id = ?", (turma_id,)).fetchone()
        turma_nome_filtro = row_t["nome"] if row_t else None
    turmas_disponiveis = conn.execute("SELECT id, nome FROM turmas WHERE ano_letivo = ? ORDER BY nome", (ano,)).fetchall()
    enriquecidos = _boletim_enriquecer_alunos(conn, trimestre, ano, turma_id=turma_id)
    conn.close()

    if ano_esc and not turma_id:
        enriquecidos = [e for e in enriquecidos if _boletim_ano_da_turma(e["turma"]) == ano_esc]

    escopo_label = f"Turma {turma_nome_filtro}" if turma_nome_filtro else (f"{ano_esc} Ano" if ano_esc else "Escola toda")

    total = len(enriquecidos)
    medias_validas = [e["media"] for e in enriquecidos if e["media"] is not None]
    media_geral = sum(medias_validas) / len(medias_validas) if medias_validas else None
    saeb_geral = _boletim_saeb_nivel(media_geral)
    n_retencao = sum(1 for e in enriquecidos if e["risco_retencao"])
    n_apoio = sum(1 for e in enriquecidos if e["apoio"])
    n_alfab = sum(1 for e in enriquecidos if e["alfab"])
    n_faltoso = sum(1 for e in enriquecidos if e["faltoso"])

    from collections import defaultdict
    por_ano = defaultdict(list)
    for e in enriquecidos:
        por_ano[_boletim_ano_da_turma(e["turma"])].append(e)
    linhas_panorama = ""
    for label in ["6°", "7°", "8°", "9°"]:
        grupo = por_ano.get(label, [])
        if not grupo:
            continue
        medias_g = [e["media"] for e in grupo if e["media"] is not None]
        media_g = sum(medias_g) / len(medias_g) if medias_g else None
        saeb_g = _boletim_saeb_nivel(media_g)
        n_ret_g = sum(1 for e in grupo if e["risco_retencao"])
        linhas_panorama += f"""<tr>
            <td style="padding:6px 8px;">{label} Ano</td><td style="padding:6px 8px; text-align:center;">{len(grupo)}</td>
            <td style="padding:6px 8px; text-align:center; font-weight:700;">{f"{media_g:.1f}" if media_g is not None else "—"}</td>
            <td style="padding:6px 8px; text-align:center;">{saeb_g["label"] if saeb_g else "—"}</td>
            <td style="padding:6px 8px; text-align:center;">{n_ret_g}</td>
        </tr>"""

    ranking_disc = []
    for d in BOLETIM_DISC_NUMERICAS:
        vals = [e["notas"].get(d) for e in enriquecidos if e["notas"].get(d) is not None]
        if vals:
            m = sum(vals) / len(vals)
            ranking_disc.append((d, m, _boletim_saeb_nivel(m)))
    ranking_disc.sort(key=lambda x: x[1], reverse=True)
    linhas_disc = "".join(
        f'<tr><td style="padding:6px 8px;">{d}</td><td style="padding:6px 8px; text-align:center; font-weight:700;">{m:.1f}</td>'
        f'<td style="padding:6px 8px; text-align:center;">{lvl["label"] if lvl else "—"}</td></tr>'
        for d, m, lvl in ranking_disc
    )

    raca_count = {}
    for e in enriquecidos:
        r = e["raca"] or "Não informado"
        raca_count[r] = raca_count.get(r, 0) + 1
    linhas_raca = "".join(
        f'<tr><td style="padding:6px 8px;">{r}</td><td style="padding:6px 8px; text-align:center;">{c}</td>'
        f'<td style="padding:6px 8px; text-align:center;">{c/total*100:.1f}%</td></tr>'
        for r, c in sorted(raca_count.items(), key=lambda x: -x[1])
    )

    saeb_geral_html = f'{saeb_geral["label"]}' if saeb_geral else "—"
    media_geral_str = f'{media_geral:.1f}' if media_geral is not None else "—"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Relatório Geral de Gestão — {trimestre}º Tri {ano}</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 24px; color:#222; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; margin-bottom:20px; }}
  th, td {{ border:1px solid #ccc; }}
  th {{ background:#f0f0f0; padding:6px 8px; text-align:left; }}
  .header {{ display:flex; align-items:center; gap:16px; margin-bottom:20px; }}
  .stat {{ display:inline-block; border:1px solid #ccc; border-radius:6px; padding:10px 16px; margin:0 10px 10px 0; }}
  .stat-val {{ font-size:22px; font-weight:800; }}
  .stat-lbl {{ font-size:11px; color:#666; text-transform:uppercase; }}
  h3 {{ margin-top:26px; margin-bottom:8px; }}
  @media print {{ .no-print {{ display:none; }} }}
</style></head>
<body>
  <div class="header">
    <img src="/static/imagens/logo_walmir.png" style="max-height:60px;" alt="Walmir">
    <div>
      <h2 style="margin:0;">Relatório Geral de Gestão</h2>
      <div style="color:#555; font-size:13px;">E.M. Walmir de Freitas Monteiro · {escopo_label} · {trimestre}º Trimestre {ano} · {total} estudante(s)</div>
    </div>
  </div>
  <div class="no-print" style="margin-bottom:16px;">
    <button onclick="window.print()">🖨️ Imprimir / Salvar PDF</button>
    <a href="/boletim/dashboard?trimestre={trimestre}&ano={ano}">← Voltar ao Dashboard</a>
    <form method="get" action="/boletim/relatorio-geral" style="display:inline-flex; gap:8px; margin-left:14px; vertical-align:middle;">
        <input type="hidden" name="trimestre" value="{trimestre}"><input type="hidden" name="ano" value="{ano}">
        <select name="turma_id" onchange="this.form.submit();">
            <option value="">Escola toda</option>
            {"".join(f'<option value="{t["id"]}"{" selected" if turma_id==t["id"] else ""}>Turma {t["nome"]}</option>' for t in turmas_disponiveis)}
        </select>
        <select name="ano_esc" onchange="this.form.submit();">
            <option value="">Todos os anos</option>
            {"".join(f'<option value="{a}"{" selected" if ano_esc==a else ""}>{a} Ano</option>' for a in ["6°","7°","8°","9°"])}
        </select>
    </form>
  </div>

  <div>
    <div class="stat"><div class="stat-val">{total}</div><div class="stat-lbl">Estudantes</div></div>
    <div class="stat"><div class="stat-val">{media_geral_str}</div><div class="stat-lbl">Média Geral ({saeb_geral_html})</div></div>
    <div class="stat"><div class="stat-val">{n_retencao}</div><div class="stat-lbl">Risco de Retenção</div></div>
    <div class="stat"><div class="stat-val">{n_apoio}</div><div class="stat-lbl">Precisam de Apoio</div></div>
    <div class="stat"><div class="stat-val">{n_alfab}</div><div class="stat-lbl">Dif. Alfabetização</div></div>
    <div class="stat"><div class="stat-val">{n_faltoso}</div><div class="stat-lbl">Faltosos</div></div>
  </div>

  <h3>Panorama por Ano de Escolaridade</h3>
  <table><thead><tr><th>Ano</th><th>Alunos</th><th>Média</th><th>SAEB</th><th>Risco Retenção</th></tr></thead>
    <tbody>{linhas_panorama}</tbody></table>

  <h3>Desempenho por Disciplina</h3>
  <table><thead><tr><th>Disciplina</th><th>Média</th><th>SAEB</th></tr></thead>
    <tbody>{linhas_disc}</tbody></table>

  <h3>Distribuição por Raça/Etnia</h3>
  <table><thead><tr><th>Raça/Etnia</th><th>Alunos</th><th>%</th></tr></thead>
    <tbody>{linhas_raca}</tbody></table>
</body></html>"""
    return HTMLResponse(content=html)


@app.get("/boletim/comparativo", response_class=HTMLResponse)
def boletim_comparativo(request: Request, ano: Optional[int] = None,
                         trimestre_a: Optional[int] = None, trimestre_b: Optional[int] = None,
                         turma_id: Optional[str] = None, ano_esc: Optional[str] = None):
    # Liberado pra todos os docentes (antes só admin) — 25/08/2026
    prof = get_current_professor(request)
    if not prof:
        return RedirectResponse("/login", status_code=303)
    turma_id = int(turma_id) if turma_id and turma_id.strip().isdigit() else None
    conn = get_db()

    trimestres_disponiveis = conn.execute("""
        SELECT DISTINCT trimestre, ano FROM boletim_medias ORDER BY ano DESC, trimestre DESC
    """).fetchall()
    anos_disponiveis = sorted({r["ano"] for r in trimestres_disponiveis}, reverse=True)
    if not anos_disponiveis:
        conn.close()
        content = '<div class="page-header"><h1>🔄 Comparativo entre Trimestres</h1></div><div class="empty">Nenhum dado importado ainda.</div>'
        return render_page("Comparativo", content, active="boletim-comparativo")
    if ano is None:
        ano = anos_disponiveis[0]

    trims_do_ano = sorted({r["trimestre"] for r in trimestres_disponiveis if r["ano"] == ano})
    if len(trims_do_ano) < 2:
        conn.close()
        turmas_vazio = ""
        content = f"""
            <div class="page-header"><h1>🔄 Comparativo entre Trimestres</h1></div>
            <div class="empty">Esse ano ({ano}) só tem {len(trims_do_ano)} trimestre com dados lançados — a comparação
            fica disponível assim que o 2º trimestre for importado ou lançado pela tela de Análise.</div>
        """
        return render_page("Comparativo", content, active="boletim-comparativo")

    if trimestre_b is None or trimestre_b not in trims_do_ano:
        trimestre_b = trims_do_ano[-1]
    if trimestre_a is None or trimestre_a not in trims_do_ano or trimestre_a == trimestre_b:
        anteriores = [t for t in trims_do_ano if t < trimestre_b]
        trimestre_a = anteriores[-1] if anteriores else trims_do_ano[0]

    turmas = conn.execute("SELECT id, nome FROM turmas WHERE ano_letivo = ? ORDER BY nome", (ano,)).fetchall()

    dados_a = {e["id"]: e for e in _boletim_enriquecer_alunos(conn, trimestre_a, ano, turma_id=turma_id)}
    dados_b = {e["id"]: e for e in _boletim_enriquecer_alunos(conn, trimestre_b, ano, turma_id=turma_id)}
    conn.close()

    if ano_esc and not turma_id:
        dados_a = {k: v for k, v in dados_a.items() if _boletim_ano_da_turma(v["turma"]) == ano_esc}
        dados_b = {k: v for k, v in dados_b.items() if _boletim_ano_da_turma(v["turma"]) == ano_esc}

    alunos_ids = set(dados_a) | set(dados_b)

    # --- Por disciplina: média A vs B, quem subiu/desceu ---
    linhas_disc = []
    for d in BOLETIM_DISC_NUMERICAS:
        vals_a = [e["notas"].get(d) for e in dados_a.values() if e["notas"].get(d) is not None]
        vals_b = [e["notas"].get(d) for e in dados_b.values() if e["notas"].get(d) is not None]
        media_a = sum(vals_a) / len(vals_a) if vals_a else None
        media_b = sum(vals_b) / len(vals_b) if vals_b else None
        delta = (media_b - media_a) if (media_a is not None and media_b is not None) else None
        linhas_disc.append((d, media_a, media_b, delta))
    linhas_disc.sort(key=lambda x: (x[3] is None, -(x[3] or 0)))

    def seta(delta):
        if delta is None:
            return '<span style="color:var(--text-muted);">—</span>'
        if delta > 0.05:
            return f'<span style="color:var(--green); font-weight:700;">▲ +{delta:.1f}</span>'
        if delta < -0.05:
            return f'<span style="color:var(--red); font-weight:700;">▼ {delta:.1f}</span>'
        return '<span style="color:var(--text-muted);">≈ estável</span>'

    disc_rows_html = "".join(
        f'<tr><td style="padding:6px 10px;">{d}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{f"{ma:.1f}" if ma is not None else "—"}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{f"{mb:.1f}" if mb is not None else "—"}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{seta(delta)}</td></tr>'
        for d, ma, mb, delta in linhas_disc
    )

    # --- Por aluno: quem mais melhorou / quem mais piorou (média geral) ---
    deltas_alunos = []
    for aid in alunos_ids:
        ea, eb = dados_a.get(aid), dados_b.get(aid)
        if not ea or not eb or ea["media"] is None or eb["media"] is None:
            continue
        deltas_alunos.append({
            "nome": eb["nome"], "turma": eb["turma"], "media_a": ea["media"], "media_b": eb["media"],
            "delta": eb["media"] - ea["media"],
        })
    melhoraram = sorted([d for d in deltas_alunos if d["delta"] > 0], key=lambda x: -x["delta"])[:15]
    pioraram = sorted([d for d in deltas_alunos if d["delta"] < 0], key=lambda x: x["delta"])[:15]

    def lista_delta_html(lista, cor):
        if not lista:
            return '<p style="font-size:12px; color:var(--text-muted);">Nenhum aluno nessa situação.</p>'
        return "".join(
            f'<div style="display:flex; justify-content:space-between; padding:4px 8px; font-size:12px; background:var(--bg-subtle); border-radius:4px; margin-bottom:3px;">'
            f'<span>{d["nome"]} <span style="color:var(--text-muted);">· {d["turma"]}</span></span>'
            f'<strong style="color:{cor};">{d["media_a"]:.1f} → {d["media_b"]:.1f} ({"+" if d["delta"]>0 else ""}{d["delta"]:.1f})</strong></div>'
            for d in lista
        )

    melhoraram_html = lista_delta_html(melhoraram, "var(--green)")
    pioraram_html = lista_delta_html(pioraram, "var(--red)")

    # --- Por raça: média A vs B ---
    def media_por_raca(dados):
        agrup = {}
        for e in dados.values():
            if e["media"] is None:
                continue
            r = e["raca"] or "Não informado"
            agrup.setdefault(r, []).append(e["media"])
        return {r: sum(v) / len(v) for r, v in agrup.items()}

    raca_a, raca_b = media_por_raca(dados_a), media_por_raca(dados_b)
    todas_racas = sorted(set(raca_a) | set(raca_b), key=lambda r: -(raca_b.get(r, raca_a.get(r, 0))))
    raca_rows_html = "".join(
        f'<tr><td style="padding:6px 10px;">{r}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{f"{raca_a[r]:.1f}" if r in raca_a else "—"}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{f"{raca_b[r]:.1f}" if r in raca_b else "—"}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{seta((raca_b[r]-raca_a[r]) if (r in raca_a and r in raca_b) else None)}</td></tr>'
        for r in todas_racas
    )

    # --- Por distorção idade-série: média A vs B ---
    def media_por_distorcao(dados):
        grupos = {"Idade adequada": [], "Com distorção": []}
        for e in dados.values():
            if e["media"] is None or not e.get("data_nascimento"):
                continue
            try:
                ano_esc = int(str(e["turma"])[0])
            except (ValueError, IndexError):
                continue
            idade = _calcular_idade_referencia(e["data_nascimento"], ano)
            if idade is None:
                continue
            idade_esperada = ano_esc + 5
            grupo = "Com distorção" if idade > idade_esperada else "Idade adequada"
            grupos[grupo].append(e["media"])
        return {g: (sum(v) / len(v) if v else None) for g, v in grupos.items()}

    dist_a, dist_b = media_por_distorcao(dados_a), media_por_distorcao(dados_b)
    dist_rows_html = "".join(
        f'<tr><td style="padding:6px 10px;">{g}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{f"{dist_a[g]:.1f}" if dist_a.get(g) is not None else "—"}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{f"{dist_b[g]:.1f}" if dist_b.get(g) is not None else "—"}</td>'
        f'<td style="padding:6px 10px; text-align:center;">{seta((dist_b[g]-dist_a[g]) if (dist_a.get(g) is not None and dist_b.get(g) is not None) else None)}</td></tr>'
        for g in ["Idade adequada", "Com distorção"]
    )
    tem_dados_idade = any(v is not None for v in list(dist_a.values()) + list(dist_b.values()))

    # --- Seletores ---
    trim_a_opts = "".join(f'<option value="{t}"{" selected" if t==trimestre_a else ""}>{t}º Trimestre</option>' for t in trims_do_ano)
    trim_b_opts = "".join(f'<option value="{t}"{" selected" if t==trimestre_b else ""}>{t}º Trimestre</option>' for t in trims_do_ano)
    turma_opts = '<option value="">Escola toda</option>' + "".join(
        f'<option value="{t["id"]}"{" selected" if turma_id==t["id"] else ""}>Turma {t["nome"]}</option>' for t in turmas
    )
    ano_esc_opts_comp = '<option value="">Todos os anos</option>' + "".join(
        f'<option value="{a}"{" selected" if ano_esc==a else ""}>{a} Ano</option>' for a in ["6°", "7°", "8°", "9°"]
    )

    content = f"""
        <div class="page-header">
            <h1>🔄 Comparativo entre Trimestres</h1>
            <p class="subtitle">Comparando {trimestre_a}º → {trimestre_b}º Trimestre de {ano}</p>
        </div>
        <form method="get" action="/boletim/comparativo" style="background:var(--bg-subtle); padding:12px 16px; border-radius:8px; margin-bottom:18px;">
            <div style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">
                <label style="margin:0; flex:1 1 140px;">De<select name="trimestre_a">{trim_a_opts}</select></label>
                <label style="margin:0; flex:1 1 140px;">Para<select name="trimestre_b">{trim_b_opts}</select></label>
                <label style="margin:0; flex:1 1 160px;">Turma<select name="turma_id">{turma_opts}</select></label>
                <label style="margin:0; flex:1 1 160px;">Ano de escolaridade<select name="ano_esc">{ano_esc_opts_comp}</select></label>
                <input type="hidden" name="ano" value="{ano}">
                <button type="submit" class="btn btn-primary">Comparar</button>
            </div>
        </form>

        <div class="card" style="margin-bottom:18px; padding:0; overflow:hidden;">
            <div style="padding:14px 16px; border-bottom:1px solid var(--border); font-weight:700; font-size:14px;">📐 Por Disciplina</div>
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <thead><tr style="background:var(--bg-subtle);">
                    <th style="padding:6px 10px; text-align:left;">Disciplina</th>
                    <th style="padding:6px 10px;">{trimestre_a}º Tri</th><th style="padding:6px 10px;">{trimestre_b}º Tri</th>
                    <th style="padding:6px 10px;">Variação</th>
                </tr></thead>
                <tbody>{disc_rows_html}</tbody>
            </table>
        </div>

        <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(280px, 1fr)); gap:16px; margin-bottom:18px;">
            <div class="card">
                <h3 style="margin-top:0; color:var(--green);">📈 Quem Mais Melhorou</h3>
                {melhoraram_html}
            </div>
            <div class="card">
                <h3 style="margin-top:0; color:var(--red);">📉 Quem Merece Atenção (caiu)</h3>
                {pioraram_html}
            </div>
        </div>

        <div class="card" style="margin-bottom:18px; padding:0; overflow:hidden;">
            <div style="padding:14px 16px; border-bottom:1px solid var(--border); font-weight:700; font-size:14px;">🧑🏾 Por Raça/Etnia</div>
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <thead><tr style="background:var(--bg-subtle);">
                    <th style="padding:6px 10px; text-align:left;">Raça/Etnia</th>
                    <th style="padding:6px 10px;">{trimestre_a}º Tri</th><th style="padding:6px 10px;">{trimestre_b}º Tri</th>
                    <th style="padding:6px 10px;">Variação</th>
                </tr></thead>
                <tbody>{raca_rows_html}</tbody>
            </table>
        </div>

        <div class="card" style="padding:0; overflow:hidden;">
            <div style="padding:14px 16px; border-bottom:1px solid var(--border); font-weight:700; font-size:14px;">📏 Por Distorção Idade-Série</div>
            {f'''<table style="width:100%; border-collapse:collapse; font-size:13px;">
                <thead><tr style="background:var(--bg-subtle);">
                    <th style="padding:6px 10px; text-align:left;">Grupo</th>
                    <th style="padding:6px 10px;">{trimestre_a}º Tri</th><th style="padding:6px 10px;">{trimestre_b}º Tri</th>
                    <th style="padding:6px 10px;">Variação</th>
                </tr></thead>
                <tbody>{dist_rows_html}</tbody>
            </table>''' if tem_dados_idade else '<div style="padding:14px 16px; color:var(--text-muted); font-size:13px;">Nenhum aluno desse recorte tem data de nascimento cadastrada — cadastre em Turmas → aluno → editar pra habilitar essa comparação.</div>'}
        </div>
    """
    return render_page("Comparativo entre Trimestres", content, active="boletim-comparativo")


@app.get("/boletim/analise", response_class=HTMLResponse)
def boletim_analise_form(request: Request, trimestre: Optional[int] = None, ano: Optional[int] = None,
                          turma_id: Optional[str] = None, disciplina_id: Optional[str] = None):
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)
    eh_admin = bool(prof.get("is_admin") or prof.get("is_gestor"))
    turma_id = int(turma_id) if turma_id and turma_id.strip().isdigit() else None
    disciplina_id = int(disciplina_id) if disciplina_id and disciplina_id.strip().isdigit() else None

    conn = get_db()

    if trimestre is None or ano is None:
        ultimo = conn.execute("SELECT trimestre, ano FROM boletim_medias ORDER BY ano DESC, trimestre DESC LIMIT 1").fetchone()
        trimestre, ano = (ultimo["trimestre"], ultimo["ano"]) if ultimo else (1, 2026)

    # Opções de turma/disciplina — admin vê tudo, professor só o que está mapeado pra ele
    if eh_admin:
        opcoes = conn.execute("""
            SELECT DISTINCT t.id AS turma_id, t.nome AS turma_nome, d.id AS disciplina_id, d.nome AS disciplina_nome
            FROM turmas t CROSS JOIN disciplinas d WHERE t.ano_letivo = ? AND d.nome != 'Geral'
            ORDER BY t.nome, d.nome
        """, (ano,)).fetchall()
    else:
        opcoes = conn.execute("""
            SELECT DISTINCT t.id AS turma_id, t.nome AS turma_nome, d.id AS disciplina_id, d.nome AS disciplina_nome
            FROM boletim_professor_turma bpt
            JOIN turmas t ON t.id = bpt.turma_id
            JOIN disciplinas d ON d.id = bpt.disciplina_id
            WHERE bpt.professor_id = ? AND t.ano_letivo = ?
            ORDER BY t.nome, d.nome
        """, (prof["id"], ano)).fetchall()

    if not opcoes:
        conn.close()
        msg = "Nenhuma turma/disciplina foi vinculada ao seu usuário ainda. Peça pra um admin te vincular, ou lance através da importação da planilha." if not eh_admin else "Nenhuma turma cadastrada ainda."
        content = f'<div class="page-header"><h1>📝 Análise — Conselho de Classe</h1></div><div class="empty">{msg}</div>'
        return render_page("Análise", content, active="boletim-analise")

    if turma_id is None or disciplina_id is None:
        turma_id, disciplina_id = opcoes[0]["turma_id"], opcoes[0]["disciplina_id"]

    # Se for professor comum, confere que ele tem permissão pra essa combinação específica
    if not eh_admin and not any(o["turma_id"] == turma_id and o["disciplina_id"] == disciplina_id for o in opcoes):
        turma_id, disciplina_id = opcoes[0]["turma_id"], opcoes[0]["disciplina_id"]

    turma = conn.execute("SELECT * FROM turmas WHERE id = ?", (turma_id,)).fetchone()
    disciplina = conn.execute("SELECT * FROM disciplinas WHERE id = ?", (disciplina_id,)).fetchone()

    alunos = conn.execute("SELECT id, nome, numero FROM alunos WHERE turma_id = ? ORDER BY numero, nome", (turma_id,)).fetchall()
    analises_existentes = {}
    for r in conn.execute("""SELECT * FROM boletim_analise WHERE trimestre=? AND ano=? AND disciplina_id=?
                              AND aluno_id IN (SELECT id FROM alunos WHERE turma_id=?)""",
                           (trimestre, ano, disciplina_id, turma_id)).fetchall():
        analises_existentes[r["aluno_id"]] = r
    conn.close()

    opcoes_opts = "".join(
        f'<option value="{o["turma_id"]}:{o["disciplina_id"]}"{" selected" if o["turma_id"]==turma_id and o["disciplina_id"]==disciplina_id else ""}>'
        f'Turma {o["turma_nome"]} — {o["disciplina_nome"]}</option>'
        for o in opcoes
    )

    linhas = ""
    for a in alunos:
        ex = analises_existentes.get(a["id"])
        emo_atual = ex["emocional"] if ex else ""
        apoio_atual = bool(ex["apoio"]) if ex else False
        alfab_atual = bool(ex["alfabetizacao"]) if ex else False
        faltoso_atual = bool(ex["faltoso"]) if ex else False
        obs_atual = ex["observacao"] if ex else ""

        emo_opts = "".join(
            f'<option value="{v}"{" selected" if emo_atual==v else ""}>{lbl}</option>'
            for v, lbl in [("", "—"), ("bem", "😊 Bem"), ("oscilando", "😐 Oscilando"), ("fragilizado", "😟 Fragilizado")]
        )
        linhas += f"""<tr>
            <td style="padding:6px;">{a["numero"] or "—"}</td>
            <td style="padding:6px; white-space:nowrap;">{a["nome"]}<input type="hidden" name="aluno_id" value="{a["id"]}"></td>
            <td style="padding:6px;"><select name="emocional_{a["id"]}">{emo_opts}</select></td>
            <td style="padding:6px; text-align:center;"><input type="checkbox" name="apoio_{a["id"]}"{" checked" if apoio_atual else ""}></td>
            <td style="padding:6px; text-align:center;"><input type="checkbox" name="alfab_{a["id"]}"{" checked" if alfab_atual else ""}></td>
            <td style="padding:6px; text-align:center;"><input type="checkbox" name="faltoso_{a["id"]}"{" checked" if faltoso_atual else ""}></td>
            <td style="padding:6px;"><input type="text" name="obs_{a["id"]}" value="{(obs_atual or "").replace(chr(34), "&quot;")}" style="width:100%; margin:0;" placeholder="Observação (opcional)"></td>
        </tr>"""

    content = f"""
        <div class="page-header">
            <h1>📝 Análise — Conselho de Classe</h1>
            <p class="subtitle">Turma {turma['nome']} · {disciplina['nome']} · {trimestre}º Trimestre {ano}</p>
        </div>
        <form method="get" action="/boletim/analise" style="background:var(--bg-subtle); padding:12px 16px; border-radius:8px; margin-bottom:18px;">
            <div style="display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;">
                <label style="margin:0; flex:0 0 140px;">Trimestre
                    <select name="trimestre" onchange="this.form.submit();">
                        <option value="1"{' selected' if trimestre==1 else ''}>1º Trimestre</option>
                        <option value="2"{' selected' if trimestre==2 else ''}>2º Trimestre</option>
                        <option value="3"{' selected' if trimestre==3 else ''}>3º Trimestre</option>
                    </select>
                </label>
                <input type="hidden" name="ano" value="{ano}">
                <label style="margin:0; flex:1 1 320px;">Turma / Disciplina
                    <select onchange="var v=this.value.split(':'); this.form.turma_id.value=v[0]; this.form.disciplina_id.value=v[1]; this.form.submit();">
                        {opcoes_opts}
                    </select>
                    <input type="hidden" name="turma_id" value="{turma_id}">
                    <input type="hidden" name="disciplina_id" value="{disciplina_id}">
                </label>
            </div>
        </form>
        <form method="post" action="/boletim/analise/salvar">
            <input type="hidden" name="trimestre" value="{trimestre}">
            <input type="hidden" name="ano" value="{ano}">
            <input type="hidden" name="turma_id" value="{turma_id}">
            <input type="hidden" name="disciplina_id" value="{disciplina_id}">
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <thead><tr style="background:var(--bg-subtle);">
                    <th style="padding:6px; text-align:left;">Nº</th><th style="padding:6px; text-align:left;">Aluno</th>
                    <th style="padding:6px;">Estado emocional</th><th style="padding:6px;">Apoio</th>
                    <th style="padding:6px;">Alfabetização</th><th style="padding:6px;">Faltoso</th><th style="padding:6px; text-align:left;">Observação</th>
                </tr></thead>
                <tbody>{linhas}</tbody>
            </table>
            <div class="page-actions">
                <button type="submit" class="btn btn-primary">💾 Salvar</button>
            </div>
        </form>
    """
    return render_page("Análise", content, active="boletim-analise")


@app.post("/boletim/analise/salvar")
async def boletim_analise_salvar(request: Request):
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)

    form = await request.form()
    trimestre = int(form.get("trimestre"))
    ano = int(form.get("ano"))
    turma_id = int(form.get("turma_id"))
    disciplina_id = int(form.get("disciplina_id"))
    alunos_ids = form.getlist("aluno_id")

    conn = get_db()
    n_salvos = 0
    for aid_str in alunos_ids:
        aid = int(aid_str)
        emocional = form.get(f"emocional_{aid}") or None
        apoio = form.get(f"apoio_{aid}") is not None
        alfab = form.get(f"alfab_{aid}") is not None
        faltoso = form.get(f"faltoso_{aid}") is not None
        obs = (form.get(f"obs_{aid}") or "").strip() or None
        conn.execute("""
            INSERT INTO boletim_analise (aluno_id, disciplina_id, professor_id, trimestre, ano, emocional, apoio, alfabetizacao, faltoso, observacao, atualizado_em)
            VALUES (?,?,?,?,?,?,?,?,?,?, CURRENT_TIMESTAMP)
            ON CONFLICT(aluno_id, disciplina_id, trimestre, ano) DO UPDATE SET
                professor_id = excluded.professor_id, emocional = excluded.emocional,
                apoio = excluded.apoio, alfabetizacao = excluded.alfabetizacao, faltoso = excluded.faltoso,
                observacao = excluded.observacao, atualizado_em = CURRENT_TIMESTAMP
        """, (aid, disciplina_id, prof["id"], trimestre, ano, emocional, apoio, alfab, faltoso, obs))
        n_salvos += 1
    conn.commit()
    conn.close()

    return RedirectResponse(
        f"/boletim/analise?trimestre={trimestre}&ano={ano}&turma_id={turma_id}&disciplina_id={disciplina_id}",
        status_code=303
    )



@app.get("/turmas/{turma_id}", response_class=HTMLResponse)
def ver_turma(request: Request, turma_id: int):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    prof = get_current_professor(request)
    is_admin = prof and prof["is_admin"]
    conn = get_db()
    turma = conn.execute("SELECT * FROM turmas WHERE id = ?", (turma_id,)).fetchone()
    if not turma:
        conn.close()
        return HTMLResponse(render_page("Não encontrada", '<h1>Turma não encontrada</h1><p><a href="/turmas">← Voltar</a></p>', active="turmas"), status_code=404)
    alunos = conn.execute("SELECT * FROM alunos WHERE turma_id = ? ORDER BY numero, nome", (turma_id,)).fetchall()
    proximo_numero = conn.execute("SELECT COALESCE(MAX(numero), 0) + 1 AS n FROM alunos WHERE turma_id = ?", (turma_id,)).fetchone()["n"]
    conn.close()

    if alunos:
        alunos_html = ""
        for a in alunos:
            num = a["numero"] if a["numero"] else "—"
            extras = []
            if a["raca"]: extras.append(a["raca"])
            if a["email"]: extras.append(a["email"])
            if a["data_nascimento"]: extras.append(f'nasc. {format_data_br(a["data_nascimento"])}')
            extra_line = f'<div style="font-size:12px; color:var(--text-muted); margin-top:2px;">{" · ".join(extras)}</div>' if extras else ""
            if is_admin:
                nome_escapado = a["nome"].replace("'", "\\'")
                acoes = (
                    f'<div style="font-size:11px; margin-top:6px;">'
                    f'<a href="/alunos/{a["id"]}/editar" style="color:var(--text-muted);">Editar</a>'
                    f'<span style="color:var(--text-subtle);"> · </span>'
                    f'<a href="/alunos/{a["id"]}/transferir" style="color:var(--text-muted);">Transferir</a>'
                    f'<span style="color:var(--text-subtle);"> · </span>'
                    f'<form action="/alunos/{a["id"]}/deletar" method="post" style="display:inline; margin:0;" '
                    f"onsubmit=\"return confirm('Excluir {nome_escapado}? Se o aluno tiver entregas registradas, você poderá forçar a exclusão na próxima tela.');\">"
                    f'<button type="submit" style="background:none; border:none; padding:0; color:var(--red); cursor:pointer; font-size:inherit; font-family:inherit;">Excluir</button>'
                    f'</form>'
                    f'</div>'
                )
            else:
                acoes = ""
            alunos_html += f'<div class="student-row"><div class="numero">{num}</div><div>{a["nome"]}{extra_line}{acoes}</div><div class="codigo">{a["codigo_unico"]}</div></div>'
    else:
        alunos_html = '<div class="empty">Nenhum aluno cadastrado nesta turma ainda.</div>'

    racas_options = '<option value="">Não informada</option>' + "".join(f'<option value="{r}">{r}</option>' for r in RACAS)

    if is_admin:
        excluir_turma_btn = (
            f'<div class="page-actions"><form action="/turmas/{turma_id}/deletar" method="post" style="margin:0;" '
            f"onsubmit=\"return confirm('Excluir esta turma?\\n\\nIsso removerá: alunos, aplicações desta turma, respostas e entregas associadas.') && "
            f"confirm('TEM CERTEZA? Esta ação é IRREVERSÍVEL e não pode ser desfeita.');\">"
            f'<button type="submit" class="btn" style="background:var(--red); color:white; border-color:var(--red);">🗑️ Excluir turma</button>'
            f'</form></div>'
        )
        form_adicionar = f"""
            <h2>Adicionar aluno</h2>
            <form action="/turmas/{turma_id}/alunos" method="post">
                <div style="display:grid; grid-template-columns: 100px 1fr; gap:12px;">
                    <label>Número<input type="number" name="numero" value="{proximo_numero}" min="1"></label>
                    <label>Nome<input type="text" name="nome" required placeholder="Nome completo"></label>
                </div>
                <div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(140px, 1fr)); gap:12px;">
                    <label>Raça<select name="raca">{racas_options}</select></label>
                    <label>E-mail<input type="email" name="email" placeholder="aluno@email.com"></label>
                    <label>Data de nascimento<input type="date" name="data_nascimento"></label>
                </div>
                <div class="page-actions">
                    <button type="submit" class="btn btn-primary">Adicionar</button>
                </div>
            </form>
        """
    else:
        excluir_turma_btn = ""
        form_adicionar = '<p class="muted-line" style="font-size:13px; margin-top:18px;">Apenas o administrador pode adicionar/editar/excluir alunos.</p>'

    content = f"""
        <div class="page-header">
            <h1>{turma["nome"]}</h1>
            <p class="subtitle">Ano letivo {turma["ano_letivo"]} · {len(alunos)} alunos</p>
            {excluir_turma_btn}
        </div>

        <h2>Alunos</h2>
        {alunos_html}

        {form_adicionar}
    """
    return render_page(f"Turma {turma['nome']}", content, active="turmas")


@app.post("/turmas/{turma_id}/alunos")
def adicionar_aluno(request: Request, 
    turma_id: int,
    nome: str = Form(...),
    numero: Optional[int] = Form(None),
    raca: str = Form(""),
    email: str = Form(""),
    data_nascimento: str = Form(""),
):
    _r = _require_admin_or_403(request)
    if _r is not None: return _r
    conn = get_db()
    codigo = gerar_codigo_aluno(conn)
    conn.execute(
        "INSERT INTO alunos (turma_id, nome, numero, codigo_unico, raca, email, data_nascimento) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (turma_id, nome.strip(), numero, codigo, raca.strip() or None, email.strip() or None, data_nascimento.strip() or None)
    )
    conn.commit()
    conn.close()
    return RedirectResponse(f"/turmas/{turma_id}", status_code=303)


# ==========================================
#  ROTAS DE APLICAÇÕES (ATUALIZADAS TAREFA A2)
# ==========================================


# ==========================================
#  ACESSO PENDENTE / BLOQUEADO
# ==========================================

@app.get("/acesso-pendente", response_class=HTMLResponse)
def acesso_pendente(request: Request):
    prof = get_current_professor(request)
    nome = prof["nome"] if prof else "Professor"
    body = f"""
        <div style="max-width:480px; margin:80px auto; text-align:center; padding:0 20px;">
            <div style="font-size:56px; margin-bottom:16px;">⏳</div>
            <h1 style="font-size:22px; margin-bottom:8px;">Acesso aguardando aprovação</h1>
            <p style="color:var(--text-muted); margin-bottom:24px;">
                Olá, <strong>{nome}</strong>! Seu cadastro foi recebido e está aguardando aprovação da gestão escolar.
            </p>
            <a href="/logout" class="btn" style="margin-top:24px;">Sair</a>
        </div>"""
    return HTMLResponse(render_page("Acesso pendente", body, active=""))


@app.get("/acesso-bloqueado", response_class=HTMLResponse)
def acesso_bloqueado(request: Request):
    prof = get_current_professor(request)
    nome = prof["nome"] if prof else "Professor"
    body = f"""
        <div style="max-width:480px; margin:80px auto; text-align:center; padding:0 20px;">
            <div style="font-size:56px; margin-bottom:16px;">🚫</div>
            <h1 style="font-size:22px; margin-bottom:8px;">Acesso bloqueado</h1>
            <p style="color:var(--text-muted);">Olá, <strong>{nome}</strong>. Seu acesso foi bloqueado. Entre em contato com o administrador.</p>
            <a href="/logout" class="btn" style="margin-top:24px;">Sair</a>
        </div>"""
    return HTMLResponse(render_page("Acesso bloqueado", body, active=""))


# ==========================================
#  PAINEL DE GESTÃO DE PROVAS
# ==========================================

STATUS_REVISAO_LABEL = {
    "rascunho":   ("✏️", "Rascunho",  "var(--text-muted)", "var(--bg-subtle)"),
    "submetida":  ("📤", "Submetida", "var(--orange)",     "var(--orange-bg)"),
    "aprovada":   ("✅", "Aprovada",  "var(--green)",      "var(--green-bg)"),
    "devolvida":  ("↩️", "Devolvida", "var(--red)",        "var(--red-bg)"),
}

def _status_badge_html(status: str) -> str:
    icon, label, color, bg = STATUS_REVISAO_LABEL.get(status, ("❓", status, "var(--text-muted)", "var(--bg-subtle)"))
    return f'<span style="background:{bg}; color:{color}; border-radius:6px; padding:2px 10px; font-size:12px; font-weight:600;">{icon} {label}</span>'


@app.post("/provas/{prova_id}/submeter")
def submeter_prova(prova_id: int):
    prof = _current_prof_ctx.get()
    if not prof:
        return RedirectResponse("/login", status_code=303)
    conn = get_db()
    prova = conn.execute("SELECT * FROM provas WHERE id = ?", (prova_id,)).fetchone()
    if prova and (prova["criada_por_professor_id"] == prof["id"] or prof.get("is_admin")):
        status_rev = prova["status_revisao"] if "status_revisao" in prova.keys() else "rascunho"
        if status_rev in ("rascunho", "devolvida"):
            conn.execute("UPDATE provas SET status_revisao = 'submetida', obs_gestao = NULL WHERE id = ?", (prova_id,))
            conn.commit()
    conn.close()
    return RedirectResponse(f"/provas/{prova_id}", status_code=303)


@app.post("/provas/{prova_id}/aprovar")
def aprovar_prova(prova_id: int, obs: str = Form("")):
    prof = _current_prof_ctx.get()
    if not prof or not (prof.get("is_admin") or prof.get("is_gestor")):
        return RedirectResponse("/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE provas SET status_revisao = 'aprovada', obs_gestao = ?, revisado_por_id = ?, revisado_em = CURRENT_TIMESTAMP WHERE id = ?",
        (obs.strip() or None, prof["id"], prova_id))
    conn.commit(); conn.close()
    return RedirectResponse("/painel-gestao", status_code=303)


@app.post("/provas/{prova_id}/devolver")
def devolver_prova(prova_id: int, obs: str = Form(...)):
    prof = _current_prof_ctx.get()
    if not prof or not (prof.get("is_admin") or prof.get("is_gestor")):
        return RedirectResponse("/login", status_code=303)
    conn = get_db()
    conn.execute("UPDATE provas SET status_revisao = 'devolvida', obs_gestao = ?, revisado_por_id = ?, revisado_em = CURRENT_TIMESTAMP WHERE id = ?",
        (obs.strip(), prof["id"], prova_id))
    conn.commit(); conn.close()
    return RedirectResponse("/painel-gestao", status_code=303)


@app.get("/painel-gestao", response_class=HTMLResponse)
def painel_gestao(request: Request, status: Optional[str] = "submetida", prof_id: Optional[int] = None):
    prof = _current_prof_ctx.get()
    if not prof or not (prof.get("is_admin") or prof.get("is_gestor")):
        return HTMLResponse(render_page("Acesso negado", '<div class="empty">Sem permissão.</div>', active="painel-gestao"), status_code=403)
    conn = get_db()
    where = []; params = []
    status_atual = status or "submetida"
    if status_atual != "todas":
        where.append("p.status_revisao = ?"); params.append(status_atual)
    if prof_id:
        where.append("p.criada_por_professor_id = ?"); params.append(prof_id)
    wc = ("WHERE " + " AND ".join(where)) if where else ""
    provas = conn.execute(f"""
        SELECT p.id, p.titulo, p.status_revisao, p.obs_gestao, p.criada_em, p.revisado_em,
               pr.nome AS criador_nome, rv.nome AS revisor_nome,
               (SELECT COUNT(*) FROM prova_questoes WHERE prova_id = p.id) AS n_questoes
        FROM provas p
        LEFT JOIN professores pr ON pr.id = p.criada_por_professor_id
        LEFT JOIN professores rv ON rv.id = p.revisado_por_id
        {wc} ORDER BY CASE p.status_revisao WHEN 'submetida' THEN 0 WHEN 'devolvida' THEN 1 WHEN 'aprovada' THEN 2 ELSE 3 END, p.id DESC
    """, params).fetchall()
    professores_lista = conn.execute("SELECT id, nome FROM professores ORDER BY nome").fetchall()
    contadores = {r["status_revisao"]: r["c"] for r in conn.execute("SELECT status_revisao, COUNT(*) AS c FROM provas GROUP BY status_revisao").fetchall()}
    conn.close()

    def _cnt(s): return contadores.get(s, 0)
    tabs_data = [
        ("submetida", "📤 Aguardando (" + str(_cnt("submetida")) + ")", "var(--orange)"),
        ("devolvida",  "↩️ Devolvidas (" + str(_cnt("devolvida")) + ")", "var(--red)"),
        ("aprovada",   "✅ Aprovadas (" + str(_cnt("aprovada")) + ")", "var(--green)"),
        ("rascunho",   "✏️ Rascunhos (" + str(_cnt("rascunho")) + ")", "var(--text-muted)"),
        ("todas",      "📋 Todas", "var(--accent)"),
    ]
    tabs_html = '<div style="display:flex; gap:0; border-bottom:2px solid var(--border); margin-bottom:18px; flex-wrap:wrap;">'
    for key, label, color in tabs_data:
        ativo = status_atual == key
        tabs_html += f'<a href="/painel-gestao?status={key}" style="padding:9px 16px; font-size:13px; font-weight:600; text-decoration:none; border-bottom:3px solid {"var(--accent)" if ativo else "transparent"}; color:{"var(--accent)" if ativo else "var(--text-muted)"}; margin-bottom:-2px; white-space:nowrap;">{label}</a>'
    tabs_html += '</div>'

    prof_opts = '<option value="">Todos os professores</option>' + "".join(f'<option value="{p["id"]}"{" selected" if prof_id == p["id"] else ""}>{p["nome"]}</option>' for p in professores_lista)
    filtros_html = f"""<form method="get" action="/painel-gestao" style="background:var(--bg-subtle); padding:10px 16px; border-radius:8px; margin-bottom:18px;">
        <input type="hidden" name="status" value="{status_atual}">
        <div style="display:flex; gap:10px; align-items:flex-end; flex-wrap:wrap;">
            <label style="margin:0;">Professor<select name="prof_id">{prof_opts}</select></label>
            <button type="submit" class="btn btn-primary" style="margin:0;">Filtrar</button>
            <a href="/painel-gestao?status={status_atual}" class="btn" style="margin:0;">Limpar</a>
        </div></form>"""

    cards_html = ""
    for p in provas:
        badge = _status_badge_html(p["status_revisao"])
        obs_html = f'<div style="margin-top:6px; background:var(--orange-bg); border-left:3px solid var(--orange); padding:6px 10px; border-radius:4px; font-size:12px;"><strong>Obs:</strong> {p["obs_gestao"]}</div>' if p["obs_gestao"] else ""
        revisor = f'<span style="font-size:11px; color:var(--text-muted);">Revisado por {p["revisor_nome"]} em {(p["revisado_em"] or "")[:10]}</span>' if p["revisor_nome"] else ""
        acoes = f'<a href="/provas/{p["id"]}" class="btn" style="padding:5px 10px; font-size:12px;">👁️ Ver</a><a href="/provas/{p["id"]}/imprimir" class="btn" style="padding:5px 10px; font-size:12px;" target="_blank">🖨️ PDF</a>'
        if p["status_revisao"] == "submetida":
            acoes += f'''<form method="post" action="/provas/{p["id"]}/aprovar" style="margin:0; display:inline-flex; gap:4px; align-items:center;">
                <input type="text" name="obs" placeholder="Obs. o